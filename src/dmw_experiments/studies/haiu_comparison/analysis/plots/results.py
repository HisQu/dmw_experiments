"""Plot DMW–Haiu comparison results from one or more exported workbooks."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
from dataclasses import dataclass
from datetime import datetime
from numbers import Real
from pathlib import Path
from typing import Any

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from matplotlib.patches import Rectangle
from matplotlib.text import Text
from openpyxl import load_workbook

import dmw_experiments.shared.analysis as ut
from dmw_experiments.studies.haiu_comparison.analysis.quality.grades import (
    CONDITION_LABELS as QUALITY_CONDITION_LABELS,
    CONDITION_ORDER as QUALITY_CONDITION_ORDER,
    build_quality_grade_analysis,
)
from dmw_experiments.studies.haiu_comparison.analysis.quality.errors import (
    build_quality_error_analysis,
)
from dmw_experiments.studies.haiu_comparison.analysis.quality.inputs import (
    load_historian_quality_error_counts,
    load_historian_quality_grades,
)
from dmw_experiments.studies.haiu_comparison.analysis.workbooks.quality import (
    export_quality_grade_analysis_workbook,
)

CONDITION_ORDER = QUALITY_CONDITION_ORDER
CONDITION_LABELS = QUALITY_CONDITION_LABELS
PAIR_SHEETS = {
    "02_DMW_Context_AB": "DMW: full vs RAG",
    "03_DMW_vs_Haiu_RAG": "DMW RAG vs standalone",
}
QUALITY_PAIR_SHEETS = {
    ("workflow_full_ontology", "workflow_rag"): "DMW: full vs RAG",
    ("workflow_rag", "haiu_rag_ontologizer"): "DMW RAG vs standalone",
}
ABSOLUTE_PAIR_SIDE_LABELS = {
    ("DMW: full vs RAG", "left"): "DMW + Full Ontology",
    ("DMW: full vs RAG", "right"): "DMW + Haiu RAG",
    ("DMW RAG vs standalone", "left"): "DMW + Haiu RAG",
    ("DMW RAG vs standalone", "right"): "Standalone Haiu RAG",
}
PAIRED_TRAJECTORY_CONDITION_ORDER = (
    "DMW + Full Ontology",
    "DMW + Haiu RAG",
    "Standalone Haiu RAG",
)
PAIRED_COMPARISON_ENDPOINTS = {
    ("DMW: full vs RAG", "left"): (0.0, "dmw_full"),
    ("DMW: full vs RAG", "right"): (1.0, "dmw_rag_context"),
    ("DMW RAG vs standalone", "left"): (2.0, "dmw_rag_system"),
    ("DMW RAG vs standalone", "right"): (3.0, "standalone"),
}
PAIRED_COMPARISON_ENDPOINT_ORDER = tuple(
    endpoint_label for _, endpoint_label in PAIRED_COMPARISON_ENDPOINTS.values()
)
PAIRED_COMPARISON_TICKS: tuple[tuple[float, str], ...] = (
    (0.5, "DMW vs DMW+HAIU"),
    (2.5, "DMW+HAIU vs HAIU"),
)
OUTCOME_CONDITION_TICK_LABELS = ("DMW", "DMW+HAIU", "HAIU")
OUTCOME_COMPARISON_TICK_LABELS = (
    "DMW vs\nDMW+HAIU",
    "DMW+HAIU\nvs HAIU",
)
REQUIRED_SHEETS = {
    "00_About",
    "01_Results",
    "02_DMW_Context_AB",
    "03_DMW_vs_Haiu_RAG",
    "04_Observations",
    "09_Token_Accounting",
}
PROVIDER_COLORS = ("#0072B2", "#D55E00", "#009E73", "#CC79A7")
ERROR_CLASS_COLORS = {
    "Invalid Turtle": "#B07AA1",
    "Context / length": "#59A14F",
    "Provider timeout": "#EDC948",
    "Connection error": "#4E79A7",
    "Provenance rejection": "#9C755F",
    "Runner deadline": "#BAB0AC",
    "Other generation error": "#E15759",
}
FIGURE_FONTSIZE = 8
OUTCOME_FIGURE_SIZE = (8.6, 2.9)
# > Two 2 in-tall panels plus dedicated legend bands in a narrower export.
PAIRED_GRADE_FIGURE_SIZE = (6.0, 4.7)
PAIRED_RESOURCE_FIGURE_SIZE = (6.0, 3.6)
SIX_PANEL_FIGURE_SIZE = (7.7, 5.2)
# > Retains the 30%-narrowed width while accommodating two interaction rows.
PROVIDER_INTERACTION_FIGURE_SIZE = (5.04, 5.4)
# > Four side-by-side panels retain 1.7 in of drawable height below two
# > vertical legend groups.
ERROR_PROFILE_FIGURE_SIZE = (8.8, 4.2)
BAR_WIDTH = 0.52
FALSE_ASSIGNMENT_BAR_WIDTH = 0.8
PAIR_TRAJECTORY_JITTER = 0.08
PAIR_TRAJECTORY_LINE_ALPHA = 0.3
GRADE_COLORS = {
    1: "#1A9850",
    2: "#66BD63",
    3: "#FEE08B",
    4: "#FDAE61",
    5: "#F46D43",
    6: "#D73027",
}
GRADE_LEGEND_LABELS = {
    1: "Grade 1 — correct, essentially complete",
    2: "Grade 2 — correct, minor omissions",
    3: "Grade 3 — incomplete, factually safe",
    4: "Grade 4 — ≥1 error; local/formal, patchable",
    5: "Grade 5 — ≥1 error; plausible historical error",
    6: "Grade 6 — ≥1 error; gross source-reading failure",
}
ERROR_INTERPRETATION_BAND_COLORS = {
    "0": "#009E73",
    "1": "#F0E442",
    "2": "#E69F00",
    "3+": "#D55E00",
}
ERROR_ASSERTION_BAND_COLORS = {
    "0": "#009E73",
    "1": "#F0E442",
    "2": "#E69F00",
    "3": "#D55E00",
    "4+": "#A50F15",
}
ERROR_COUNT_SEGMENT_ANNOTATION_MINIMUM_SHARE = 0.075
PAIR_CHANGE_MAGNITUDE_COLORS = {
    "improved_by_more_than_2": "#006D2C",
    "improved_by_2": "#31A354",
    "improved_by_1": "#A1D99B",
    "unchanged": "#7F7F7F",
    "worsened_by_1": "#FCAE91",
    "worsened_by_2": "#FB6A4A",
    "worsened_by_more_than_2": "#CB181D",
}
ERROR_CHANGE_MAGNITUDE_LABELS = {
    "improved_by_more_than_2": ">2 fewer errors",
    "improved_by_2": "2 fewer errors",
    "improved_by_1": "1 fewer error",
    "unchanged": "Unchanged",
    "worsened_by_1": "1 more error",
    "worsened_by_2": "2 more errors",
    "worsened_by_more_than_2": ">2 more errors",
}
GRADE_CHANGE_MAGNITUDE_LABELS = {
    "improved_by_more_than_2": "Improved by >2 grades",
    "improved_by_2": "Improved by 2 grades",
    "improved_by_1": "Improved by 1 grade",
    "unchanged": "Unchanged",
    "worsened_by_1": "Worsened by 1 grade",
    "worsened_by_2": "Worsened by 2 grades",
    "worsened_by_more_than_2": "Worsened by >2 grades",
}
DIRECT_PAIR_ENDPOINTS = {
    ("DMW full ontology vs DMW RAG", "first"): (0.0, "dmw_full"),
    ("DMW full ontology vs DMW RAG", "second"): (1.0, "dmw_rag_context"),
    ("DMW RAG vs standalone Haiu RAG", "first"): (2.0, "dmw_rag_system"),
    ("DMW RAG vs standalone Haiu RAG", "second"): (3.0, "standalone"),
}
DIRECT_PAIR_ENDPOINT_LABELS = {
    "dmw_full": "DMW",
    "dmw_rag_context": "DMW+\nHAIU",
    "dmw_rag_system": "DMW+\nHAIU",
    "standalone": "HAIU",
}
FIGURE_CAPTIONS = {
    "outcomes": (
        "Completion outcomes by provider and condition. Valid Turtle shares "
        "and issue categories use all attempted regesta; paired-comparison "
        "shares use only the planned condition pairs."
    ),
    "paired-absolute-metrics": (
        "Within-regest changes in generated ontology metrics. Every trajectory "
        "connects the same provider/regest across a planned direct comparison; "
        "sample sizes therefore vary by metric according to valid outputs."
    ),
    "paired-quality-grades": (
        "Historian quality grades. Panel A gives ordinal grade distributions "
        "only for matched provider–regest pairs. The two DMW + Haiu RAG bars "
        "are the same condition, each restricted to the valid pairs of its "
        "adjacent comparison. Panel B separates right-condition improvements "
        "and worsenings of exactly 1 grade, exactly 2 grades, and more than "
        "2 grades around unchanged pairs. Lower grades are better."
    ),
    "quality-grade-provider-interaction": (
        "Provider interaction in regesta complete for all three conditions in "
        "both providers. Thin lines are paired regesta; thick transparent "
        "lines are provider trends with paired-bootstrap grade or Wilson "
        "false-assignment-rate intervals. The reported sign and McNemar tests "
        "are exploratory."
    ),
    "false-assignment-error-profile": (
        "False-assignment counts in matched provider–regest pairs, pooled only "
        "after matching within provider. Panels A and B show the distributions "
        "of independent false interpretations and false atomic assertions; "
        "Panels C and D separate right-condition changes of exactly 1 error, "
        "exactly 2 errors, and more than 2 errors around unchanged pairs. "
        "Panel C treats the recorded 3+ interpretation band as its lower bound "
        "of 3, so its magnitude bins are conservative when either endpoint is "
        "3+ and an unchanged 3+ pair need not have equal exact counts."
    ),
}


@dataclass(frozen=True)
class WorkbookResults:
    """Hold the plot-ready tables and identity from one workbook.

    :param path: Source workbook path.
    :param provider: Short provider-profile identifier.
    :param provider_label: Human-readable provider and quantization label.
    :param status: Export status shown in ``00_About``.
    :param results: Per-condition aggregate results.
    :param observations: Per-regest condition observations.
    :param pairs: Valid and invalid rows from both primary comparisons.
    """

    path: Path
    provider: str
    provider_label: str
    status: str
    results: pd.DataFrame
    observations: pd.DataFrame
    pairs: pd.DataFrame


def load_workbook_results(path: Path) -> WorkbookResults:
    """Read and validate the workbook tables used by the figures.

    :param path: Exported ``overview.xlsx`` file.
    :return: Validated workbook identity and plot-ready tables.
    """
    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"Workbook not found: {resolved}")

    excel_file = pd.ExcelFile(resolved)
    missing = sorted(REQUIRED_SHEETS.difference(excel_file.sheet_names))
    if missing:
        raise ValueError(
            f"{resolved} is missing required worksheets: {', '.join(missing)}"
        )

    about = _read_about(resolved)
    profile = _parse_provider_profile(about, resolved)
    provider = str(profile["name"])
    provider_label = _provider_label(profile)

    results = pd.read_excel(excel_file, sheet_name="01_Results")
    observations = pd.read_excel(excel_file, sheet_name="04_Observations")
    excluded_outcome_error_cells = _local_runtime_recovery_cells(resolved)
    observations["_outcome_error_excluded"] = observations.apply(
        lambda row: (
            (
                str(row["condition"]),
                _regest_id_text(row["regest_id"]),
            )
            in excluded_outcome_error_cells
        ),
        axis="columns",
    )
    pair_frames: list[pd.DataFrame] = []
    for sheet_name, comparison_label in PAIR_SHEETS.items():
        frame = pd.read_excel(excel_file, sheet_name=sheet_name)
        frame["comparison"] = comparison_label
        pair_frames.append(frame)
    pairs = pd.concat(pair_frames, ignore_index=True)

    _require_columns(
        results,
        {
            "condition",
            "observations",
            "valid_completed_count",
            "valid_completed_rate",
        },
        path=resolved,
        sheet_name="01_Results",
    )
    _require_columns(
        observations,
        {
            "condition",
            "regest_id",
            "success",
            "turtle_syntax_valid",
            "output_truncated",
            "failure_code",
            "error_message",
            "duration_seconds",
            "prompt_tokens",
            "prompt_tokens_complete",
            "output_tokens",
        },
        path=resolved,
        sheet_name="04_Observations",
    )
    _require_columns(
        pairs,
        {
            "valid_pair",
            "left_condition",
            "right_condition",
            "left_duration_seconds",
            "right_duration_seconds",
            "left_reuse_share",
            "right_reuse_share",
            "duration_delta_seconds",
            "reuse_share_delta",
            "left_novel_schema_declaration_count",
            "right_novel_schema_declaration_count",
            "left_triples",
            "right_triples",
        },
        path=resolved,
        sheet_name="02/03 paired results",
    )

    for frame in (results, observations, pairs):
        frame["provider"] = provider
        frame["provider_label"] = provider_label
    results["condition_label"] = ut.series_column(results, "condition").map(
        lambda value: CONDITION_LABELS.get(str(value))
    )
    observations["condition_label"] = ut.series_column(
        observations, "condition"
    ).map(lambda value: CONDITION_LABELS.get(str(value)))
    result_condition_labels = ut.series_column(results, "condition_label")
    observation_condition_labels = ut.series_column(
        observations, "condition_label"
    )
    unknown_conditions = sorted(
        set(
            ut.series_column(
                results.loc[result_condition_labels.isna()],
                "condition",
            )
        )
        | set(
            ut.series_column(
                observations.loc[observation_condition_labels.isna()],
                "condition",
            )
        )
    )
    if unknown_conditions:
        raise ValueError(
            f"{resolved} contains unsupported conditions: "
            f"{', '.join(str(value) for value in unknown_conditions)}"
        )

    return WorkbookResults(
        path=resolved,
        provider=provider,
        provider_label=provider_label,
        status=about.get("Status", "UNKNOWN STATUS"),
        results=results,
        observations=observations,
        pairs=pairs,
    )


def _local_runtime_recovery_cells(workbook_path: Path) -> set[tuple[str, str]]:
    """Find cells archived by local-runtime recovery amendments.

    :param workbook_path: Derived overview workbook beneath a run directory.
    :return: Condition and regest keys selected by a local-runtime amendment.
    """
    run_dir = workbook_path.parent.parent
    amendments_dir = run_dir / "summaries" / "amendments"
    if not amendments_dir.is_dir():
        return set()

    selected_cells: set[tuple[str, str]] = set()
    for amendment_path in sorted(amendments_dir.glob("*.json")):
        amendment = json.loads(amendment_path.read_text(encoding="utf-8"))
        if amendment.get("kind") != "local_runtime_recovery":
            continue
        amendment_id = amendment.get("amendment_id")
        if not isinstance(amendment_id, str) or not amendment_id:
            raise ValueError(
                f"{amendment_path} has no local-runtime amendment ID."
            )
        archive_root = run_dir / "superseded" / amendment_id / "raw"
        for artifact_path in archive_root.glob("*/*.json"):
            selected_cells.add((artifact_path.parent.name, artifact_path.stem))
    return selected_cells


def _regest_id_text(value: object) -> str:
    """Normalize numeric spreadsheet identifiers to their raw-file spelling.

    :param value: Regest identifier loaded from an observation worksheet.
    :return: Canonical text form used in raw artifact paths.
    """
    if isinstance(value, Real) and not isinstance(value, bool):
        numeric_value = float(value)
        if not np.isnan(numeric_value) and numeric_value.is_integer():
            return str(int(numeric_value))
    return str(value)


def plot_workbooks(
    workbook_paths: list[Path],
    *,
    output_root: Path | None = None,
    timestamp: str | None = None,
    quality_review_workbook: Path | None = None,
    quality_reveal_key: Path | None = None,
) -> Path:
    """Create timestamped figures from matched provider workbooks.

    :param workbook_paths: Exported ``overview.xlsx`` files to compare.
    :param output_root: Parent directory for the timestamped plot directory.
    :param timestamp: Optional deterministic timestamp for tests or reruns.
    :param quality_review_workbook: Optional graded historian-review workbook.
    :param quality_reveal_key: Condition mapping for the masked review IDs.
    :return: Directory containing PNG, PDF, and manifest outputs.
    """
    if not workbook_paths:
        raise ValueError("At least one workbook is required.")
    if (quality_review_workbook is None) != (quality_reveal_key is None):
        raise ValueError(
            "Historian quality plotting requires both the graded workbook "
            "and its reveal key."
        )

    workbooks = [load_workbook_results(path) for path in workbook_paths]
    providers = [workbook.provider for workbook in workbooks]
    duplicates = sorted(
        provider for provider in set(providers) if providers.count(provider) > 1
    )
    if duplicates:
        raise ValueError(
            "Each provider profile may appear only once; duplicates: "
            f"{', '.join(duplicates)}"
        )

    generated_at = datetime.now().astimezone()
    directory_timestamp = timestamp or generated_at.strftime("%Y%m%dT%H%M%S%Z")
    root = (
        Path(output_root).expanduser().resolve()
        if output_root is not None
        else _common_output_root([workbook.path for workbook in workbooks])
    )
    output_dir = root / f"plots-{directory_timestamp}"
    if output_dir.exists():
        raise FileExistsError(
            f"Plot output directory already exists: {output_dir}"
        )
    output_dir.mkdir(parents=True)

    ut.configure_matplotlib_defaults(fontsize=FIGURE_FONTSIZE)
    palette = {
        workbook.provider_label: PROVIDER_COLORS[index]
        for index, workbook in enumerate(workbooks)
    }
    results = pd.concat(
        [workbook.results for workbook in workbooks], ignore_index=True
    )
    observations = pd.concat(
        [workbook.observations for workbook in workbooks], ignore_index=True
    )
    pairs = pd.concat(
        [workbook.pairs for workbook in workbooks], ignore_index=True
    )
    provider_order = [workbook.provider_label for workbook in workbooks]
    status_text = _status_text(workbooks)

    figures = {
        "outcomes": _plot_outcomes(
            results,
            pairs,
            observations=observations,
            provider_order=provider_order,
            palette=palette,
            status_text=status_text,
        ),
        "paired-absolute-metrics": _plot_paired_absolute_metrics(
            pairs,
            observations=observations,
            provider_order=provider_order,
            palette=palette,
            status_text=status_text,
        ),
    }
    quality_grade_manifest: dict[str, Any] | None = None
    if quality_review_workbook is not None and quality_reveal_key is not None:
        from dmw_experiments.studies.haiu_comparison.analysis.plots.quality import (
            _has_error_profile_pairs,
            _plot_false_assignment_error_profile,
            _plot_quality_grade_overview,
            _plot_quality_grade_provider_interaction,
        )

        grades = load_historian_quality_grades(
            workbook_path=quality_review_workbook,
            reveal_key_path=quality_reveal_key,
            provider_order=provider_order,
        )
        quality_analysis = build_quality_grade_analysis(grades)
        error_counts = load_historian_quality_error_counts(
            workbook_path=quality_review_workbook,
            reveal_key_path=quality_reveal_key,
            provider_order=provider_order,
        )
        error_analysis = build_quality_error_analysis(error_counts)
        quality_analysis_workbook = export_quality_grade_analysis_workbook(
            analysis=quality_analysis,
            path=output_dir
            / f"quality-grade-analysis-{directory_timestamp}.xlsx",
            quality_review_workbook=quality_review_workbook,
            quality_reveal_key=quality_reveal_key,
            error_analysis=(
                error_analysis
                if not error_analysis.observations.empty
                else None
            ),
        )
        figures["paired-quality-grades"] = _plot_quality_grade_overview(
            quality_analysis,
            status_text=status_text,
        )
        provider_interaction = _plot_quality_grade_provider_interaction(
            quality_analysis,
            provider_order=provider_order,
            palette=palette,
            status_text=status_text,
        )
        if provider_interaction is not None:
            figures["quality-grade-provider-interaction"] = provider_interaction
        if _has_error_profile_pairs(error_analysis):
            figures["false-assignment-error-profile"] = (
                _plot_false_assignment_error_profile(error_analysis)
            )
        quality_grade_manifest = {
            "workbook": _source_display_path(quality_review_workbook),
            "reveal_key": _source_display_path(quality_reveal_key),
            "analysis_workbook": quality_analysis_workbook.name,
            "graded_observations": len(quality_analysis.observations),
            "complete_triplets": len(quality_analysis.complete_triplets),
            "providers": sorted(
                set(
                    ut.series_column(
                        quality_analysis.observations,
                        "provider_label",
                    ).astype(str)
                )
            ),
            "error_count_observations": len(error_analysis.observations),
            "interpretation_count_observations": int(
                error_analysis.observations["false_interpretations"]
                .notna()
                .sum()
            ),
            "assertion_count_observations": int(
                error_analysis.observations["false_assertions"].notna().sum()
            ),
            "matched_interpretation_pairs": len(
                error_analysis.matched_interpretation_pairs
            ),
            "matched_assertion_pairs": len(
                error_analysis.matched_assertion_pairs
            ),
        }
    try:
        for stem, figure in figures.items():
            ut.export_figure(
                figure,
                output_dir / stem,
                raster_dpi=200,
            )
    finally:
        for figure in figures.values():
            plt.close(figure)
    caption_path = output_dir / "image-captions.md"
    _write_image_captions(caption_path, figure_stems=tuple(figures))

    manifest = {
        "schema_version": 1,
        "generated_at": generated_at.isoformat(),
        "status": status_text,
        "inputs": [
            {
                "source": _source_display_path(workbook.path),
                "provider_profile": workbook.provider,
                "provider_label": workbook.provider_label,
                "export_status": workbook.status,
                "observations": len(workbook.observations),
            }
            for workbook in workbooks
        ],
        "figures": [
            f"{stem}.{suffix}"
            for stem in figures
            for suffix in ut.DEFAULT_FIGURE_FORMATS
        ],
        "caption_file": caption_path.name,
        "paired_metric_rule": "Only rows with valid_pair=true are plotted.",
    }
    if quality_grade_manifest is not None:
        manifest["historian_quality_grades"] = quality_grade_manifest
    (output_dir / "plot_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return output_dir


def _read_about(path: Path) -> dict[str, str]:
    """Read the two-column identity table from ``00_About``.

    :param path: Workbook path.
    :return: Non-empty key/value rows.
    """
    worksheet = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )["00_About"]
    return {
        str(key): str(value)
        for key, value in worksheet.iter_rows(values_only=True)
        if key is not None and value is not None
    }


def _parse_provider_profile(
    about: dict[str, str],
    path: Path,
) -> dict[str, Any]:
    """Parse the provider-profile JSON stored in the workbook.

    :param about: Key/value rows from ``00_About``.
    :param path: Workbook path used in validation messages.
    :return: Provider-profile mapping.
    """
    raw_profile = about.get("Provider profile")
    if raw_profile is None:
        raise ValueError(f"{path} has no Provider profile in 00_About.")
    try:
        profile = json.loads(raw_profile)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"{path} contains invalid Provider profile JSON."
        ) from exc
    if not isinstance(profile, dict) or not profile.get("name"):
        raise ValueError(f"{path} has no provider-profile name.")
    return profile


def _provider_label(profile: dict[str, Any]) -> str:
    """Build a concise label without hard-coding provider profiles.

    :param profile: Parsed provider-profile metadata.
    :return: Display label containing provider and quantization when present.
    """
    provider = str(profile.get("chat_provider") or profile["name"])
    provider_names = {
        "academiccloud": "AcademicCloud",
        "lmstudio": "LM Studio",
    }
    label = provider_names.get(provider.lower(), provider)
    quantization = profile.get("quantization")
    return f"{label} {quantization}" if quantization else label


def _require_columns(
    frame: pd.DataFrame,
    required: set[str],
    *,
    path: Path,
    sheet_name: str,
) -> None:
    """Reject workbook schema drift before creating partial figures.

    :param frame: Loaded worksheet data.
    :param required: Columns needed by the plotting code.
    :param path: Workbook path used in the error.
    :param sheet_name: Reader-facing worksheet identifier.
    :return: None.
    """
    missing = sorted(required.difference(frame.columns))
    if missing:
        raise ValueError(
            f"{path} worksheet {sheet_name} is missing columns: "
            f"{', '.join(missing)}"
        )


def _common_output_root(paths: list[Path]) -> Path:
    """Choose the nearest shared directory for plot output.

    :param paths: Resolved workbook paths.
    :return: Shared ancestor of the workbook-containing directories.
    """
    common = Path(os.path.commonpath([str(path.parent) for path in paths]))
    return common


def _status_text(workbooks: list[WorkbookResults]) -> str:
    """Summarize whether any source is a partial diagnostic export.

    :param workbooks: Loaded workbook results.
    :return: Figure subtitle describing evidentiary status.
    """
    statuses = {workbook.status for workbook in workbooks}
    if any("PARTIAL" in status.upper() for status in statuses):
        return "PARTIAL DIAGNOSTIC EXPORT — not publication evidence"
    if len(statuses) == 1:
        return next(iter(statuses))
    return " / ".join(sorted(statuses))


def _plot_outcomes(
    results: pd.DataFrame,
    pairs: pd.DataFrame,
    *,
    observations: pd.DataFrame,
    provider_order: list[str],
    palette: dict[str, str],
    status_text: str,
) -> Figure:
    """Plot valid ontology rates and valid paired-comparison denominators.

    :param results: Per-condition aggregate results from every provider.
    :param pairs: Rows from both primary paired comparisons.
    :param observations: Canonical per-regest condition records.
    :param provider_order: Provider labels in the requested display order.
    :param palette: Provider-label color mapping.
    :param status_text: Evidence-status subtitle.
    :return: Completed outcome figure.
    """
    figure, axes = plt.subplots(
        1,
        2 + len(provider_order),
        figsize=OUTCOME_FIGURE_SIZE,
        gridspec_kw={
            "width_ratios": (1.1, 1.0, *([1.0] * len(provider_order))),
        },
    )
    condition_order = [CONDITION_LABELS[value] for value in CONDITION_ORDER]

    sns.barplot(
        data=results,
        x="condition_label",
        y="valid_completed_rate",
        hue="provider_label",
        order=condition_order,
        hue_order=provider_order,
        palette=palette,
        errorbar=None,
        width=BAR_WIDTH,
        ax=axes[0],
    )
    axes[0].set_title("Valid completed ontologies")
    axes[0].set_xlabel("")
    axes[0].set_ylabel("Share of attempted regests")
    ut.configure_percent_axis(axes[0], ymin=0, ymax=1.0, major_max=1.0)
    ut.apply_grid(axes[0])
    axes[0].set_xticks(
        range(len(condition_order)),
        OUTCOME_CONDITION_TICK_LABELS,
    )
    success_texts = _annotate_success_bars(
        axes[0],
        results,
        provider_order,
        condition_order,
    )

    pair_summary = (
        pairs.groupby(["comparison", "provider_label"], observed=True)[
            "valid_pair"
        ]
        .agg(valid_pairs="sum", attempted_pairs="size")
        .reset_index()
    )
    pair_summary["valid_pair_rate"] = (
        pair_summary["valid_pairs"] / pair_summary["attempted_pairs"]
    )
    comparison_order = list(PAIR_SHEETS.values())
    sns.barplot(
        data=pair_summary,
        x="comparison",
        y="valid_pair_rate",
        hue="provider_label",
        order=comparison_order,
        hue_order=provider_order,
        palette=palette,
        errorbar=None,
        width=BAR_WIDTH,
        ax=axes[1],
    )
    axes[1].set_title("Valid paired comparisons")
    axes[1].set_xlabel("")
    axes[1].set_ylabel("Share of attempted pairs")
    ut.configure_percent_axis(axes[1], ymin=0, ymax=1.0, major_max=1.0)
    ut.apply_grid(axes[1])
    axes[1].set_xticks(
        range(len(comparison_order)),
        OUTCOME_COMPARISON_TICK_LABELS,
    )
    pair_texts = _annotate_pair_bars(
        axes[1],
        pair_summary,
        provider_order,
        comparison_order,
    )

    issue_assessments = [
        _plot_error_assessment(
            ax,
            observations=observations,
            provider_label=provider,
        )
        for ax, provider in zip(axes[2:], provider_order, strict=True)
    ]
    present_error_classes = set().union(
        *(
            error_classes
            for error_classes, _texts, _excluded_count in issue_assessments
        )
    )

    _finish_figure(
        figure,
        axes,
        status_text=status_text,
        provider_count=len(provider_order),
        layout_top=0.72,
        subplot_wspace=0.52,
        show_legend=False,
    )
    ut.ensure_top_text_headroom(axes[0], success_texts)
    ut.ensure_top_text_headroom(axes[1], pair_texts)
    for ax, (_error_classes, issue_texts, _excluded_count) in zip(
        axes[2:],
        issue_assessments,
        strict=True,
    ):
        ut.ensure_top_text_headroom(ax, issue_texts)
    provider_handles, provider_labels = ut.square_legend_entries(palette)
    provider_legend_ax = figure.add_axes((0.03, 0.82, 0.45, 0.06))
    ut.add_top_band_figure_legend(
        figure,
        provider_legend_ax,
        handles=provider_handles,
        labels=provider_labels,
        ncol=len(provider_labels),
        frameon=False,
        show_handles=True,
    )
    error_class_colors = {
        error_class: color
        for error_class, color in ERROR_CLASS_COLORS.items()
        if error_class in present_error_classes
    }
    if error_class_colors:
        error_handles, error_labels = ut.square_legend_entries(
            error_class_colors
        )
        error_legend_ax = figure.add_axes((0.53, 0.77, 0.44, 0.13))
        ut.add_top_band_figure_legend(
            figure,
            error_legend_ax,
            handles=error_handles,
            labels=error_labels,
            max_rows=2,
            frameon=False,
            show_handles=True,
        )
    return figure


def _plot_error_assessment(
    ax: Axes,
    *,
    observations: pd.DataFrame,
    provider_label: str,
) -> tuple[set[str], list[Text], int]:
    """Show every outcome that does not provide valid Turtle by condition.

    :param ax: Panel that receives one provider's stacked failure counts.
    :param observations: Canonical per-regest condition records.
    :param provider_label: Display identity of the provider represented here.
    :return: Error categories displayed in this provider panel, their count
        labels, and the number of documented configuration-invalid errors
        excluded from this outcome-error assessment.
    """
    provider_rows = observations.loc[
        observations["provider_label"] == provider_label
    ].copy()
    provider_rows = provider_rows.loc[
        ~(
            provider_rows["success"].astype(bool)
            & provider_rows["turtle_syntax_valid"].eq(True)
        )
    ]
    configuration_invalid = provider_rows.apply(
        _is_configuration_invalid_outcome_error,
        axis="columns",
    )
    excluded_count = int(configuration_invalid.sum())
    provider_rows = provider_rows.loc[~configuration_invalid]
    provider_rows["error_class"] = provider_rows.apply(
        _classify_failure,
        axis="columns",
    )
    bottom = pd.Series(0, index=CONDITION_ORDER, dtype="int64")
    for error_class, color in ERROR_CLASS_COLORS.items():
        counts = (
            provider_rows.loc[
                provider_rows["error_class"] == error_class,
                "condition",
            ]
            .value_counts()
            .reindex(CONDITION_ORDER, fill_value=0)
        )
        ax.bar(
            range(len(CONDITION_ORDER)),
            counts,
            bottom=bottom,
            color=color,
            width=BAR_WIDTH,
            label=error_class,
        )
        bottom += counts
    max_issue_count = max(1, int(bottom.max()))
    ax.set_title(f"Outcome issues\n{provider_label}", pad=10)
    ax.set_xlabel("")
    ax.set_xticks(range(len(CONDITION_ORDER)), OUTCOME_CONDITION_TICK_LABELS)
    ax.set_ylabel("Regests without valid Turtle")
    ax.set_ylim(0, max_issue_count)
    ut.apply_grid(ax)
    count_texts: list[Text] = []
    for x_position, count in enumerate(bottom):
        count_texts.append(
            ax.annotate(
                f"n={int(count)}",
                (x_position, int(count)),
                xytext=(0, 2),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=6.5,
                alpha=0.7,
            )
        )
    return set(provider_rows["error_class"]), count_texts, excluded_count


def _classify_failure(row: pd.Series) -> str:
    """Map one non-valid-Turtle outcome to a stable error category.

    :param row: Observation fields from the exported workbook.
    :return: Reader-facing error category for the stacked assessment panel.
    """
    turtle_syntax_valid = row.get("turtle_syntax_valid")
    if (
        bool(row.get("success"))
        and isinstance(turtle_syntax_valid, (bool, np.bool_, Real))
        and not bool(turtle_syntax_valid)
    ):
        return "Invalid Turtle"

    message = str(row.get("error_message") or "").lower()
    failure_code = str(row.get("failure_code") or "").lower()
    if (
        bool(row.get("output_truncated"))
        or "length limit" in message
        or failure_code == "model_context_window_exceeded"
    ):
        return "Context / length"
    if "timeout" in message or "timed out" in message:
        return "Provider timeout"
    if "connection" in message:
        return "Connection error"
    if (
        "fresh ontology observation" in message
        or "reused a prior ontology" in message
    ):
        return "Provenance rejection"
    if "condition_wall_clock_timeout" in failure_code:
        return "Runner deadline"
    return "Other generation error"


def _is_configuration_invalid_outcome_error(row: pd.Series) -> bool:
    """Skip abandoned LM Studio configuration attempts in outcome errors.

    :param row: One exported condition observation.
    :return: Whether the row is an invalid local configuration attempt.
    """
    provider = str(row.get("provider_label") or "").lower()
    if not provider.startswith("lm studio"):
        return False
    if bool(row.get("success")):
        return False
    if bool(row.get("_outcome_error_excluded")):
        return _classify_failure(row) == "Other generation error"
    recovery = row.get("local_runtime_recovery")
    if recovery is None or (isinstance(recovery, float) and np.isnan(recovery)):
        return False
    if isinstance(recovery, str) and recovery.strip().lower() in {
        "",
        "{}",
        "none",
        "nan",
    }:
        return False
    return _classify_failure(row) == "Other generation error"


def _annotate_success_bars(
    ax: Axes,
    results: pd.DataFrame,
    provider_order: list[str],
    condition_order: list[str],
) -> list[Text]:
    """Label success-rate bars with valid and attempted counts.

    :param ax: Outcome-rate axes.
    :param results: Per-condition aggregate results.
    :param provider_order: Provider labels in display order.
    :param condition_order: Condition labels in display order.
    :return: Created count labels for final dynamic boundary clearance.
    """
    rows = _bar_rows(
        results,
        category_column="condition_label",
        category_order=condition_order,
        provider_order=provider_order,
    )
    patches = [patch for patch in ax.patches if isinstance(patch, Rectangle)]
    count_texts: list[Text] = []
    for patch, row in zip(patches, rows, strict=False):
        if not patch.get_height():
            continue
        label = (
            f"{int(row['valid_completed_count'])}/{int(row['observations'])}"
        )
        count_texts.append(
            ax.annotate(
                label,
                (patch.get_x() + patch.get_width() / 2, patch.get_height()),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=6.5,
                alpha=0.7,
            )
        )
    return count_texts


def _annotate_pair_bars(
    ax: Axes,
    summary: pd.DataFrame,
    provider_order: list[str],
    comparison_order: list[str],
) -> list[Text]:
    """Label paired-result bars with valid and attempted counts.

    :param ax: Valid-pair-rate axes.
    :param summary: Aggregated valid and attempted pair counts.
    :param provider_order: Provider labels in display order.
    :param comparison_order: Primary comparisons in display order.
    :return: Created count labels for final dynamic boundary clearance.
    """
    rows = _bar_rows(
        summary,
        category_column="comparison",
        category_order=comparison_order,
        provider_order=provider_order,
    )
    patches = [patch for patch in ax.patches if isinstance(patch, Rectangle)]
    count_texts: list[Text] = []
    for patch, row in zip(patches, rows, strict=False):
        label = f"{int(row['valid_pairs'])}/{int(row['attempted_pairs'])}"
        count_texts.append(
            ax.annotate(
                label,
                (patch.get_x() + patch.get_width() / 2, patch.get_height()),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=6.5,
                alpha=0.7,
            )
        )
    return count_texts


def _bar_rows(
    frame: pd.DataFrame,
    *,
    category_column: str,
    category_order: list[str],
    provider_order: list[str],
) -> list[pd.Series]:
    """Order table rows in the same hue-major order as seaborn bars.

    :param frame: Aggregate values represented by bars.
    :param category_column: Column mapped to the x-axis.
    :param category_order: X-axis categories in display order.
    :param provider_order: Provider labels in hue order.
    :return: Rows aligned with Matplotlib's bar-patch order.
    """
    indexed = frame.set_index([category_column, "provider_label"])
    rows: list[pd.Series] = []
    for provider in provider_order:
        for category in category_order:
            rows.append(indexed.loc[(category, provider)])
    return rows


def _plot_resource_usage(
    observations: pd.DataFrame,
    *,
    provider_order: list[str],
    palette: dict[str, str],
    status_text: str,
) -> Figure:
    """Plot paired complete prompt and output token observations.

    :param observations: Per-regest condition observations.
    :param provider_order: Provider labels in display order.
    :param palette: Provider-label color mapping.
    :param status_text: Evidence-status subtitle.
    :return: Completed resource-observation figure.
    """
    figure, axes = plt.subplots(1, 2, figsize=PAIRED_RESOURCE_FIGURE_SIZE)
    plot_specs = (
        (
            "prompt_tokens",
            "Total input tokens for Turtle generation",
            "Tokens",
            "log",
        ),
        ("output_tokens", "Output tokens", "Tokens", "symlog"),
    )
    for ax, (column, title, y_label, scale) in zip(
        axes,
        plot_specs,
        strict=True,
    ):
        panel_data = _paired_resource_panel_data(
            observations,
            metric=column,
        )
        if panel_data.empty:
            _empty_panel(ax, title)
            continue
        sns.lineplot(
            data=panel_data,
            x="plot_x",
            y=column,
            hue="provider_label",
            units="regest_id",
            estimator=None,
            sort=True,
            hue_order=provider_order,
            palette=palette,
            alpha=0.42,
            linewidth=1.0,
            ax=ax,
        )
        _plot_paired_trajectory_points(
            ax,
            panel_data=panel_data,
            metric=column,
            provider_order=provider_order,
            palette=palette,
        )
        ax.set_title(title)
        ax.set_xlabel("")
        ax.set_ylabel(y_label)
        ax.set_xticks(
            range(len(PAIRED_TRAJECTORY_CONDITION_ORDER)),
            PAIRED_TRAJECTORY_CONDITION_ORDER,
        )
        ax.set_xlim(-0.5, len(PAIRED_TRAJECTORY_CONDITION_ORDER) - 0.5)
        ax.set_box_aspect(1)
        if scale == "log" and (panel_data[column] > 0).all():
            ax.set_yscale("log")
            ut.configure_log_minor_ticks(ax)
        elif scale == "symlog":
            ax.set_yscale("symlog", linthresh=100)
            ax.set_ylim(bottom=0)
        ut.apply_grid(ax)
        ut.rotate_x_ticklabels(ax, rotation=25)

    _finish_figure(
        figure,
        axes,
        status_text=f"{status_text} · successful complete two-stage observations",
        provider_count=len(provider_order),
    )
    return figure


def _paired_resource_panel_data(
    observations: pd.DataFrame,
    *,
    metric: str,
    comparison_endpoints: bool = False,
) -> pd.DataFrame:
    """Select complete two-stage token observations that form paired segments.

    Resource observations are independent of Turtle syntax validity. They
    require both conditions to complete successfully with a numeric aggregate
    Stage-2 input measurement, preventing partial failed stages from being
    compared to a completed Turtle-generation request.

    :param observations: Per-regest condition observations from all providers.
    :param metric: Aggregate token field to display.
    :param comparison_endpoints: Keep both comparisons separate instead of
        pooling their shared DMW-plus-RAG observations.
    :return: Resource-valid observations that belong to at least one pair.
    """
    numeric = ut.numeric_column(observations, metric)
    eligible = observations.loc[
        ut.series_column(observations, "success").fillna(False).astype(bool)
        & ut.series_column(observations, "prompt_tokens_complete")
        .fillna(False)
        .astype(bool)
        & numeric.notna()
    ].copy()
    eligible[metric] = numeric.loc[eligible.index]
    paired_frames: list[pd.DataFrame] = []
    for comparison, left, right in zip(
        PAIR_SHEETS.values(),
        CONDITION_ORDER[:-1],
        CONDITION_ORDER[1:],
        strict=True,
    ):
        condition_rows = eligible.loc[
            ut.series_column(eligible, "condition").isin((left, right))
        ].copy()
        complete_keys = (
            condition_rows.groupby(["provider_label", "regest_id"])["condition"]
            .agg(set)
            .loc[lambda conditions: conditions == {left, right}]
            .index
        )
        condition_rows = condition_rows.set_index(
            ["provider_label", "regest_id"]
        )
        paired_frames.append(
            condition_rows.loc[condition_rows.index.isin(complete_keys)]
            .reset_index()
            .assign(
                comparison=comparison,
                pair_side=lambda frame: ut.series_column(
                    frame,
                    "condition",
                ).map({left: "left", right: "right"}),
            )
        )
    paired = pd.concat(paired_frames, ignore_index=True)
    if comparison_endpoints:
        endpoint_mapping = paired.apply(
            lambda row: PAIRED_COMPARISON_ENDPOINTS[
                (row["comparison"], row["pair_side"])
            ],
            axis="columns",
        )
        paired["plot_x"] = endpoint_mapping.map(lambda value: value[0])
        paired["endpoint_label"] = endpoint_mapping.map(lambda value: value[1])
        paired["trajectory_id"] = (
            ut.series_column(paired, "provider_label").astype(str)
            + "\x00"
            + ut.series_column(paired, "regest_id").astype(str)
            + "\x00"
            + ut.series_column(paired, "comparison").astype(str)
        )
        return _with_trajectory_jitter(
            paired,
            trajectory_column="trajectory_id",
        )

    paired = paired.drop_duplicates(
        subset=["provider_label", "regest_id", "condition"],
    )
    condition_positions = {
        condition: float(index)
        for index, condition in enumerate(CONDITION_ORDER)
    }
    paired["condition_label"] = ut.series_column(paired, "condition").map(
        CONDITION_LABELS
    )
    paired["plot_x"] = (
        ut.series_column(paired, "condition")
        .map(condition_positions)
        .astype(float)
    )
    return _with_trajectory_jitter(paired)


def _plot_paired_absolute_metrics(
    pairs: pd.DataFrame,
    *,
    observations: pd.DataFrame,
    provider_order: list[str],
    palette: dict[str, str],
    status_text: str,
) -> Figure:
    """Plot both valid paired comparisons as within-regest trajectories.

    :param pairs: Rows from both primary paired comparisons.
    :param provider_order: Provider labels in display order.
    :param palette: Provider-label color mapping.
    :param status_text: Evidence-status subtitle.
    :return: Completed valid-pair absolute metric figure.
    """
    absolute_pairs = _absolute_pair_observations(pairs)
    figure, axes = plt.subplots(
        2,
        3,
        figsize=SIX_PANEL_FIGURE_SIZE,
        squeeze=False,
    )
    plot_specs = (
        (
            "quality",
            "reuse_share",
            "Schema-reference reuse",
            "Reuse share",
        ),
        (
            "quality",
            "novel_schema_resources",
            "Novel schema resources",
            "Absent reference IRIs",
        ),
        ("quality", "triples", "Generated graph size", "Triples"),
        (
            "quality",
            "duration_minutes",
            "Duration",
            "Minutes (wall clock)",
        ),
        (
            "resource",
            "prompt_tokens",
            "Total input tokens for Turtle generation",
            "Tokens",
        ),
        ("resource", "output_tokens", "Output tokens", "Tokens"),
    )
    for index, (kind, column, title, y_label) in enumerate(plot_specs):
        ax = axes.flat[index]
        panel_data = (
            _paired_comparison_panel_data(absolute_pairs, metric=column)
            if kind == "quality"
            else _paired_resource_panel_data(
                observations,
                metric=column,
                comparison_endpoints=True,
            )
        )
        if panel_data.empty:
            _empty_panel(ax, title)
            continue
        sns.lineplot(
            data=panel_data,
            x="plot_x",
            y=column,
            hue="provider_label",
            units=_trajectory_unit_column(panel_data),
            estimator=None,
            sort=True,
            hue_order=provider_order,
            palette=palette,
            alpha=PAIR_TRAJECTORY_LINE_ALPHA,
            linewidth=0.5,
            ax=ax,
        )
        _plot_paired_trajectory_points(
            ax,
            panel_data=panel_data,
            metric=column,
            provider_order=provider_order,
            palette=palette,
        )
        ax.set_title(title)
        ax.set_xlabel("")
        ax.set_ylabel(y_label)
        _configure_paired_comparison_axis(ax)
        ax.set_box_aspect(0.95)
        uses_log_scale = (
            column
            in {
                "duration_minutes",
                "prompt_tokens",
                "output_tokens",
            }
            and (panel_data[column] > 0).all()
        )
        if column == "reuse_share":
            ut.configure_percent_axis(ax, ymin=0, ymax=1.05, major_max=1.0)
        elif uses_log_scale:
            ax.set_yscale("log")
            ut.configure_log_minor_ticks(ax)
        else:
            ax.set_ylim(bottom=0)
        ut.apply_grid(ax)
        ax.tick_params(axis="x", pad=3.5)
        annotation_base_ylim = ax.get_ylim()
        annotation_clearance_anchor = (
            float(panel_data[column].min()) if uses_log_scale else 0.0
        )
        for mean_line_index, provider_label in enumerate(
            provider_order,
            start=0,
        ):
            provider_data = panel_data.loc[
                ut.series_column(panel_data, "provider_label") == provider_label
            ]
            ut.annotate_xaxis_group_statistic(
                ax,
                data=provider_data,
                x="endpoint_label",
                y=column,
                x_order=PAIRED_COMPARISON_ENDPOINT_ORDER,
                statistic="mean",
                color=palette[provider_label],
                label="µ:",
                label_color=palette[provider_label],
                value_formatter=_format_compact_annotation_number,
                base_ylim=annotation_base_ylim,
                clearance_anchor=annotation_clearance_anchor,
                line_index=mean_line_index,
                line_spacing_points=8.0,
                fontsize=FIGURE_FONTSIZE,
                minimum_visible_y=(
                    annotation_base_ylim[0]
                    if annotation_base_ylim[0] == 0.0
                    else None
                ),
            )
        ut.annotate_xaxis_group_statistic(
            ax,
            data=panel_data,
            x="endpoint_label",
            y=column,
            x_order=PAIRED_COMPARISON_ENDPOINT_ORDER,
            statistic="count",
            hue="provider_label",
            hue_order=provider_order,
            palette=palette,
            label="n:",
            value_formatter=lambda value: str(int(value)),
            base_ylim=(ax.get_ylim()[0], annotation_base_ylim[1]),
            placement="top",
            box_width=0.5,
            fontsize=FIGURE_FONTSIZE,
        )
    _finish_figure(
        figure,
        axes,
        status_text=f"{status_text} · valid pairs only",
        provider_count=len(provider_order),
        layout_top=0.88,
        subplot_wspace=0.24,
        subplot_hspace=0.2,
        legend_colors={
            provider_label: palette[provider_label]
            for provider_label in provider_order
        },
    )
    return figure


def _absolute_pair_observations(pairs: pd.DataFrame) -> pd.DataFrame:
    """Expand valid pair rows into one absolute observation per side.

    :param pairs: Rows from both primary paired comparisons.
    :return: Long-form absolute metric observations.
    """
    records: list[dict[str, Any]] = []
    for pair in pairs.to_dict(orient="records"):
        if pair["valid_pair"] is not True:
            continue
        comparison = str(pair["comparison"])
        for side in ("left", "right"):
            label_key = (comparison, side)
            if label_key not in ABSOLUTE_PAIR_SIDE_LABELS:
                raise ValueError(
                    f"Unsupported pair side for absolute plot: {label_key!r}"
                )
            records.append(
                {
                    "regest_id": pair["regest_id"],
                    "provider_label": pair["provider_label"],
                    "comparison": comparison,
                    "pair_side": side,
                    "condition_label": ABSOLUTE_PAIR_SIDE_LABELS[label_key],
                    "reuse_share": pair[f"{side}_reuse_share"],
                    "novel_schema_resources": pair[
                        f"{side}_novel_schema_declaration_count"
                    ],
                    "triples": pair[f"{side}_triples"],
                    "duration_minutes": (
                        float(pair[f"{side}_duration_seconds"]) / 60.0
                    ),
                }
            )
    return ut.frame_from_records(
        records,
        columns=(
            "regest_id",
            "provider_label",
            "comparison",
            "pair_side",
            "condition_label",
            "reuse_share",
            "novel_schema_resources",
            "triples",
            "duration_minutes",
        ),
    )


def _paired_line_panel_data(
    observations: pd.DataFrame,
    *,
    comparison: str,
    metric: str,
) -> pd.DataFrame:
    """Select complete paired metric observations in categorical x-axis order.

    :param observations: Expanded valid-pair observations from both conditions.
    :param comparison: One named two-condition comparison.
    :param metric: Numeric observation field plotted on the y-axis.
    :return: Complete paired rows ordered from the left to the right condition.
    """
    expected_conditions = [
        ABSOLUTE_PAIR_SIDE_LABELS[(comparison, side)]
        for side in ("left", "right")
    ]
    numeric = ut.numeric_column(observations, metric)
    selected = observations.loc[
        (ut.series_column(observations, "comparison") == comparison)
        & numeric.notna()
    ].copy()
    selected[metric] = numeric.loc[selected.index]
    complete_pair_ids = (
        selected.groupby(["provider_label", "regest_id"])["condition_label"]
        .agg(set)
        .loc[lambda conditions: conditions == set(expected_conditions)]
        .index
    )
    selected = selected.set_index(["provider_label", "regest_id"])
    selected = selected.loc[
        selected.index.isin(complete_pair_ids)
    ].reset_index()
    selected["condition_label"] = pd.Categorical(
        selected["condition_label"],
        categories=expected_conditions,
        ordered=True,
    )
    return selected.sort_values(
        ["provider_label", "regest_id", "condition_label"]
    )


def _paired_comparison_panel_data(
    observations: pd.DataFrame,
    *,
    metric: str,
) -> pd.DataFrame:
    """Prepare two separate paired comparisons for one metric.

    :param observations: Expanded valid-pair observations from both conditions.
    :param metric: Numeric observation field plotted on the y-axis.
    :return: One endpoint per valid pair with four comparison-specific x slots.
    """
    frames = [
        _paired_line_panel_data(
            observations,
            comparison=comparison,
            metric=metric,
        )
        for comparison in PAIR_SHEETS.values()
    ]
    panel_data = pd.concat(frames, ignore_index=True)
    endpoint_mapping = panel_data.apply(
        lambda row: PAIRED_COMPARISON_ENDPOINTS[
            (row["comparison"], row["pair_side"])
        ],
        axis="columns",
    )
    panel_data["plot_x"] = endpoint_mapping.map(lambda value: value[0])
    panel_data["endpoint_label"] = endpoint_mapping.map(lambda value: value[1])
    panel_data["trajectory_id"] = (
        ut.series_column(panel_data, "provider_label").astype(str)
        + "\x00"
        + ut.series_column(panel_data, "regest_id").astype(str)
        + "\x00"
        + ut.series_column(panel_data, "comparison").astype(str)
    )
    return _with_trajectory_jitter(
        panel_data,
        trajectory_column="trajectory_id",
    )


def _with_trajectory_jitter(
    panel_data: pd.DataFrame,
    *,
    trajectory_column: str = "regest_id",
) -> pd.DataFrame:
    """Offset each paired trajectory consistently without changing its rank.

    :param panel_data: Long-form paired observations with centered ``plot_x``.
    :param trajectory_column: Stable value identifying one connected line.
    :return: Observations with a deterministic shared horizontal offset per
        provider/regest trajectory.
    """
    result = panel_data.copy()
    trajectory_keys = zip(
        ut.series_column(result, "provider_label"),
        ut.series_column(result, trajectory_column),
        strict=True,
    )
    result["plot_x"] = ut.numeric_column(result, "plot_x") + pd.Series(
        [
            _trajectory_jitter(provider, regest_id)
            for provider, regest_id in trajectory_keys
        ],
        index=result.index,
    )
    return result.sort_values(["provider_label", "regest_id", "plot_x"])


def _trajectory_unit_column(panel_data: pd.DataFrame) -> str:
    """Select the field that prevents lines crossing comparison groups.

    :param panel_data: Long-form paired observations for one panel.
    :return: Seaborn ``units`` column name.
    """
    return (
        "trajectory_id"
        if "trajectory_id" in panel_data.columns
        else "regest_id"
    )


def _configure_paired_comparison_axis(
    ax: Axes,
    *,
    tick_labels: tuple[str, str] | None = None,
) -> None:
    """Render two pairwise-comparison groups over four hidden endpoints.

    :param ax: Absolute paired-metric subplot.
    :param tick_labels: Optional compact labels for the two comparison groups.
    :return: None.
    """
    tick_positions = tuple(item[0] for item in PAIRED_COMPARISON_TICKS)
    labels = tick_labels or tuple(item[1] for item in PAIRED_COMPARISON_TICKS)
    ax.set_xticks(tick_positions, labels)
    ax.set_xlim(-0.5, 3.5)
    ax.axvline(
        1.5,
        color="#b0b0b0",
        linewidth=0.5,
        linestyle=":",
        zorder=0,
        label="_comparison-divider",
    )


def _trajectory_jitter(provider_label: object, regest_id: object) -> float:
    """Return a reproducible small x-offset for one paired trajectory.

    :param provider_label: Display identity of the provider.
    :param regest_id: Stable matched sample identifier.
    :return: Offset within ``±PAIR_TRAJECTORY_JITTER`` category units.
    """
    key = f"{provider_label}\x00{regest_id}".encode("utf-8")
    value = int.from_bytes(hashlib.sha256(key).digest()[:8], "big")
    unit_interval = value / ((1 << 64) - 1)
    return (unit_interval - 0.5) * 2.0 * PAIR_TRAJECTORY_JITTER


def _plot_paired_trajectory_points(
    ax: Axes,
    *,
    panel_data: pd.DataFrame,
    metric: str,
    provider_order: list[str],
    palette: dict[str, str],
) -> None:
    """Draw centralized-style endpoints at the jittered line coordinates.

    :param ax: Panel receiving the paired endpoints.
    :param panel_data: Jittered long-form paired observations.
    :param metric: Numeric y-axis field.
    :param provider_order: Provider labels in display order.
    :param palette: Provider colours keyed by display label.
    :return: None.
    """
    strip_kws = ut.DEFAULT_OVERLAY_STRIP_KWS
    for provider_label in provider_order:
        provider_data = panel_data.loc[
            ut.series_column(panel_data, "provider_label") == provider_label
        ]
        if provider_data.empty:
            continue
        collection = ax.scatter(
            provider_data["plot_x"],
            provider_data[metric],
            color=palette[provider_label],
            s=float(strip_kws["size"]) ** 2,
            linewidths=strip_kws["linewidth"],
            edgecolors=palette[provider_label],
            alpha=strip_kws["alpha"],
            zorder=strip_kws["zorder"],
        )
        ut.restyle_overlay_strip_collections(
            [collection],
            face_alpha=0.0,
            edge_alpha=0.6,
        )


def _format_compact_annotation_number(value: float) -> str:
    """Format a statistic without scientific notation for compact plot bands.

    :param value: Mean statistic displayed in a panel annotation.
    :return: Plain number below one thousand or a rounded ``k`` value above it.
    """
    magnitude = abs(float(value))
    if magnitude < 1_000:
        return f"{value:.3g}"
    compact = value / 1_000.0
    return f"{compact:.3g}k"


def _empty_panel(ax: Axes, title: str) -> None:
    """Mark a panel whose source workbook has no plottable observations.

    :param ax: Axes to replace with an explanatory message.
    :param title: Panel title.
    :return: None.
    """
    ax.set_title(title)
    ax.text(
        0.5,
        0.5,
        "No valid observations",
        transform=ax.transAxes,
        ha="center",
        va="center",
    )
    ax.set_axis_off()


def _finish_figure(
    figure: Figure,
    axes: Any,
    *,
    status_text: str,
    provider_count: int,
    layout_top: float = 0.90,
    layout_bottom: float = 0.01,
    subplot_wspace: float | None = None,
    subplot_hspace: float | None = None,
    legend_colors: dict[str, str] | None = None,
    show_legend: bool = True,
) -> None:
    """Apply one compact shared legend.

    :param figure: Figure to finalize.
    :param axes: Axes container holding provider legends.
    :param status_text: Retained source-export status, omitted from the figure.
    :param provider_count: Number of provider legend entries to retain.
    :param layout_top: Upper boundary reserved for the plotting grid.
    :param layout_bottom: Lower boundary reserved for local legends or labels.
    :param subplot_wspace: Optional relative horizontal space between subplots.
    :param subplot_hspace: Optional relative vertical space between subplots.
    :param legend_colors: Optional explicit legend labels and their colors.
    :param show_legend: Whether to place the collected legend in a shared top
        band.  Figures with multiple legend groups can reserve their own bands.
    :return: None.
    """
    del status_text
    handles, labels = ut.collect_legend_entries(axes)
    ut.remove_axis_legends(axes)
    if legend_colors is not None:
        handles, labels = ut.square_legend_entries(legend_colors)
    figure.tight_layout(rect=(0.01, layout_bottom, 0.99, layout_top))
    subplot_adjustments = {
        name: value
        for name, value in {
            "wspace": subplot_wspace,
            "hspace": subplot_hspace,
        }.items()
        if value is not None
    }
    if subplot_adjustments:
        figure.subplots_adjust(**subplot_adjustments)
    ut.refresh_above_xaxis_annotations(axes)
    if handles and show_legend:
        legend_ax = figure.add_axes((0.2, layout_top + 0.01, 0.6, 0.035))
        ut.add_top_band_figure_legend(
            figure,
            legend_ax,
            handles=handles[:provider_count],
            labels=labels[:provider_count],
            ncol=provider_count,
            frameon=False,
            show_handles=True,
        )


def _write_image_captions(
    path: Path,
    *,
    figure_stems: tuple[str, ...],
) -> None:
    """Write one reader-facing caption for every exported figure.

    Captions keep study-design explanations out of the visual plotting area,
    where they compete with panel titles, legends, and data annotations.

    :param path: Markdown destination beside the exported figure files.
    :param figure_stems: Figure stems in the same order as their exports.
    :return: None after writing the complete caption document.
    :raises ValueError: If an exported figure has no maintained caption.
    """
    captions: list[str] = []
    for stem in figure_stems:
        caption = FIGURE_CAPTIONS.get(stem)
        if caption is None:
            raise ValueError(f"No image caption is defined for {stem!r}.")
        captions.extend((f"## `{stem}`", "", caption, ""))
    path.write_text(
        "# Figure captions\n\n"
        "Use these captions with the correspondingly named PDF, PNG, or SVG "
        "figure files.\n\n" + "\n".join(captions),
        encoding="utf-8",
    )


def _source_display_path(path: Path) -> str:
    """Return a useful source label without recording a host-absolute path.

    :param path: Absolute source workbook path.
    :return: Run, analysis directory, and workbook filename.
    """
    parents = path.parents
    if len(parents) >= 2:
        return str(Path(parents[1].name) / parents[0].name / path.name)
    return path.name


def _build_parser() -> argparse.ArgumentParser:
    """Build the command-line interface.

    :return: Configured argument parser.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Plot DMW–Haiu results from one or more exported overview.xlsx "
            "workbooks."
        )
    )
    parser.add_argument(
        "workbooks",
        nargs="+",
        type=Path,
        help="Exported overview.xlsx workbook paths.",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        help=(
            "Parent for plots-TIMESTAMP. Defaults to the nearest common "
            "workbook directory."
        ),
    )
    parser.add_argument(
        "--quality-review-workbook",
        type=Path,
        help=(
            "Optional evaluated historian-review workbook. Requires "
            "--quality-reveal-key and adds historian-quality figures."
        ),
    )
    parser.add_argument(
        "--quality-reveal-key",
        type=Path,
        help=(
            "Condition reveal key for --quality-review-workbook. Required "
            "when that workbook is supplied."
        ),
    )
    return parser


def main() -> None:
    """Run the workbook plotting command.

    :return: None.
    """
    args = _build_parser().parse_args()
    output_dir = plot_workbooks(
        args.workbooks,
        output_root=args.output_root,
        quality_review_workbook=args.quality_review_workbook,
        quality_reveal_key=args.quality_reveal_key,
    )
    print(output_dir)


if __name__ == "__main__":
    main()
