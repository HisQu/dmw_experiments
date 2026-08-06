"""Plot DMW–Haiu comparison results from one or more exported workbooks."""

from __future__ import annotations

import argparse
import hashlib
import json
import os
from dataclasses import dataclass
from datetime import datetime
from numbers import Real
from pathlib import Path
from typing import Any, Mapping

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from matplotlib.patches import Rectangle
from matplotlib.text import Text
from openpyxl import load_workbook

import haiu.utils as ut
from dmw_experiments.studies.datamodel_workflow_haiu_comparison.quality_grade_analysis import (
    CONDITION_LABELS as QUALITY_CONDITION_LABELS,
    CONDITION_ORDER as QUALITY_CONDITION_ORDER,
    QUALITY_COMPARISONS,
    QualityGradeAnalysis,
    build_quality_grade_analysis,
)
from dmw_experiments.studies.datamodel_workflow_haiu_comparison.quality_error_analysis import (
    QualityErrorAnalysis,
    build_quality_error_analysis,
)
from dmw_experiments.studies.datamodel_workflow_haiu_comparison.quality_grade_inputs import (
    load_historian_quality_error_counts,
    load_historian_quality_grades,
)
from dmw_experiments.studies.datamodel_workflow_haiu_comparison.quality_grade_workbook import (
    export_quality_grade_analysis_workbook,
)

CONDITION_ORDER = QUALITY_CONDITION_ORDER
CONDITION_LABELS = QUALITY_CONDITION_LABELS
PAIR_SHEETS = {
    "02_DMW_Context_AB": "DMW: full vs RAG",
    "03_DMW_vs_Haiu_RAG": "DMW RAG vs standalone",
}
QUALITY_PAIR_SHEETS = {
    ("workflow_full_ontology", "workflow_rag"): "DMW: full vs RAG",
    ("workflow_rag", "haiu_rag_ontologizer"): "DMW RAG vs standalone",
}
ABSOLUTE_PAIR_SIDE_LABELS = {
    ("DMW: full vs RAG", "left"): "DMW + Full Ontology",
    ("DMW: full vs RAG", "right"): "DMW + Haiu RAG",
    ("DMW RAG vs standalone", "left"): "DMW + Haiu RAG",
    ("DMW RAG vs standalone", "right"): "Standalone Haiu RAG",
}
PAIRED_TRAJECTORY_CONDITION_ORDER = (
    "DMW + Full Ontology",
    "DMW + Haiu RAG",
    "Standalone Haiu RAG",
)
PAIRED_COMPARISON_ENDPOINTS = {
    ("DMW: full vs RAG", "left"): (0.0, "dmw_full"),
    ("DMW: full vs RAG", "right"): (1.0, "dmw_rag_context"),
    ("DMW RAG vs standalone", "left"): (2.0, "dmw_rag_system"),
    ("DMW RAG vs standalone", "right"): (3.0, "standalone"),
}
PAIRED_COMPARISON_ENDPOINT_ORDER = tuple(
    endpoint_label for _, endpoint_label in PAIRED_COMPARISON_ENDPOINTS.values()
)
PAIRED_COMPARISON_TICKS: tuple[tuple[float, str], ...] = (
    (0.5, "DMW vs DMW+HAIU"),
    (2.5, "DMW+HAIU vs HAIU"),
)
OUTCOME_CONDITION_TICK_LABELS = ("DMW", "DMW+HAIU", "HAIU")
OUTCOME_COMPARISON_TICK_LABELS = (
    "DMW vs\nDMW+HAIU",
    "DMW+HAIU\nvs HAIU",
)
REQUIRED_SHEETS = {
    "00_About",
    "01_Results",
    "02_DMW_Context_AB",
    "03_DMW_vs_Haiu_RAG",
    "04_Observations",
    "09_Token_Accounting",
}
PROVIDER_COLORS = ("#0072B2", "#D55E00", "#009E73", "#CC79A7")
ERROR_CLASS_COLORS = {
    "Invalid Turtle": "#B07AA1",
    "Context / length": "#59A14F",
    "Provider timeout": "#EDC948",
    "Connection error": "#4E79A7",
    "Provenance rejection": "#9C755F",
    "Runner deadline": "#BAB0AC",
    "Other generation error": "#E15759",
}
FIGURE_FONTSIZE = 8
OUTCOME_FIGURE_SIZE = (8.6, 2.9)
# > Two 2 in-tall panels plus dedicated legend bands in a narrower export.
PAIRED_GRADE_FIGURE_SIZE = (6.0, 4.7)
PAIRED_RESOURCE_FIGURE_SIZE = (6.0, 3.6)
SIX_PANEL_FIGURE_SIZE = (7.7, 5.2)
# > Retains the 30%-narrowed width while accommodating two interaction rows.
PROVIDER_INTERACTION_FIGURE_SIZE = (5.04, 5.4)
# > Four side-by-side panels retain a compact 1.7 in of drawable height.
ERROR_PROFILE_FIGURE_SIZE = (8.8, 3.5)
BAR_WIDTH = 0.52
FALSE_ASSIGNMENT_BAR_WIDTH = 0.8
PAIR_TRAJECTORY_JITTER = 0.08
PAIR_TRAJECTORY_LINE_ALPHA = 0.3
GRADE_COLORS = {
    1: "#1A9850",
    2: "#66BD63",
    3: "#FEE08B",
    4: "#FDAE61",
    5: "#F46D43",
    6: "#D73027",
}
GRADE_LEGEND_LABELS = {
    1: "Grade 1 — correct, essentially complete",
    2: "Grade 2 — correct, minor omissions",
    3: "Grade 3 — incomplete, factually safe",
    4: "Grade 4 — ≥1 error; local/formal, patchable",
    5: "Grade 5 — ≥1 error; plausible historical error",
    6: "Grade 6 — ≥1 error; gross source-reading failure",
}
ERROR_INTERPRETATION_BAND_COLORS = {
    "0": "#009E73",
    "1": "#F0E442",
    "2": "#E69F00",
    "3+": "#D55E00",
}
ERROR_ASSERTION_BAND_COLORS = {
    "0": "#009E73",
    "1": "#F0E442",
    "2": "#E69F00",
    "3": "#D55E00",
    "4+": "#A50F15",
}
ERROR_COUNT_SEGMENT_ANNOTATION_MINIMUM_SHARE = 0.075
ERROR_COUNT_CHANGE_COLORS = {
    "improved": "#1B9E77",
    "unchanged": "#7F7F7F",
    "worsened": "#D73027",
}
ERROR_COUNT_CHANGE_LABELS = {
    "improved": "Fewer errors",
    "unchanged": "Unchanged",
    "worsened": "More errors",
}
GRADE_CHANGE_LABELS = {
    "improved": "Improved",
    "unchanged": "Unchanged",
    "worsened": "Worsened",
}
DIRECT_PAIR_ENDPOINTS = {
    ("DMW full ontology vs DMW RAG", "first"): (0.0, "dmw_full"),
    ("DMW full ontology vs DMW RAG", "second"): (1.0, "dmw_rag_context"),
    ("DMW RAG vs standalone Haiu RAG", "first"): (2.0, "dmw_rag_system"),
    ("DMW RAG vs standalone Haiu RAG", "second"): (3.0, "standalone"),
}
DIRECT_PAIR_ENDPOINT_LABELS = {
    "dmw_full": "DMW",
    "dmw_rag_context": "DMW+\nHAIU",
    "dmw_rag_system": "DMW+\nHAIU",
    "standalone": "HAIU",
}
FIGURE_CAPTIONS = {
    "outcomes": (
        "Completion outcomes by provider and condition. Valid Turtle shares "
        "and issue categories use all attempted regesta; paired-comparison "
        "shares use only the planned condition pairs."
    ),
    "paired-absolute-metrics": (
        "Within-regest changes in generated ontology metrics. Every trajectory "
        "connects the same provider/regest across a planned direct comparison; "
        "sample sizes therefore vary by metric according to valid outputs."
    ),
    "paired-quality-grades": (
        "Historian quality grades. Panel A gives ordinal grade distributions "
        "only for matched provider–regest pairs. The two DMW + Haiu RAG bars "
        "are the same condition, each restricted to the valid pairs of its "
        "adjacent comparison. Panel B classifies the right condition as "
        "improved, unchanged, or worsened. Lower grades are better."
    ),
    "quality-grade-provider-interaction": (
        "Provider interaction in regesta complete for all three conditions in "
        "both providers. Thin lines are paired regesta; thick transparent "
        "lines are provider trends with paired-bootstrap grade or Wilson "
        "false-assignment-rate intervals. The reported sign and McNemar tests "
        "are exploratory."
    ),
    "false-assignment-error-profile": (
        "False-assignment counts in matched provider–regest pairs, pooled only "
        "after matching within provider. Panels A and B show the distributions "
        "of independent false interpretations and false atomic assertions; "
        "Panels C and D show whether the right condition has fewer, the same, "
        "or more errors."
    ),
}


@dataclass(frozen=True)
class WorkbookResults:
    """Hold the plot-ready tables and identity from one workbook.

    :param path: Source workbook path.
    :param provider: Short provider-profile identifier.
    :param provider_label: Human-readable provider and quantization label.
    :param status: Export status shown in ``00_About``.
    :param results: Per-condition aggregate results.
    :param observations: Per-regest condition observations.
    :param pairs: Valid and invalid rows from both primary comparisons.
    """

    path: Path
    provider: str
    provider_label: str
    status: str
    results: pd.DataFrame
    observations: pd.DataFrame
    pairs: pd.DataFrame


def load_workbook_results(path: Path) -> WorkbookResults:
    """Read and validate the workbook tables used by the figures.

    :param path: Exported ``overview.xlsx`` file.
    :return: Validated workbook identity and plot-ready tables.
    """
    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"Workbook not found: {resolved}")

    excel_file = pd.ExcelFile(resolved)
    missing = sorted(REQUIRED_SHEETS.difference(excel_file.sheet_names))
    if missing:
        raise ValueError(
            f"{resolved} is missing required worksheets: {', '.join(missing)}"
        )

    about = _read_about(resolved)
    profile = _parse_provider_profile(about, resolved)
    provider = str(profile["name"])
    provider_label = _provider_label(profile)

    results = pd.read_excel(excel_file, sheet_name="01_Results")
    observations = pd.read_excel(excel_file, sheet_name="04_Observations")
    excluded_outcome_error_cells = _local_runtime_recovery_cells(resolved)
    observations["_outcome_error_excluded"] = observations.apply(
        lambda row: (
            (
                str(row["condition"]),
                _regest_id_text(row["regest_id"]),
            )
            in excluded_outcome_error_cells
        ),
        axis="columns",
    )
    pair_frames: list[pd.DataFrame] = []
    for sheet_name, comparison_label in PAIR_SHEETS.items():
        frame = pd.read_excel(excel_file, sheet_name=sheet_name)
        frame["comparison"] = comparison_label
        pair_frames.append(frame)
    pairs = pd.concat(pair_frames, ignore_index=True)

    _require_columns(
        results,
        {
            "condition",
            "observations",
            "valid_completed_count",
            "valid_completed_rate",
        },
        path=resolved,
        sheet_name="01_Results",
    )
    _require_columns(
        observations,
        {
            "condition",
            "regest_id",
            "success",
            "turtle_syntax_valid",
            "output_truncated",
            "failure_code",
            "error_message",
            "duration_seconds",
            "prompt_tokens",
            "prompt_tokens_complete",
            "output_tokens",
        },
        path=resolved,
        sheet_name="04_Observations",
    )
    _require_columns(
        pairs,
        {
            "valid_pair",
            "left_condition",
            "right_condition",
            "left_duration_seconds",
            "right_duration_seconds",
            "left_reuse_share",
            "right_reuse_share",
            "duration_delta_seconds",
            "reuse_share_delta",
            "left_novel_schema_declaration_count",
            "right_novel_schema_declaration_count",
            "left_triples",
            "right_triples",
        },
        path=resolved,
        sheet_name="02/03 paired results",
    )

    for frame in (results, observations, pairs):
        frame["provider"] = provider
        frame["provider_label"] = provider_label
    results["condition_label"] = ut.series_column(results, "condition").map(
        lambda value: CONDITION_LABELS.get(str(value))
    )
    observations["condition_label"] = ut.series_column(
        observations, "condition"
    ).map(lambda value: CONDITION_LABELS.get(str(value)))
    result_condition_labels = ut.series_column(results, "condition_label")
    observation_condition_labels = ut.series_column(
        observations, "condition_label"
    )
    unknown_conditions = sorted(
        set(
            ut.series_column(
                results.loc[result_condition_labels.isna()],
                "condition",
            )
        )
        | set(
            ut.series_column(
                observations.loc[observation_condition_labels.isna()],
                "condition",
            )
        )
    )
    if unknown_conditions:
        raise ValueError(
            f"{resolved} contains unsupported conditions: "
            f"{', '.join(str(value) for value in unknown_conditions)}"
        )

    return WorkbookResults(
        path=resolved,
        provider=provider,
        provider_label=provider_label,
        status=about.get("Status", "UNKNOWN STATUS"),
        results=results,
        observations=observations,
        pairs=pairs,
    )


def _local_runtime_recovery_cells(workbook_path: Path) -> set[tuple[str, str]]:
    """Find cells archived by local-runtime recovery amendments.

    :param workbook_path: Derived overview workbook beneath a run directory.
    :return: Condition and regest keys selected by a local-runtime amendment.
    """
    run_dir = workbook_path.parent.parent
    amendments_dir = run_dir / "summaries" / "amendments"
    if not amendments_dir.is_dir():
        return set()

    selected_cells: set[tuple[str, str]] = set()
    for amendment_path in sorted(amendments_dir.glob("*.json")):
        amendment = json.loads(amendment_path.read_text(encoding="utf-8"))
        if amendment.get("kind") != "local_runtime_recovery":
            continue
        amendment_id = amendment.get("amendment_id")
        if not isinstance(amendment_id, str) or not amendment_id:
            raise ValueError(
                f"{amendment_path} has no local-runtime amendment ID."
            )
        archive_root = run_dir / "superseded" / amendment_id / "raw"
        for artifact_path in archive_root.glob("*/*.json"):
            selected_cells.add((artifact_path.parent.name, artifact_path.stem))
    return selected_cells


def _regest_id_text(value: object) -> str:
    """Normalize numeric spreadsheet identifiers to their raw-file spelling.

    :param value: Regest identifier loaded from an observation worksheet.
    :return: Canonical text form used in raw artifact paths.
    """
    if isinstance(value, Real) and not isinstance(value, bool):
        numeric_value = float(value)
        if not np.isnan(numeric_value) and numeric_value.is_integer():
            return str(int(numeric_value))
    return str(value)


def plot_workbooks(
    workbook_paths: list[Path],
    *,
    output_root: Path | None = None,
    timestamp: str | None = None,
    quality_review_workbook: Path | None = None,
    quality_reveal_key: Path | None = None,
) -> Path:
    """Create timestamped figures from matched provider workbooks.

    :param workbook_paths: Exported ``overview.xlsx`` files to compare.
    :param output_root: Parent directory for the timestamped plot directory.
    :param timestamp: Optional deterministic timestamp for tests or reruns.
    :param quality_review_workbook: Optional graded historian-review workbook.
    :param quality_reveal_key: Condition mapping for the masked review IDs.
    :return: Directory containing PNG, PDF, and manifest outputs.
    """
    if not workbook_paths:
        raise ValueError("At least one workbook is required.")
    if (quality_review_workbook is None) != (quality_reveal_key is None):
        raise ValueError(
            "Historian quality plotting requires both the graded workbook "
            "and its reveal key."
        )

    workbooks = [load_workbook_results(path) for path in workbook_paths]
    providers = [workbook.provider for workbook in workbooks]
    duplicates = sorted(
        provider for provider in set(providers) if providers.count(provider) > 1
    )
    if duplicates:
        raise ValueError(
            "Each provider profile may appear only once; duplicates: "
            f"{', '.join(duplicates)}"
        )

    generated_at = datetime.now().astimezone()
    directory_timestamp = timestamp or generated_at.strftime("%Y%m%dT%H%M%S%Z")
    root = (
        Path(output_root).expanduser().resolve()
        if output_root is not None
        else _common_output_root([workbook.path for workbook in workbooks])
    )
    output_dir = root / f"plots-{directory_timestamp}"
    if output_dir.exists():
        raise FileExistsError(
            f"Plot output directory already exists: {output_dir}"
        )
    output_dir.mkdir(parents=True)

    ut.configure_matplotlib_defaults(fontsize=FIGURE_FONTSIZE)
    palette = {
        workbook.provider_label: PROVIDER_COLORS[index]
        for index, workbook in enumerate(workbooks)
    }
    results = pd.concat(
        [workbook.results for workbook in workbooks], ignore_index=True
    )
    observations = pd.concat(
        [workbook.observations for workbook in workbooks], ignore_index=True
    )
    pairs = pd.concat(
        [workbook.pairs for workbook in workbooks], ignore_index=True
    )
    provider_order = [workbook.provider_label for workbook in workbooks]
    status_text = _status_text(workbooks)

    figures = {
        "outcomes": _plot_outcomes(
            results,
            pairs,
            observations=observations,
            provider_order=provider_order,
            palette=palette,
            status_text=status_text,
        ),
        "paired-absolute-metrics": _plot_paired_absolute_metrics(
            pairs,
            observations=observations,
            provider_order=provider_order,
            palette=palette,
            status_text=status_text,
        ),
    }
    quality_grade_manifest: dict[str, Any] | None = None
    if quality_review_workbook is not None and quality_reveal_key is not None:
        grades = load_historian_quality_grades(
            workbook_path=quality_review_workbook,
            reveal_key_path=quality_reveal_key,
            provider_order=provider_order,
        )
        quality_analysis = build_quality_grade_analysis(grades)
        error_counts = load_historian_quality_error_counts(
            workbook_path=quality_review_workbook,
            reveal_key_path=quality_reveal_key,
            provider_order=provider_order,
        )
        error_analysis = build_quality_error_analysis(error_counts)
        quality_analysis_workbook = export_quality_grade_analysis_workbook(
            analysis=quality_analysis,
            path=output_dir / "quality-grade-analysis.xlsx",
            quality_review_workbook=quality_review_workbook,
            quality_reveal_key=quality_reveal_key,
            error_analysis=(
                error_analysis
                if not error_analysis.observations.empty
                else None
            ),
        )
        figures["paired-quality-grades"] = _plot_quality_grade_overview(
            quality_analysis,
            status_text=status_text,
        )
        provider_interaction = _plot_quality_grade_provider_interaction(
            quality_analysis,
            provider_order=provider_order,
            palette=palette,
            status_text=status_text,
        )
        if provider_interaction is not None:
            figures["quality-grade-provider-interaction"] = provider_interaction
        if _has_error_profile_pairs(error_analysis):
            figures["false-assignment-error-profile"] = (
                _plot_false_assignment_error_profile(error_analysis)
            )
        quality_grade_manifest = {
            "workbook": _source_display_path(quality_review_workbook),
            "reveal_key": _source_display_path(quality_reveal_key),
            "analysis_workbook": quality_analysis_workbook.name,
            "graded_observations": len(quality_analysis.observations),
            "complete_triplets": len(quality_analysis.complete_triplets),
            "providers": sorted(
                set(
                    ut.series_column(
                        quality_analysis.observations,
                        "provider_label",
                    ).astype(str)
                )
            ),
            "error_count_observations": len(error_analysis.observations),
            "interpretation_count_observations": int(
                error_analysis.observations["false_interpretations"]
                .notna()
                .sum()
            ),
            "assertion_count_observations": int(
                error_analysis.observations["false_assertions"].notna().sum()
            ),
            "matched_interpretation_pairs": len(
                error_analysis.matched_interpretation_pairs
            ),
            "matched_assertion_pairs": len(
                error_analysis.matched_assertion_pairs
            ),
        }
    try:
        for stem, figure in figures.items():
            ut.export_figure(
                figure,
                output_dir / stem,
                raster_dpi=200,
            )
    finally:
        for figure in figures.values():
            plt.close(figure)
    caption_path = output_dir / "image-captions.md"
    _write_image_captions(caption_path, figure_stems=tuple(figures))

    manifest = {
        "schema_version": 1,
        "generated_at": generated_at.isoformat(),
        "status": status_text,
        "inputs": [
            {
                "source": _source_display_path(workbook.path),
                "provider_profile": workbook.provider,
                "provider_label": workbook.provider_label,
                "export_status": workbook.status,
                "observations": len(workbook.observations),
            }
            for workbook in workbooks
        ],
        "figures": [
            f"{stem}.{suffix}"
            for stem in figures
            for suffix in ut.DEFAULT_FIGURE_FORMATS
        ],
        "caption_file": caption_path.name,
        "paired_metric_rule": "Only rows with valid_pair=true are plotted.",
    }
    if quality_grade_manifest is not None:
        manifest["historian_quality_grades"] = quality_grade_manifest
    (output_dir / "plot_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return output_dir


def _read_about(path: Path) -> dict[str, str]:
    """Read the two-column identity table from ``00_About``.

    :param path: Workbook path.
    :return: Non-empty key/value rows.
    """
    worksheet = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )["00_About"]
    return {
        str(key): str(value)
        for key, value in worksheet.iter_rows(values_only=True)
        if key is not None and value is not None
    }


def _parse_provider_profile(
    about: dict[str, str],
    path: Path,
) -> dict[str, Any]:
    """Parse the provider-profile JSON stored in the workbook.

    :param about: Key/value rows from ``00_About``.
    :param path: Workbook path used in validation messages.
    :return: Provider-profile mapping.
    """
    raw_profile = about.get("Provider profile")
    if raw_profile is None:
        raise ValueError(f"{path} has no Provider profile in 00_About.")
    try:
        profile = json.loads(raw_profile)
    except json.JSONDecodeError as exc:
        raise ValueError(
            f"{path} contains invalid Provider profile JSON."
        ) from exc
    if not isinstance(profile, dict) or not profile.get("name"):
        raise ValueError(f"{path} has no provider-profile name.")
    return profile


def _provider_label(profile: dict[str, Any]) -> str:
    """Build a concise label without hard-coding provider profiles.

    :param profile: Parsed provider-profile metadata.
    :return: Display label containing provider and quantization when present.
    """
    provider = str(profile.get("chat_provider") or profile["name"])
    provider_names = {
        "academiccloud": "AcademicCloud",
        "lmstudio": "LM Studio",
    }
    label = provider_names.get(provider.lower(), provider)
    quantization = profile.get("quantization")
    return f"{label} {quantization}" if quantization else label


def _require_columns(
    frame: pd.DataFrame,
    required: set[str],
    *,
    path: Path,
    sheet_name: str,
) -> None:
    """Reject workbook schema drift before creating partial figures.

    :param frame: Loaded worksheet data.
    :param required: Columns needed by the plotting code.
    :param path: Workbook path used in the error.
    :param sheet_name: Reader-facing worksheet identifier.
    :return: None.
    """
    missing = sorted(required.difference(frame.columns))
    if missing:
        raise ValueError(
            f"{path} worksheet {sheet_name} is missing columns: "
            f"{', '.join(missing)}"
        )


def _common_output_root(paths: list[Path]) -> Path:
    """Choose the nearest shared directory for plot output.

    :param paths: Resolved workbook paths.
    :return: Shared ancestor of the workbook-containing directories.
    """
    common = Path(os.path.commonpath([str(path.parent) for path in paths]))
    return common


def _status_text(workbooks: list[WorkbookResults]) -> str:
    """Summarize whether any source is a partial diagnostic export.

    :param workbooks: Loaded workbook results.
    :return: Figure subtitle describing evidentiary status.
    """
    statuses = {workbook.status for workbook in workbooks}
    if any("PARTIAL" in status.upper() for status in statuses):
        return "PARTIAL DIAGNOSTIC EXPORT — not publication evidence"
    if len(statuses) == 1:
        return next(iter(statuses))
    return " / ".join(sorted(statuses))


def _plot_outcomes(
    results: pd.DataFrame,
    pairs: pd.DataFrame,
    *,
    observations: pd.DataFrame,
    provider_order: list[str],
    palette: dict[str, str],
    status_text: str,
) -> Figure:
    """Plot valid ontology rates and valid paired-comparison denominators.

    :param results: Per-condition aggregate results from every provider.
    :param pairs: Rows from both primary paired comparisons.
    :param observations: Canonical per-regest condition records.
    :param provider_order: Provider labels in the requested display order.
    :param palette: Provider-label color mapping.
    :param status_text: Evidence-status subtitle.
    :return: Completed outcome figure.
    """
    figure, axes = plt.subplots(
        1,
        2 + len(provider_order),
        figsize=OUTCOME_FIGURE_SIZE,
        gridspec_kw={
            "width_ratios": (1.1, 1.0, *([1.0] * len(provider_order))),
        },
    )
    condition_order = [CONDITION_LABELS[value] for value in CONDITION_ORDER]

    sns.barplot(
        data=results,
        x="condition_label",
        y="valid_completed_rate",
        hue="provider_label",
        order=condition_order,
        hue_order=provider_order,
        palette=palette,
        errorbar=None,
        width=BAR_WIDTH,
        ax=axes[0],
    )
    axes[0].set_title("Valid completed ontologies")
    axes[0].set_xlabel("")
    axes[0].set_ylabel("Share of attempted regests")
    ut.configure_percent_axis(axes[0], ymin=0, ymax=1.0, major_max=1.0)
    ut.apply_grid(axes[0])
    axes[0].set_xticks(
        range(len(condition_order)),
        OUTCOME_CONDITION_TICK_LABELS,
    )
    success_texts = _annotate_success_bars(
        axes[0],
        results,
        provider_order,
        condition_order,
    )

    pair_summary = (
        pairs.groupby(["comparison", "provider_label"], observed=True)[
            "valid_pair"
        ]
        .agg(valid_pairs="sum", attempted_pairs="size")
        .reset_index()
    )
    pair_summary["valid_pair_rate"] = (
        pair_summary["valid_pairs"] / pair_summary["attempted_pairs"]
    )
    comparison_order = list(PAIR_SHEETS.values())
    sns.barplot(
        data=pair_summary,
        x="comparison",
        y="valid_pair_rate",
        hue="provider_label",
        order=comparison_order,
        hue_order=provider_order,
        palette=palette,
        errorbar=None,
        width=BAR_WIDTH,
        ax=axes[1],
    )
    axes[1].set_title("Valid paired comparisons")
    axes[1].set_xlabel("")
    axes[1].set_ylabel("Share of attempted pairs")
    ut.configure_percent_axis(axes[1], ymin=0, ymax=1.0, major_max=1.0)
    ut.apply_grid(axes[1])
    axes[1].set_xticks(
        range(len(comparison_order)),
        OUTCOME_COMPARISON_TICK_LABELS,
    )
    pair_texts = _annotate_pair_bars(
        axes[1],
        pair_summary,
        provider_order,
        comparison_order,
    )

    issue_assessments = [
        _plot_error_assessment(
            ax,
            observations=observations,
            provider_label=provider,
        )
        for ax, provider in zip(axes[2:], provider_order, strict=True)
    ]
    present_error_classes = set().union(
        *(
            error_classes
            for error_classes, _texts, _excluded_count in issue_assessments
        )
    )

    _finish_figure(
        figure,
        axes,
        status_text=status_text,
        provider_count=len(provider_order),
        layout_top=0.72,
        subplot_wspace=0.52,
        show_legend=False,
    )
    ut.ensure_top_text_headroom(axes[0], success_texts)
    ut.ensure_top_text_headroom(axes[1], pair_texts)
    for ax, (_error_classes, issue_texts, _excluded_count) in zip(
        axes[2:],
        issue_assessments,
        strict=True,
    ):
        ut.ensure_top_text_headroom(ax, issue_texts)
    provider_handles, provider_labels = ut.square_legend_entries(palette)
    provider_legend_ax = figure.add_axes((0.03, 0.82, 0.45, 0.06))
    ut.add_top_band_figure_legend(
        figure,
        provider_legend_ax,
        handles=provider_handles,
        labels=provider_labels,
        ncol=len(provider_labels),
        frameon=False,
        show_handles=True,
    )
    error_class_colors = {
        error_class: color
        for error_class, color in ERROR_CLASS_COLORS.items()
        if error_class in present_error_classes
    }
    if error_class_colors:
        error_handles, error_labels = ut.square_legend_entries(
            error_class_colors
        )
        error_legend_ax = figure.add_axes((0.53, 0.77, 0.44, 0.13))
        ut.add_top_band_figure_legend(
            figure,
            error_legend_ax,
            handles=error_handles,
            labels=error_labels,
            max_rows=2,
            frameon=False,
            show_handles=True,
        )
    return figure


def _plot_error_assessment(
    ax: Axes,
    *,
    observations: pd.DataFrame,
    provider_label: str,
) -> tuple[set[str], list[Text], int]:
    """Show every outcome that does not provide valid Turtle by condition.

    :param ax: Panel that receives one provider's stacked failure counts.
    :param observations: Canonical per-regest condition records.
    :param provider_label: Display identity of the provider represented here.
    :return: Error categories displayed in this provider panel, their count
        labels, and the number of documented configuration-invalid errors
        excluded from this outcome-error assessment.
    """
    provider_rows = observations.loc[
        observations["provider_label"] == provider_label
    ].copy()
    provider_rows = provider_rows.loc[
        ~(
            provider_rows["success"].astype(bool)
            & provider_rows["turtle_syntax_valid"].eq(True)
        )
    ]
    configuration_invalid = provider_rows.apply(
        _is_configuration_invalid_outcome_error,
        axis="columns",
    )
    excluded_count = int(configuration_invalid.sum())
    provider_rows = provider_rows.loc[~configuration_invalid]
    provider_rows["error_class"] = provider_rows.apply(
        _classify_failure,
        axis="columns",
    )
    bottom = pd.Series(0, index=CONDITION_ORDER, dtype="int64")
    for error_class, color in ERROR_CLASS_COLORS.items():
        counts = (
            provider_rows.loc[
                provider_rows["error_class"] == error_class,
                "condition",
            ]
            .value_counts()
            .reindex(CONDITION_ORDER, fill_value=0)
        )
        ax.bar(
            range(len(CONDITION_ORDER)),
            counts,
            bottom=bottom,
            color=color,
            width=BAR_WIDTH,
            label=error_class,
        )
        bottom += counts
    max_issue_count = max(1, int(bottom.max()))
    ax.set_title(f"Outcome issues\n{provider_label}", pad=10)
    ax.set_xlabel("")
    ax.set_xticks(range(len(CONDITION_ORDER)), OUTCOME_CONDITION_TICK_LABELS)
    ax.set_ylabel("Regests without valid Turtle")
    ax.set_ylim(0, max_issue_count)
    ut.apply_grid(ax)
    count_texts: list[Text] = []
    for x_position, count in enumerate(bottom):
        count_texts.append(
            ax.annotate(
                f"n={int(count)}",
                (x_position, int(count)),
                xytext=(0, 2),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=6.5,
                alpha=0.7,
            )
        )
    return set(provider_rows["error_class"]), count_texts, excluded_count


def _classify_failure(row: pd.Series) -> str:
    """Map one non-valid-Turtle outcome to a stable error category.

    :param row: Observation fields from the exported workbook.
    :return: Reader-facing error category for the stacked assessment panel.
    """
    turtle_syntax_valid = row.get("turtle_syntax_valid")
    if (
        bool(row.get("success"))
        and isinstance(turtle_syntax_valid, (bool, np.bool_, Real))
        and not bool(turtle_syntax_valid)
    ):
        return "Invalid Turtle"

    message = str(row.get("error_message") or "").lower()
    failure_code = str(row.get("failure_code") or "").lower()
    if (
        bool(row.get("output_truncated"))
        or "length limit" in message
        or failure_code == "model_context_window_exceeded"
    ):
        return "Context / length"
    if "timeout" in message or "timed out" in message:
        return "Provider timeout"
    if "connection" in message:
        return "Connection error"
    if (
        "fresh ontology observation" in message
        or "reused a prior ontology" in message
    ):
        return "Provenance rejection"
    if "condition_wall_clock_timeout" in failure_code:
        return "Runner deadline"
    return "Other generation error"


def _is_configuration_invalid_outcome_error(row: pd.Series) -> bool:
    """Skip abandoned LM Studio configuration attempts in outcome errors.

    :param row: One exported condition observation.
    :return: Whether the row is an invalid local configuration attempt.
    """
    provider = str(row.get("provider_label") or "").lower()
    if not provider.startswith("lm studio"):
        return False
    if bool(row.get("success")):
        return False
    if bool(row.get("_outcome_error_excluded")):
        return _classify_failure(row) == "Other generation error"
    recovery = row.get("local_runtime_recovery")
    if recovery is None or (isinstance(recovery, float) and np.isnan(recovery)):
        return False
    if isinstance(recovery, str) and recovery.strip().lower() in {
        "",
        "{}",
        "none",
        "nan",
    }:
        return False
    return _classify_failure(row) == "Other generation error"


def _annotate_success_bars(
    ax: Axes,
    results: pd.DataFrame,
    provider_order: list[str],
    condition_order: list[str],
) -> list[Text]:
    """Label success-rate bars with valid and attempted counts.

    :param ax: Outcome-rate axes.
    :param results: Per-condition aggregate results.
    :param provider_order: Provider labels in display order.
    :param condition_order: Condition labels in display order.
    :return: Created count labels for final dynamic boundary clearance.
    """
    rows = _bar_rows(
        results,
        category_column="condition_label",
        category_order=condition_order,
        provider_order=provider_order,
    )
    patches = [patch for patch in ax.patches if isinstance(patch, Rectangle)]
    count_texts: list[Text] = []
    for patch, row in zip(patches, rows, strict=False):
        if not patch.get_height():
            continue
        label = (
            f"{int(row['valid_completed_count'])}/{int(row['observations'])}"
        )
        count_texts.append(
            ax.annotate(
                label,
                (patch.get_x() + patch.get_width() / 2, patch.get_height()),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=6.5,
                alpha=0.7,
            )
        )
    return count_texts


def _annotate_pair_bars(
    ax: Axes,
    summary: pd.DataFrame,
    provider_order: list[str],
    comparison_order: list[str],
) -> list[Text]:
    """Label paired-result bars with valid and attempted counts.

    :param ax: Valid-pair-rate axes.
    :param summary: Aggregated valid and attempted pair counts.
    :param provider_order: Provider labels in display order.
    :param comparison_order: Primary comparisons in display order.
    :return: Created count labels for final dynamic boundary clearance.
    """
    rows = _bar_rows(
        summary,
        category_column="comparison",
        category_order=comparison_order,
        provider_order=provider_order,
    )
    patches = [patch for patch in ax.patches if isinstance(patch, Rectangle)]
    count_texts: list[Text] = []
    for patch, row in zip(patches, rows, strict=False):
        label = f"{int(row['valid_pairs'])}/{int(row['attempted_pairs'])}"
        count_texts.append(
            ax.annotate(
                label,
                (patch.get_x() + patch.get_width() / 2, patch.get_height()),
                xytext=(0, 3),
                textcoords="offset points",
                ha="center",
                va="bottom",
                fontsize=6.5,
                alpha=0.7,
            )
        )
    return count_texts


def _bar_rows(
    frame: pd.DataFrame,
    *,
    category_column: str,
    category_order: list[str],
    provider_order: list[str],
) -> list[pd.Series]:
    """Order table rows in the same hue-major order as seaborn bars.

    :param frame: Aggregate values represented by bars.
    :param category_column: Column mapped to the x-axis.
    :param category_order: X-axis categories in display order.
    :param provider_order: Provider labels in hue order.
    :return: Rows aligned with Matplotlib's bar-patch order.
    """
    indexed = frame.set_index([category_column, "provider_label"])
    rows: list[pd.Series] = []
    for provider in provider_order:
        for category in category_order:
            rows.append(indexed.loc[(category, provider)])
    return rows


def _plot_resource_usage(
    observations: pd.DataFrame,
    *,
    provider_order: list[str],
    palette: dict[str, str],
    status_text: str,
) -> Figure:
    """Plot paired complete prompt and output token observations.

    :param observations: Per-regest condition observations.
    :param provider_order: Provider labels in display order.
    :param palette: Provider-label color mapping.
    :param status_text: Evidence-status subtitle.
    :return: Completed resource-observation figure.
    """
    figure, axes = plt.subplots(1, 2, figsize=PAIRED_RESOURCE_FIGURE_SIZE)
    plot_specs = (
        (
            "prompt_tokens",
            "Total input tokens for Turtle generation",
            "Tokens",
            "log",
        ),
        ("output_tokens", "Output tokens", "Tokens", "symlog"),
    )
    for ax, (column, title, y_label, scale) in zip(
        axes,
        plot_specs,
        strict=True,
    ):
        panel_data = _paired_resource_panel_data(
            observations,
            metric=column,
        )
        if panel_data.empty:
            _empty_panel(ax, title)
            continue
        sns.lineplot(
            data=panel_data,
            x="plot_x",
            y=column,
            hue="provider_label",
            units="regest_id",
            estimator=None,
            sort=True,
            hue_order=provider_order,
            palette=palette,
            alpha=0.42,
            linewidth=1.0,
            ax=ax,
        )
        _plot_paired_trajectory_points(
            ax,
            panel_data=panel_data,
            metric=column,
            provider_order=provider_order,
            palette=palette,
        )
        ax.set_title(title)
        ax.set_xlabel("")
        ax.set_ylabel(y_label)
        ax.set_xticks(
            range(len(PAIRED_TRAJECTORY_CONDITION_ORDER)),
            PAIRED_TRAJECTORY_CONDITION_ORDER,
        )
        ax.set_xlim(-0.5, len(PAIRED_TRAJECTORY_CONDITION_ORDER) - 0.5)
        ax.set_box_aspect(1)
        if scale == "log" and (panel_data[column] > 0).all():
            ax.set_yscale("log")
            ut.configure_log_minor_ticks(ax)
        elif scale == "symlog":
            ax.set_yscale("symlog", linthresh=100)
            ax.set_ylim(bottom=0)
        ut.apply_grid(ax)
        ut.rotate_x_ticklabels(ax, rotation=25)

    _finish_figure(
        figure,
        axes,
        status_text=f"{status_text} · successful complete two-stage observations",
        provider_count=len(provider_order),
    )
    return figure


def _paired_resource_panel_data(
    observations: pd.DataFrame,
    *,
    metric: str,
    comparison_endpoints: bool = False,
) -> pd.DataFrame:
    """Select complete two-stage token observations that form paired segments.

    Resource observations are independent of Turtle syntax validity. They
    require both conditions to complete successfully with a numeric aggregate
    Stage-2 input measurement, preventing partial failed stages from being
    compared to a completed Turtle-generation request.

    :param observations: Per-regest condition observations from all providers.
    :param metric: Aggregate token field to display.
    :param comparison_endpoints: Keep both comparisons separate instead of
        pooling their shared DMW-plus-RAG observations.
    :return: Resource-valid observations that belong to at least one pair.
    """
    numeric = ut.numeric_column(observations, metric)
    eligible = observations.loc[
        ut.series_column(observations, "success").fillna(False).astype(bool)
        & ut.series_column(observations, "prompt_tokens_complete")
        .fillna(False)
        .astype(bool)
        & numeric.notna()
    ].copy()
    eligible[metric] = numeric.loc[eligible.index]
    paired_frames: list[pd.DataFrame] = []
    for comparison, left, right in zip(
        PAIR_SHEETS.values(),
        CONDITION_ORDER[:-1],
        CONDITION_ORDER[1:],
        strict=True,
    ):
        condition_rows = eligible.loc[
            ut.series_column(eligible, "condition").isin((left, right))
        ].copy()
        complete_keys = (
            condition_rows.groupby(["provider_label", "regest_id"])["condition"]
            .agg(set)
            .loc[lambda conditions: conditions == {left, right}]
            .index
        )
        condition_rows = condition_rows.set_index(
            ["provider_label", "regest_id"]
        )
        paired_frames.append(
            condition_rows.loc[condition_rows.index.isin(complete_keys)]
            .reset_index()
            .assign(
                comparison=comparison,
                pair_side=lambda frame: ut.series_column(
                    frame,
                    "condition",
                ).map({left: "left", right: "right"}),
            )
        )
    paired = pd.concat(paired_frames, ignore_index=True)
    if comparison_endpoints:
        endpoint_mapping = paired.apply(
            lambda row: PAIRED_COMPARISON_ENDPOINTS[
                (row["comparison"], row["pair_side"])
            ],
            axis="columns",
        )
        paired["plot_x"] = endpoint_mapping.map(lambda value: value[0])
        paired["endpoint_label"] = endpoint_mapping.map(lambda value: value[1])
        paired["trajectory_id"] = (
            ut.series_column(paired, "provider_label").astype(str)
            + "\x00"
            + ut.series_column(paired, "regest_id").astype(str)
            + "\x00"
            + ut.series_column(paired, "comparison").astype(str)
        )
        return _with_trajectory_jitter(
            paired,
            trajectory_column="trajectory_id",
        )

    paired = paired.drop_duplicates(
        subset=["provider_label", "regest_id", "condition"],
    )
    condition_positions = {
        condition: float(index)
        for index, condition in enumerate(CONDITION_ORDER)
    }
    paired["condition_label"] = ut.series_column(paired, "condition").map(
        CONDITION_LABELS
    )
    paired["plot_x"] = (
        ut.series_column(paired, "condition")
        .map(condition_positions)
        .astype(float)
    )
    return _with_trajectory_jitter(paired)


def _plot_paired_absolute_metrics(
    pairs: pd.DataFrame,
    *,
    observations: pd.DataFrame,
    provider_order: list[str],
    palette: dict[str, str],
    status_text: str,
) -> Figure:
    """Plot both valid paired comparisons as within-regest trajectories.

    :param pairs: Rows from both primary paired comparisons.
    :param provider_order: Provider labels in display order.
    :param palette: Provider-label color mapping.
    :param status_text: Evidence-status subtitle.
    :return: Completed valid-pair absolute metric figure.
    """
    absolute_pairs = _absolute_pair_observations(pairs)
    figure, axes = plt.subplots(
        2,
        3,
        figsize=SIX_PANEL_FIGURE_SIZE,
        squeeze=False,
    )
    plot_specs = (
        (
            "quality",
            "reuse_share",
            "Schema-reference reuse",
            "Reuse share",
        ),
        (
            "quality",
            "novel_schema_resources",
            "Novel schema resources",
            "Absent reference IRIs",
        ),
        ("quality", "triples", "Generated graph size", "Triples"),
        (
            "quality",
            "duration_minutes",
            "Duration",
            "Minutes (wall clock)",
        ),
        (
            "resource",
            "prompt_tokens",
            "Total input tokens for Turtle generation",
            "Tokens",
        ),
        ("resource", "output_tokens", "Output tokens", "Tokens"),
    )
    for index, (kind, column, title, y_label) in enumerate(plot_specs):
        ax = axes.flat[index]
        panel_data = (
            _paired_comparison_panel_data(absolute_pairs, metric=column)
            if kind == "quality"
            else _paired_resource_panel_data(
                observations,
                metric=column,
                comparison_endpoints=True,
            )
        )
        if panel_data.empty:
            _empty_panel(ax, title)
            continue
        sns.lineplot(
            data=panel_data,
            x="plot_x",
            y=column,
            hue="provider_label",
            units=_trajectory_unit_column(panel_data),
            estimator=None,
            sort=True,
            hue_order=provider_order,
            palette=palette,
            alpha=PAIR_TRAJECTORY_LINE_ALPHA,
            linewidth=0.5,
            ax=ax,
        )
        _plot_paired_trajectory_points(
            ax,
            panel_data=panel_data,
            metric=column,
            provider_order=provider_order,
            palette=palette,
        )
        ax.set_title(title)
        ax.set_xlabel("")
        ax.set_ylabel(y_label)
        _configure_paired_comparison_axis(ax)
        ax.set_box_aspect(0.95)
        uses_log_scale = (
            column
            in {
                "duration_minutes",
                "prompt_tokens",
                "output_tokens",
            }
            and (panel_data[column] > 0).all()
        )
        if column == "reuse_share":
            ut.configure_percent_axis(ax, ymin=0, ymax=1.05, major_max=1.0)
        elif uses_log_scale:
            ax.set_yscale("log")
            ut.configure_log_minor_ticks(ax)
        else:
            ax.set_ylim(bottom=0)
        ut.apply_grid(ax)
        ax.tick_params(axis="x", pad=3.5)
        annotation_base_ylim = ax.get_ylim()
        annotation_clearance_anchor = (
            float(panel_data[column].min()) if uses_log_scale else 0.0
        )
        for mean_line_index, provider_label in enumerate(
            provider_order,
            start=0,
        ):
            provider_data = panel_data.loc[
                ut.series_column(panel_data, "provider_label") == provider_label
            ]
            ut.annotate_xaxis_group_statistic(
                ax,
                data=provider_data,
                x="endpoint_label",
                y=column,
                x_order=PAIRED_COMPARISON_ENDPOINT_ORDER,
                statistic="mean",
                color=palette[provider_label],
                label="µ:",
                label_color=palette[provider_label],
                value_formatter=_format_compact_annotation_number,
                base_ylim=annotation_base_ylim,
                clearance_anchor=annotation_clearance_anchor,
                line_index=mean_line_index,
                line_spacing_points=8.0,
                fontsize=FIGURE_FONTSIZE,
                minimum_visible_y=(
                    annotation_base_ylim[0]
                    if annotation_base_ylim[0] == 0.0
                    else None
                ),
            )
        ut.annotate_xaxis_group_statistic(
            ax,
            data=panel_data,
            x="endpoint_label",
            y=column,
            x_order=PAIRED_COMPARISON_ENDPOINT_ORDER,
            statistic="count",
            hue="provider_label",
            hue_order=provider_order,
            palette=palette,
            label="n:",
            value_formatter=lambda value: str(int(value)),
            base_ylim=(ax.get_ylim()[0], annotation_base_ylim[1]),
            placement="top",
            box_width=0.5,
            fontsize=FIGURE_FONTSIZE,
        )
    _finish_figure(
        figure,
        axes,
        status_text=f"{status_text} · valid pairs only",
        provider_count=len(provider_order),
        layout_top=0.88,
        subplot_wspace=0.24,
        subplot_hspace=0.2,
        legend_colors={
            provider_label: palette[provider_label]
            for provider_label in provider_order
        },
    )
    return figure


def _plot_quality_grade_overview(
    analysis: QualityGradeAnalysis,
    *,
    status_text: str,
) -> Figure:
    """Plot the grade distribution and pooled paired grade changes.

    :param analysis: Validated and pre-calculated historian-grade tables.
    :param status_text: Evidence-status subtitle retained by figure formatting.
    :return: Two-panel historian-quality figure.
    """
    figure, axes = plt.subplots(
        1,
        2,
        figsize=PAIRED_GRADE_FIGURE_SIZE,
        squeeze=False,
    )
    flat_axes = axes.ravel()
    _plot_grade_distribution(
        flat_axes[0],
        analysis=analysis,
        title="A. Grade distribution",
    )
    _plot_pair_change_distribution(
        flat_axes[1],
        distribution=analysis.pooled_grade_change_distribution,
        pair_data=analysis.direct_duel_pairs,
        sample_value_column="first_grade",
        title="B. Paired grade changes",
        ylabel="Share of matched pairs",
    )
    _finish_figure(
        figure,
        flat_axes,
        status_text=f"{status_text} · completed grades only; direct pairs only",
        provider_count=0,
        layout_top=0.55,
        layout_bottom=0.1245,
        subplot_wspace=0.4,
        show_legend=False,
    )
    # > ``tight_layout`` reserves extra space for the shared n annotations.
    # > Restore the specified 2 in panel height after those annotations settle.
    figure.subplots_adjust(
        left=0.12,
        right=0.98,
        bottom=0.1245,
        top=0.55,
        wspace=0.4,
    )
    _add_grade_distribution_legend(figure)
    _add_paired_change_legend(figure)
    for axis in flat_axes:
        ut.ensure_top_text_headroom(axis, axis.texts)
    return figure


def _plot_pairwise_quality_grade_trajectories(
    ax: Axes,
    *,
    analysis: QualityGradeAnalysis,
    provider_order: list[str],
    palette: dict[str, str],
) -> None:
    """Draw one trajectory for every valid planned quality comparison.

    DMW versus DMW + Haiu RAG and DMW + Haiu RAG versus standalone Haiu RAG
    occupy independent endpoint pairs. A complete three-condition review group
    contributes one line to each direct comparison; it never creates a DMW
    versus standalone-Haiu trajectory.

    :param ax: Axis receiving the matched grade trajectories.
    :param analysis: Validated historian-grade calculation tables.
    :param provider_order: Provider labels in display order.
    :param palette: Provider colors keyed by display label.
    :return: None after rendering the panel.
    """
    panel_data = _quality_pair_grade_panel_data(analysis.direct_duel_pairs)
    if panel_data.empty:
        _empty_panel(ax, "A. Pairwise matched quality grades")
        return
    plotted_providers = [
        provider_label
        for provider_label in provider_order
        if provider_label in set(panel_data["provider_label"].astype(str))
    ]
    sns.lineplot(
        data=panel_data,
        x="plot_x",
        y="grade",
        hue="provider_label",
        units=_trajectory_unit_column(panel_data),
        estimator=None,
        sort=True,
        hue_order=plotted_providers,
        palette=palette,
        alpha=PAIR_TRAJECTORY_LINE_ALPHA,
        linewidth=0.5,
        ax=ax,
    )
    _plot_paired_trajectory_points(
        ax,
        panel_data=panel_data,
        metric="grade",
        provider_order=plotted_providers,
        palette=palette,
    )
    ax.set_title("A. Pairwise matched quality grades")
    ax.set_xlabel("")
    ax.set_ylabel("Grade (1 best, 6 worst)")
    _configure_paired_comparison_axis(
        ax,
        tick_labels=OUTCOME_COMPARISON_TICK_LABELS,
    )
    ax.set_ylim(0.5, 6.5)
    ax.set_yticks(range(1, 7))
    ax.set_box_aspect(0.95)
    ut.apply_grid(ax)
    _annotate_grade_trajectory_panel(
        ax,
        panel_data=panel_data,
        provider_order=plotted_providers,
        palette=palette,
    )


def _quality_pair_grade_panel_data(
    direct_duel_pairs: pd.DataFrame,
) -> pd.DataFrame:
    """Expand quality duels into independent, two-endpoint trajectories.

    :param direct_duel_pairs: Auditable paired-grade rows for the two planned
        comparisons.
    :return: Jittered endpoint rows ready for Seaborn lines and centralized
        annotations. A regest may contribute to both planned comparisons.
    """
    columns = (
        "provider_label",
        "regest_id",
        "comparison",
        "pair_side",
        "grade",
        "plot_x",
        "endpoint_label",
        "trajectory_id",
    )
    frames: list[pd.DataFrame] = []
    for (
        first_condition,
        second_condition,
    ), comparison in QUALITY_PAIR_SHEETS.items():
        pair_rows = direct_duel_pairs.loc[
            (direct_duel_pairs["first_condition"] == first_condition)
            & (direct_duel_pairs["second_condition"] == second_condition)
        ]
        if pair_rows.empty:
            continue
        left = pair_rows.loc[
            :, ["provider_label", "regest_id", "first_grade"]
        ].rename(columns={"first_grade": "grade"})
        left["comparison"] = comparison
        left["pair_side"] = "left"
        right = pair_rows.loc[
            :, ["provider_label", "regest_id", "second_grade"]
        ].rename(columns={"second_grade": "grade"})
        right["comparison"] = comparison
        right["pair_side"] = "right"
        frames.extend((left, right))
    if not frames:
        return ut.empty_frame(columns)
    panel_data = pd.concat(frames, ignore_index=True)
    endpoint_mapping = panel_data.apply(
        lambda row: PAIRED_COMPARISON_ENDPOINTS[
            (row["comparison"], row["pair_side"])
        ],
        axis="columns",
    )
    panel_data["plot_x"] = endpoint_mapping.map(lambda value: value[0])
    panel_data["endpoint_label"] = endpoint_mapping.map(lambda value: value[1])
    panel_data["trajectory_id"] = (
        ut.series_column(panel_data, "provider_label").astype(str)
        + "\x00"
        + ut.series_column(panel_data, "regest_id").astype(str)
        + "\x00"
        + ut.series_column(panel_data, "comparison").astype(str)
    )
    return _with_trajectory_jitter(
        panel_data,
        trajectory_column="trajectory_id",
    )


def _annotate_grade_trajectory_panel(
    ax: Axes,
    *,
    panel_data: pd.DataFrame,
    provider_order: list[str],
    palette: dict[str, str],
) -> None:
    """Add the shared mean and denominator annotations to grade trajectories.

    :param ax: Axis receiving the centralized annotation bands.
    :param panel_data: Long-form grade trajectories with paired endpoints.
    :param provider_order: Provider labels in display order.
    :param palette: Provider colors keyed by display label.
    :return: None after adding mean and count annotations.
    """
    annotation_data = panel_data.copy()
    annotation_data["endpoint_label"] = annotation_data[
        "endpoint_label"
    ].astype(str)
    annotation_base_ylim = ax.get_ylim()
    for line_index, provider_label in enumerate(provider_order):
        provider_data = annotation_data.loc[
            ut.series_column(annotation_data, "provider_label")
            == provider_label
        ]
        ut.annotate_xaxis_group_statistic(
            ax,
            data=provider_data,
            x="endpoint_label",
            y="grade",
            x_order=PAIRED_COMPARISON_ENDPOINT_ORDER,
            statistic="mean",
            color=palette[provider_label],
            label="µ:",
            label_color=palette[provider_label],
            value_formatter=lambda value: f"{value:.2f}",
            base_ylim=annotation_base_ylim,
            clearance_anchor=1.0,
            line_index=line_index,
            line_spacing_points=8.0,
            fontsize=FIGURE_FONTSIZE,
            minimum_visible_y=0.5,
        )
    ut.annotate_xaxis_group_statistic(
        ax,
        data=annotation_data,
        x="endpoint_label",
        y="grade",
        x_order=PAIRED_COMPARISON_ENDPOINT_ORDER,
        statistic="count",
        hue="provider_label",
        hue_order=provider_order,
        palette=palette,
        label="n:",
        value_formatter=lambda value: str(int(value)),
        base_ylim=(ax.get_ylim()[0], annotation_base_ylim[1]),
        placement="top",
        box_width=0.5,
        fontsize=FIGURE_FONTSIZE,
    )


def _plot_grade_distribution(
    ax: Axes,
    *,
    analysis: QualityGradeAnalysis,
    title: str,
) -> None:
    """Render the ordinal grade distribution at paired-comparison endpoints.

    :param ax: Axis receiving stacked grade-share bars.
    :param analysis: Validated historian-grade calculation tables. The paired
        distribution repeats DMW + Haiu RAG because its two direct
        comparisons have distinct valid populations.
    :param title: Lettered reader-facing panel title.
    :return: None after rendering the panel.
    """
    distribution = analysis.paired_grade_distribution
    if distribution.empty:
        _empty_panel(ax, title)
        return
    panel_data = _direct_pair_endpoint_data(distribution)
    endpoint_order = tuple(
        endpoint
        for endpoint in DIRECT_PAIR_ENDPOINT_LABELS
        if endpoint in set(panel_data["endpoint_label"])
    )
    positions = np.array(
        [
            float(
                panel_data.loc[
                    panel_data["endpoint_label"] == endpoint,
                    "plot_x",
                ].iloc[0]
            )
            for endpoint in endpoint_order
        ]
    )
    cumulative = np.zeros(len(endpoint_order))
    for grade in range(1, 7):
        shares = np.array(
            [
                float(
                    panel_data.loc[
                        (panel_data["endpoint_label"] == endpoint)
                        & (panel_data["grade"] == grade),
                        "share",
                    ].iloc[0]
                )
                for endpoint in endpoint_order
            ]
        )
        counts = np.array(
            [
                int(
                    panel_data.loc[
                        (panel_data["endpoint_label"] == endpoint)
                        & (panel_data["grade"] == grade),
                        "count",
                    ].iloc[0]
                )
                for endpoint in endpoint_order
            ]
        )
        bars = ax.bar(
            positions,
            shares,
            bottom=cumulative,
            color=GRADE_COLORS[grade],
            edgecolor="white",
            linewidth=0.45,
            width=0.68,
        )
        ax.bar_label(
            bars,
            labels=[
                str(count)
                if share >= ERROR_COUNT_SEGMENT_ANNOTATION_MINIMUM_SHARE
                else ""
                for count, share in zip(counts, shares, strict=True)
            ],
            label_type="center",
            color=("#1F1F1F" if grade in {2, 3, 4} else "white"),
            fontsize=FIGURE_FONTSIZE - 1,
            fontweight="bold",
            padding=0,
        )
        cumulative += shares
    ax.set_title(title)
    ax.set_xlabel("")
    ax.set_ylabel("Share of paired models")
    ut.configure_percent_axis(ax, ymin=0, ymax=1.0, major_max=1.0)
    ut.apply_grid(ax)
    _configure_direct_pair_endpoint_axis(ax)
    paired_grades = _quality_pair_grade_panel_data(analysis.direct_duel_pairs)
    ut.annotate_xaxis_group_statistic(
        ax,
        data=paired_grades,
        x="endpoint_label",
        y="grade",
        x_order=PAIRED_COMPARISON_ENDPOINT_ORDER,
        statistic="count",
        label="n:",
        value_formatter=lambda value: str(int(value)),
        base_ylim=(0.0, 1.0),
        placement="top",
        box_width=0.6,
        fontsize=FIGURE_FONTSIZE,
    )


def _add_grade_distribution_legend(figure: Figure) -> None:
    """Add the six-row ordinal-grade color key above Panel A.

    :param figure: Two-panel grade figure receiving the reserved top-band key.
    :return: None after adding the local grade key.
    """
    legend_axis = figure.add_axes((0.04, 0.60, 0.48, 0.36))
    legend = ut.add_top_band_figure_legend(
        figure,
        legend_axis,
        handles=[
            Rectangle(
                (0, 0),
                1,
                1,
                facecolor=GRADE_COLORS[grade],
                edgecolor="none",
            )
            for grade in range(1, 7)
        ],
        labels=[GRADE_LEGEND_LABELS[grade] for grade in range(1, 7)],
        ncol=1,
        frameon=False,
        show_handles=True,
        loc="upper left",
        bbox_to_anchor=(0.0, 1.0),
    )
    for text in legend.get_texts():
        text.set_fontsize(FIGURE_FONTSIZE - 1)


def _add_paired_change_legend(figure: Figure) -> None:
    """Place the pooled improved/unchanged/worsened key above Panel B.

    :param figure: Two-panel grade figure receiving the shared direction key.
    :return: None after adding the compact top-band legend.
    """
    legend_axis = figure.add_axes((0.58, 0.80, 0.38, 0.05))
    handles = [
        Rectangle(
            (0, 0),
            1,
            1,
            facecolor=ERROR_COUNT_CHANGE_COLORS[direction],
            edgecolor="none",
        )
        for direction in ERROR_COUNT_CHANGE_COLORS
    ]
    ut.add_top_band_figure_legend(
        figure,
        legend_axis,
        handles=handles,
        labels=[
            GRADE_CHANGE_LABELS[direction]
            for direction in ERROR_COUNT_CHANGE_COLORS
        ],
        ncol=3,
        frameon=False,
        show_handles=True,
    )


def _plot_false_assignment_incidence(
    ax: Axes,
    *,
    analysis: QualityGradeAnalysis,
    provider_order: list[str],
    palette: dict[str, str],
) -> list[Text]:
    """Compare provider-specific incidence of rubric-defined false assignments.

    Grades 4–6 contain at least one false assignment under the historian
    rubric. DMW + Haiu RAG is deliberately repeated at the two direct-pair
    endpoints so the bar plot matches Panel A's paired-comparison layout.

    :param ax: Axis receiving the provider-hued incidence bars.
    :param analysis: Validated historian-grade calculation tables.
    :param provider_order: Provider labels in display order.
    :param palette: Provider colors keyed by display label.
    :return: Numerator/denominator annotation artists for final headroom.
    """
    panel_data = _false_assignment_incidence_panel_data(
        analysis.false_assignment_pair_summary
    )
    if panel_data.empty:
        _empty_panel(ax, "C. False-assignment incidence")
        return []
    plotted_providers = [
        provider_label
        for provider_label in provider_order
        if provider_label
        in set(ut.series_column(panel_data, "provider_label").astype(str))
    ]
    sns.barplot(
        data=panel_data,
        x="endpoint_label",
        y="false_assignment_share",
        hue="provider_label",
        order=PAIRED_COMPARISON_ENDPOINT_ORDER,
        hue_order=plotted_providers,
        palette=palette,
        errorbar=None,
        width=FALSE_ASSIGNMENT_BAR_WIDTH,
        ax=ax,
    )
    bar_rows = _bar_rows(
        panel_data,
        category_column="endpoint_label",
        category_order=list(PAIRED_COMPARISON_ENDPOINT_ORDER),
        provider_order=plotted_providers,
    )
    annotation_specs: list[tuple[float, float, str]] = []
    bar_patches = [
        patch
        for patch in ax.patches
        if isinstance(patch, Rectangle) and patch.get_width() > 0
    ]
    for patch, row in zip(bar_patches, bar_rows, strict=True):
        incidence = float(row["false_assignment_share"])
        lower = float(row["wilson_95_lower_share"])
        upper = float(row["wilson_95_upper_share"])
        ax.errorbar(
            patch.get_x() + (patch.get_width() / 2.0),
            incidence,
            yerr=[[incidence - lower], [upper - incidence]],
            fmt="none",
            ecolor="#4D4D4D",
            elinewidth=0.8,
            capsize=2.0,
            capthick=0.8,
            alpha=1.0,
            zorder=3.0,
        )
        bar_center = patch.get_x() + (patch.get_width() / 2.0)
        annotation_specs.append(
            (
                bar_center,
                upper,
                f"{int(row['false_assignment_count'])}/{int(row['models'])}",
            )
        )
    label_texts = ut.annotate_labels(
        ax,
        annotation_specs,
        fontsize=FIGURE_FONTSIZE - 1,
        fontweight="bold",
        colors=["#333333"] * len(annotation_specs),
    )
    for text in label_texts:
        text.set_zorder(3.2)
    ax.set_title("C. False-assignment incidence")
    ax.set_xlabel("")
    ax.set_ylabel("Share of analyses\nwith ≥1 false assignment")
    _configure_paired_comparison_axis(ax)
    ut.configure_percent_axis(
        ax,
        ymin=0,
        ymax=1.0,
        major_max=1.0,
        major_step=0.1,
        major_step_for_label=1,
    )
    ut.apply_grid(ax)
    return label_texts


def _false_assignment_incidence_panel_data(
    false_assignment_pair_summary: pd.DataFrame,
) -> pd.DataFrame:
    """Place matched false-assignment rates at paired-comparison endpoints.

    :param false_assignment_pair_summary: Provider-, comparison-, and
        condition-specific false-assignment calculations with Wilson bounds.
    :return: Long-form endpoint rows for a four-position paired bar plot.
    """
    if false_assignment_pair_summary.empty:
        return ut.empty_frame(
            [
                "provider_label",
                "condition",
                "endpoint_label",
                "models",
                "false_assignment_count",
                "false_assignment_share",
                "wilson_95_lower_share",
                "wilson_95_upper_share",
            ]
        )
    panel_data = false_assignment_pair_summary.copy()
    endpoint_mapping = panel_data.apply(
        lambda row: PAIRED_COMPARISON_ENDPOINTS[
            (
                QUALITY_PAIR_SHEETS[
                    (row["first_condition"], row["second_condition"])
                ],
                row["pair_side"],
            )
        ],
        axis="columns",
    )
    panel_data["plot_x"] = endpoint_mapping.map(lambda value: value[0])
    panel_data["endpoint_label"] = endpoint_mapping.map(lambda value: value[1])
    return panel_data


def _plot_false_assignment_error_profile(
    analysis: QualityErrorAnalysis,
) -> Figure:
    """Plot pooled matched false-assignment profiles and paired changes.

    Every pair is first formed within the same provider and regest. The figure
    then pools those valid provider-local pairs, which improves readability
    without creating an invalid AcademicCloud-to-LM-Studio comparison. Exact
    atomic assertions and independent false interpretations retain separate
    pair populations because the review fields are optional.

    :param analysis: Validated historian false-assignment count tables.
    :return: Four-panel pooled direct-pair false-assignment figure.
    """
    figure = plt.figure(figsize=ERROR_PROFILE_FIGURE_SIZE)
    grid = figure.add_gridspec(1, 4)
    axes = (
        figure.add_subplot(grid[0, 0]),
        figure.add_subplot(grid[0, 1]),
        figure.add_subplot(grid[0, 2]),
        figure.add_subplot(grid[0, 3]),
    )
    _plot_error_count_incidence(
        axes[0],
        incidence=analysis.pooled_interpretation_incidence,
        pairs=analysis.matched_interpretation_pairs,
        value_column="false_interpretations",
        ylabel="Share of matched models",
        title="A. Independent false\ninterpretations",
    )
    _plot_error_count_incidence(
        axes[1],
        incidence=analysis.pooled_assertion_incidence,
        pairs=analysis.matched_assertion_pairs,
        value_column="false_assertions",
        ylabel=None,
        title="B. False atomic\nassertions",
        band_colors=ERROR_ASSERTION_BAND_COLORS,
    )
    _plot_pair_change_distribution(
        axes[2],
        distribution=analysis.pooled_interpretation_change_distribution,
        pair_data=analysis.interpretation_pair_differences,
        sample_value_column="error_count_difference",
        title="C. Paired interpretation\nchanges",
        ylabel="Share of matched pairs",
    )
    _plot_pair_change_distribution(
        axes[3],
        distribution=analysis.pooled_assertion_change_distribution,
        pair_data=analysis.assertion_pair_differences,
        sample_value_column="error_count_difference",
        title="D. Paired false-assertion\nchanges",
        ylabel=None,
    )
    figure.subplots_adjust(
        left=0.065,
        right=0.99,
        bottom=0.18,
        top=0.67,
        wspace=0.46,
    )
    _add_error_profile_legends(figure)
    return figure


def _has_error_profile_pairs(analysis: QualityErrorAnalysis) -> bool:
    """Check whether either count measure has a valid direct pair to plot.

    :param analysis: Validated historian false-assignment count tables.
    :return: Whether a pooled profile has at least one provider-local pair.
    """
    return any(
        not pairs.empty
        for pairs in (
            analysis.matched_interpretation_pairs,
            analysis.matched_assertion_pairs,
        )
    )


def _plot_error_count_incidence(
    ax: Axes,
    *,
    incidence: pd.DataFrame,
    pairs: pd.DataFrame,
    value_column: str,
    ylabel: str | None,
    title: str,
    band_colors: Mapping[str, str] = ERROR_INTERPRETATION_BAND_COLORS,
) -> None:
    """Render pooled direct-pair count bands as stacked endpoint bars.

    :param ax: Axis receiving four paired-comparison endpoint bars.
    :param incidence: Pooled direct-pair rows in the four error-count bands.
    :param pairs: Raw direct provider–regest pairs for sample-size annotation.
    :param value_column: Numeric pair field counted by the n annotation.
    :param ylabel: Optional reader-facing share-axis title.
    :param title: Lettered panel title.
    :param band_colors: Display color keyed by every count band in incidence.
    :return: None after drawing the stacked distributions.
    """
    if incidence.empty:
        _empty_panel(ax, title)
        return
    panel_data = _direct_pair_endpoint_data(incidence)
    endpoint_order = tuple(
        endpoint
        for endpoint in DIRECT_PAIR_ENDPOINT_LABELS
        if endpoint in set(panel_data["endpoint_label"])
    )
    positions = np.array(
        [
            float(
                panel_data.loc[
                    panel_data["endpoint_label"] == endpoint,
                    "plot_x",
                ].iloc[0]
            )
            for endpoint in endpoint_order
        ]
    )
    cumulative = np.zeros(len(endpoint_order), dtype=float)
    for band, color in band_colors.items():
        shares = np.array(
            [
                float(
                    panel_data.loc[
                        (panel_data["endpoint_label"] == endpoint)
                        & (panel_data["error_count_band"] == band),
                        "share",
                    ].iloc[0]
                )
                for endpoint in endpoint_order
            ]
        )
        counts = np.array(
            [
                int(
                    panel_data.loc[
                        (panel_data["endpoint_label"] == endpoint)
                        & (panel_data["error_count_band"] == band),
                        "count",
                    ].iloc[0]
                )
                for endpoint in endpoint_order
            ]
        )
        bars = ax.bar(
            positions,
            shares,
            bottom=cumulative,
            color=color,
            edgecolor="white",
            linewidth=0.45,
            width=0.68,
        )
        ax.bar_label(
            bars,
            labels=[
                str(count)
                if share >= ERROR_COUNT_SEGMENT_ANNOTATION_MINIMUM_SHARE
                else ""
                for count, share in zip(counts, shares, strict=True)
            ],
            label_type="center",
            color=("#1F1F1F" if band in {"1", "2"} else "white"),
            fontsize=FIGURE_FONTSIZE - 1,
            fontweight="bold",
            padding=0,
        )
        cumulative += shares
    raw_panel_data = _error_profile_pair_panel_data(
        pairs,
        value_column=value_column,
    )
    ax.set_title(title)
    if ylabel is not None:
        ax.set_ylabel(ylabel)
    ut.configure_percent_axis(ax, ymin=0.0, ymax=1.0, major_max=1.0)
    ut.apply_grid(ax)
    _configure_direct_pair_endpoint_axis(ax)
    ut.annotate_xaxis_group_statistic(
        ax,
        data=raw_panel_data,
        x="endpoint_label",
        y=value_column,
        x_order=tuple(DIRECT_PAIR_ENDPOINT_LABELS),
        statistic="count",
        label="n:",
        value_formatter=lambda value: str(int(value)),
        base_ylim=ax.get_ylim(),
        placement="top",
        fontsize=FIGURE_FONTSIZE - 1,
        minimum_visible_y=0.0,
    )


def _plot_pair_change_distribution(
    ax: Axes,
    *,
    distribution: pd.DataFrame,
    pair_data: pd.DataFrame,
    sample_value_column: str,
    title: str,
    ylabel: str | None,
) -> None:
    """Render pooled direct-pair improvements, ties, and deteriorations.

    :param ax: Axis receiving two planned direct-comparison bars.
    :param distribution: Precalculated pooled pair counts and shares by change
        direction.
    :param pair_data: Raw direct pairs used for the matched n annotations.
    :param sample_value_column: Populated pair field used only to count the
        direct matched denominator above each comparison.
    :param title: Lettered panel title.
    :param ylabel: Optional reader-facing share-axis title.
    :return: None after drawing the mutually exclusive paired-change bars.
    """
    if distribution.empty:
        _empty_panel(ax, title)
        return
    comparison_order = tuple(
        comparison.key
        for comparison in QUALITY_COMPARISONS
        if comparison.key in set(distribution["comparison"])
    )
    positions = np.arange(len(comparison_order), dtype=float)
    cumulative = np.zeros(len(comparison_order), dtype=float)
    for direction, color in ERROR_COUNT_CHANGE_COLORS.items():
        shares: list[float] = []
        counts: list[int] = []
        for comparison in comparison_order:
            row = distribution.loc[
                (distribution["comparison"] == comparison)
                & (distribution["change_direction"] == direction)
            ].iloc[0]
            shares.append(float(row["share"]))
            counts.append(int(row["count"]))
        share_values = np.array(shares)
        ax.bar(
            positions,
            share_values,
            bottom=cumulative,
            width=0.62,
            color=color,
            edgecolor="white",
            linewidth=0.45,
            label=ERROR_COUNT_CHANGE_LABELS[direction],
        )
        for position, count, share, bottom in zip(
            positions,
            counts,
            share_values,
            cumulative,
            strict=True,
        ):
            if count:
                ax.text(
                    position,
                    bottom + (share / 2.0),
                    str(count),
                    ha="center",
                    va="center",
                    fontsize=FIGURE_FONTSIZE - 1,
                    fontweight="bold",
                    color=("#1F1F1F" if direction == "unchanged" else "white"),
                    zorder=4,
                )
        cumulative += share_values
    ax.set_title(title)
    if ylabel is not None:
        ax.set_ylabel(ylabel)
    ut.configure_percent_axis(ax, ymin=0.0, ymax=1.0, major_max=1.0)
    ut.set_fixed_x_ticklabels(
        ax,
        tuple(
            OUTCOME_COMPARISON_TICK_LABELS[index]
            for index, comparison in enumerate(QUALITY_COMPARISONS)
            if comparison.key in comparison_order
        ),
        positions=tuple(positions),
    )
    ut.apply_grid(ax)
    ut.annotate_xaxis_group_statistic(
        ax,
        data=pair_data,
        x="comparison",
        y=sample_value_column,
        x_order=comparison_order,
        statistic="count",
        label="n:",
        value_formatter=lambda value: str(int(value)),
        base_ylim=ax.get_ylim(),
        placement="top",
        fontsize=FIGURE_FONTSIZE - 1,
        minimum_visible_y=0.0,
    )


def _error_profile_pair_panel_data(
    pairs: pd.DataFrame,
    *,
    value_column: str,
) -> pd.DataFrame:
    """Expand direct pair values into jittered two-endpoint trajectories.

    :param pairs: Direct pair rows for one count metric.
    :param value_column: Count field stored as ``first_<field>`` and
        ``second_<field>`` in ``pairs``.
    :return: Long-form endpoints, one connected trajectory per provider,
        direct comparison, and regest.
    """
    columns = (
        "provider_label",
        "regest_id",
        "comparison",
        "pair_side",
        "condition",
        value_column,
        "plot_x",
        "endpoint_label",
        "trajectory_id",
    )
    if pairs.empty:
        return ut.empty_frame(columns)
    frames: list[pd.DataFrame] = []
    for side in ("first", "second"):
        frame = pairs.loc[
            :,
            (
                "provider_label",
                "regest_id",
                "comparison",
                f"{side}_condition",
                f"{side}_{value_column}",
            ),
        ].rename(
            columns={
                f"{side}_condition": "condition",
                f"{side}_{value_column}": value_column,
            }
        )
        frame["pair_side"] = side
        frames.append(frame)
    panel_data = pd.concat(frames, ignore_index=True)
    endpoint_mapping = panel_data.apply(
        lambda row: DIRECT_PAIR_ENDPOINTS[
            (str(row["comparison"]), str(row["pair_side"]))
        ],
        axis="columns",
    )
    panel_data["plot_x"] = endpoint_mapping.map(lambda value: value[0])
    panel_data["endpoint_label"] = endpoint_mapping.map(lambda value: value[1])
    panel_data["trajectory_id"] = (
        panel_data["provider_label"].astype(str)
        + "\x00"
        + panel_data["regest_id"].astype(str)
        + "\x00"
        + panel_data["comparison"].astype(str)
    )
    return _with_trajectory_jitter(
        panel_data,
        trajectory_column="trajectory_id",
    ).loc[:, columns]


def _direct_pair_endpoint_data(panel_data: pd.DataFrame) -> pd.DataFrame:
    """Attach common direct-comparison positions to summary endpoints.

    :param panel_data: Direct-pair summary rows with comparison and pair side.
    :return: Rows with their stable endpoint positions and display identities.
    """
    result = panel_data.copy()
    endpoint_mapping = result.apply(
        lambda row: DIRECT_PAIR_ENDPOINTS[
            (str(row["comparison"]), str(row["pair_side"]))
        ],
        axis="columns",
    )
    result["plot_x"] = endpoint_mapping.map(lambda value: value[0])
    result["endpoint_label"] = endpoint_mapping.map(lambda value: value[1])
    return result


def _configure_direct_pair_endpoint_axis(ax: Axes) -> None:
    """Show four direct-pair endpoints without redundant comparison labels.

    :param ax: Subplot receiving the shared direct-pair endpoint axis.
    :return: None after adding endpoint labels and the comparison divider.
    """
    endpoint_order = tuple(DIRECT_PAIR_ENDPOINT_LABELS)
    positions = tuple(range(len(endpoint_order)))
    ax.set_xlim(-0.5, 3.5)
    ax.set_xticks(
        positions,
        [DIRECT_PAIR_ENDPOINT_LABELS[endpoint] for endpoint in endpoint_order],
        fontsize=FIGURE_FONTSIZE - 2,
    )
    ax.axvline(
        1.5,
        color="#B0B0B0",
        linewidth=0.5,
        linestyle=":",
        zorder=0,
    )


def _add_error_profile_legends(figure: Figure) -> None:
    """Add one shared count-band key and a paired-change key above the figure.

    :param figure: Figure receiving compact reader-facing legends.
    :return: None after adding the two top-band legends.
    """
    count_legend_axis = figure.add_axes((0.12, 0.90, 0.76, 0.045))
    count_handles = [
        Rectangle(
            (0, 0),
            1,
            1,
            facecolor=color,
            edgecolor="none",
        )
        for color in (
            ERROR_INTERPRETATION_BAND_COLORS["0"],
            ERROR_INTERPRETATION_BAND_COLORS["1"],
            ERROR_INTERPRETATION_BAND_COLORS["2"],
            ERROR_INTERPRETATION_BAND_COLORS["3+"],
            ERROR_ASSERTION_BAND_COLORS["4+"],
        )
    ]
    ut.add_top_band_figure_legend(
        figure,
        count_legend_axis,
        handles=count_handles,
        labels=(
            "0 errors",
            "1 error",
            "2 errors",
            "3/3+ errors",
            "4+ errors",
        ),
        ncol=5,
        frameon=False,
        show_handles=True,
    )
    direction_legend_axis = figure.add_axes((0.55, 0.80, 0.41, 0.045))
    direction_handles = [
        Rectangle(
            (0, 0),
            1,
            1,
            facecolor=ERROR_COUNT_CHANGE_COLORS[direction],
            edgecolor="none",
        )
        for direction in ERROR_COUNT_CHANGE_COLORS
    ]
    ut.add_top_band_figure_legend(
        figure,
        direction_legend_axis,
        handles=direction_handles,
        labels=[
            ERROR_COUNT_CHANGE_LABELS[direction]
            for direction in ERROR_COUNT_CHANGE_COLORS
        ],
        ncol=3,
        frameon=False,
        show_handles=True,
    )


def _plot_quality_grade_provider_interaction(
    analysis: QualityGradeAnalysis,
    *,
    provider_order: list[str],
    palette: dict[str, str],
    status_text: str,
) -> Figure | None:
    """Plot matched grade and false-assignment provider interactions.

    :param analysis: Validated historian-grade calculation tables.
    :param provider_order: Provider labels in display order.
    :param palette: Provider colors keyed by display label.
    :param status_text: Evidence-status subtitle retained by figure formatting.
    :return: Exploratory two-row interaction figure, or ``None`` when no
        two-provider complete regesta are available.
    """
    pairs = analysis.provider_interaction_pairs
    if pairs.empty:
        return None
    provider_pairs = pairs.loc[
        :, ["first_provider", "second_provider"]
    ].drop_duplicates()
    if len(provider_pairs) != 1:
        return None
    first_provider, second_provider = provider_pairs.iloc[0]
    display_order = [
        provider_label
        for provider_label in provider_order
        if provider_label in {first_provider, second_provider}
    ]
    if len(display_order) != 2:
        display_order = [str(first_provider), str(second_provider)]
    figure, axes = plt.subplots(
        2,
        len(CONDITION_ORDER),
        figsize=PROVIDER_INTERACTION_FIGURE_SIZE,
        sharey="row",
        squeeze=False,
    )
    for condition_index, condition in enumerate(CONDITION_ORDER):
        condition_pairs = pairs.loc[pairs["condition"] == condition]
        grade_axis = axes[0, condition_index]
        grade_panel_data = _provider_interaction_panel_data(
            condition_pairs,
            provider_order=display_order,
        )
        _plot_provider_interaction_trajectories(
            grade_axis,
            panel_data=grade_panel_data,
            provider_order=display_order,
            palette=palette,
        )
        _plot_provider_interaction_central_trend(
            grade_axis,
            trend_data=_provider_grade_trend_panel_data(
                analysis.provider_interaction_trend_summary,
                condition=condition,
            ),
            provider_order=display_order,
            estimate_column="mean_grade",
            lower_column="bootstrap_95_lower_mean_grade",
            upper_column="bootstrap_95_upper_mean_grade",
        )
        grade_axis.set_title(
            {
                "workflow_full_ontology": "DMW full",
                "workflow_rag": "DMW + Haiu",
                "haiu_rag_ontologizer": "Standalone Haiu",
            }[condition]
        )
        _configure_provider_interaction_axis(
            grade_axis,
            provider_order=display_order,
        )
        grade_axis.set_ylim(0.5, 6.5)
        grade_axis.set_yticks(range(1, 7))
        ut.apply_grid(grade_axis)
        grade_summary = analysis.provider_interaction_summary.loc[
            analysis.provider_interaction_summary["condition"] == condition
        ]
        _annotate_provider_interaction_test(
            grade_axis,
            summary=grade_summary,
            p_value_column="exact_sign_test_p_value",
            tested_pairs_column="exact_sign_test_non_tied_pairs",
            tested_pairs_name="non-tied",
            test_name="sign test",
        )

        false_assignment_axis = axes[1, condition_index]
        false_assignment_panel_data = _provider_false_assignment_panel_data(
            condition_pairs,
            provider_order=display_order,
        )
        _plot_provider_interaction_trajectories(
            false_assignment_axis,
            panel_data=false_assignment_panel_data,
            provider_order=display_order,
            palette=palette,
        )
        _plot_provider_interaction_central_trend(
            false_assignment_axis,
            trend_data=_provider_false_assignment_trend_panel_data(
                analysis.provider_false_assignment_interaction_summary,
                condition=condition,
                provider_order=display_order,
            ),
            provider_order=display_order,
            estimate_column="false_assignment_share",
            lower_column="wilson_95_lower_share",
            upper_column="wilson_95_upper_share",
        )
        _configure_provider_interaction_axis(
            false_assignment_axis,
            provider_order=display_order,
        )
        false_assignment_axis.set_ylim(-0.12, 1.12)
        false_assignment_axis.set_yticks((0, 1), ("No", "≥1 false\nassignment"))
        ut.apply_grid(false_assignment_axis)
        false_assignment_summary = (
            analysis.provider_false_assignment_interaction_summary.loc[
                analysis.provider_false_assignment_interaction_summary[
                    "condition"
                ]
                == condition
            ]
        )
        _annotate_provider_interaction_test(
            false_assignment_axis,
            summary=false_assignment_summary,
            p_value_column="exact_mcnemar_p_value",
            tested_pairs_column="exact_mcnemar_discordant_pairs",
            tested_pairs_name="discordant",
            test_name="McNemar test",
        )
    axes[0, 0].set_ylabel("A. Grade\n(1 best, 6 worst)")
    axes[1, 0].set_ylabel("B. False-assignment\nindicator and rate")
    _finish_figure(
        figure,
        axes,
        status_text=f"{status_text} · shared complete regesta only",
        provider_count=0,
        layout_top=0.99,
        layout_bottom=0.10,
        subplot_wspace=0.28,
        subplot_hspace=0.65,
        show_legend=False,
    )
    return figure


def _plot_provider_interaction_trajectories(
    ax: Axes,
    *,
    panel_data: pd.DataFrame,
    provider_order: list[str],
    palette: dict[str, str],
) -> None:
    """Draw low-emphasis paired regest trajectories behind a trend overlay.

    :param ax: Provider-interaction axis receiving individual trajectories.
    :param panel_data: Jittered two-provider values for one condition.
    :param provider_order: Provider labels in the displayed x-axis order.
    :param palette: Provider colors keyed by display label.
    :return: None after plotting individual paired evidence.
    """
    for _regest_id, regest_data in panel_data.groupby(
        "regest_id",
        sort=False,
    ):
        ax.plot(
            regest_data["plot_x"],
            regest_data["grade"],
            color="#7F7F7F",
            linewidth=0.5,
            alpha=PAIR_TRAJECTORY_LINE_ALPHA,
            zorder=1,
        )
    _plot_paired_trajectory_points(
        ax,
        panel_data=panel_data,
        metric="grade",
        provider_order=provider_order,
        palette=palette,
    )


def _provider_interaction_panel_data(
    pairs: pd.DataFrame,
    *,
    provider_order: list[str],
) -> pd.DataFrame:
    """Expand wide provider pairs into transparent-overlay point records.

    :param pairs: One-condition provider pair rows.
    :param provider_order: Two provider labels in x-axis order.
    :return: Long-form grade points with one shared jitter offset per regest.
    """
    positions = {
        provider_label: float(index)
        for index, provider_label in enumerate(provider_order)
    }
    records: list[dict[str, object]] = []
    for pair in pairs.to_dict(orient="records"):
        grades = {
            str(pair["first_provider"]): int(pair["first_provider_grade"]),
            str(pair["second_provider"]): int(pair["second_provider_grade"]),
        }
        regest_id = str(pair["regest_id"])
        offset = _trajectory_jitter("provider-interaction", regest_id)
        for provider_label in provider_order:
            records.append(
                {
                    "provider_label": provider_label,
                    "regest_id": regest_id,
                    "grade": grades[provider_label],
                    "plot_x": positions[provider_label] + offset,
                }
            )
    return ut.frame_from_records(
        records,
        columns=("provider_label", "regest_id", "grade", "plot_x"),
    ).sort_values(["regest_id", "plot_x"])


def _provider_false_assignment_panel_data(
    pairs: pd.DataFrame,
    *,
    provider_order: list[str],
) -> pd.DataFrame:
    """Convert paired grades into the rubric's binary false-assignment field.

    :param pairs: One-condition provider grade pairs.
    :param provider_order: Two provider labels in x-axis order.
    :return: Jittered binary endpoint records where one means at least one
        false assignment under the grade rubric.
    """
    false_assignment_pairs = pairs.copy()
    false_assignment_pairs["first_provider_grade"] = (
        false_assignment_pairs["first_provider_grade"] >= 4
    ).astype(int)
    false_assignment_pairs["second_provider_grade"] = (
        false_assignment_pairs["second_provider_grade"] >= 4
    ).astype(int)
    return _provider_interaction_panel_data(
        false_assignment_pairs,
        provider_order=provider_order,
    )


def _provider_grade_trend_panel_data(
    summary: pd.DataFrame,
    *,
    condition: str,
) -> pd.DataFrame:
    """Select one condition's paired-bootstrap grade trend endpoints.

    :param summary: All matched-provider grade trend calculations.
    :param condition: Canonical condition identifier shown in one panel.
    :return: Provider endpoint rows with descriptive mean-grade intervals.
    """
    return summary.loc[summary["condition"] == condition].copy()


def _provider_false_assignment_trend_panel_data(
    summary: pd.DataFrame,
    *,
    condition: str,
    provider_order: list[str],
) -> pd.DataFrame:
    """Expand one matched false-assignment summary into provider endpoints.

    :param summary: One-row-per-condition matched provider rate calculations.
    :param condition: Canonical condition identifier shown in one panel.
    :param provider_order: Provider labels in x-axis order.
    :return: Provider endpoint rows with Wilson 95% rate intervals.
    """
    rows = summary.loc[summary["condition"] == condition]
    if rows.empty:
        return ut.empty_frame(
            [
                "provider_label",
                "false_assignment_share",
                "wilson_95_lower_share",
                "wilson_95_upper_share",
            ]
        )
    row = rows.iloc[0]
    first_provider = str(row["first_provider"])
    second_provider = str(row["second_provider"])
    values = {
        first_provider: (
            row["first_provider_false_assignment_share"],
            row["first_provider_wilson_95_lower_share"],
            row["first_provider_wilson_95_upper_share"],
        ),
        second_provider: (
            row["second_provider_false_assignment_share"],
            row["second_provider_wilson_95_lower_share"],
            row["second_provider_wilson_95_upper_share"],
        ),
    }
    return ut.frame_from_records(
        [
            {
                "provider_label": provider_label,
                "false_assignment_share": values[provider_label][0],
                "wilson_95_lower_share": values[provider_label][1],
                "wilson_95_upper_share": values[provider_label][2],
            }
            for provider_label in provider_order
            if provider_label in values
        ],
        columns=(
            "provider_label",
            "false_assignment_share",
            "wilson_95_lower_share",
            "wilson_95_upper_share",
        ),
    )


def _plot_provider_interaction_central_trend(
    ax: Axes,
    *,
    trend_data: pd.DataFrame,
    provider_order: list[str],
    estimate_column: str,
    lower_column: str,
    upper_column: str,
) -> None:
    """Overlay a high-emphasis matched-provider trend and its 95% intervals.

    :param ax: Provider-interaction axis receiving the central trend.
    :param trend_data: One estimate and interval per provider endpoint.
    :param provider_order: Provider labels in x-axis order.
    :param estimate_column: Central estimate field in ``trend_data``.
    :param lower_column: Inclusive 95% lower-bound field.
    :param upper_column: Inclusive 95% upper-bound field.
    :return: None after drawing the central line above individual trajectories.
    """
    if trend_data.empty:
        return
    indexed = trend_data.set_index("provider_label")
    plotted_providers = [
        provider_label
        for provider_label in provider_order
        if provider_label in indexed.index
    ]
    positions = np.arange(len(plotted_providers), dtype=float)
    estimates = np.array(
        [
            float(indexed.loc[provider_label, estimate_column])
            for provider_label in plotted_providers
        ]
    )
    lower = np.array(
        [
            float(indexed.loc[provider_label, lower_column])
            for provider_label in plotted_providers
        ]
    )
    upper = np.array(
        [
            float(indexed.loc[provider_label, upper_column])
            for provider_label in plotted_providers
        ]
    )
    ax.errorbar(
        positions,
        estimates,
        yerr=np.vstack((estimates - lower, upper - estimates)),
        fmt="none",
        ecolor="#3A3A3A",
        elinewidth=0.8,
        capsize=2.0,
        capthick=0.8,
        zorder=4,
    )
    ax.plot(
        positions,
        estimates,
        color="#202020",
        linewidth=2.8,
        alpha=0.45,
        marker="o",
        markersize=3.2,
        zorder=5,
    )


def _configure_provider_interaction_axis(
    ax: Axes,
    *,
    provider_order: list[str],
) -> None:
    """Apply compact provider labels to one interaction subplot.

    :param ax: Interaction subplot with provider positions zero and one.
    :param provider_order: Provider labels in display order.
    :return: None after configuring the shared x-axis.
    """
    ut.set_fixed_x_ticklabels(
        ax,
        provider_order,
        positions=range(len(provider_order)),
    )
    ut.rotate_x_ticklabels(ax, rotation=20, ha="right")
    ax.set_xlim(-0.25, len(provider_order) - 0.75)


def _annotate_provider_interaction_test(
    ax: Axes,
    *,
    summary: pd.DataFrame,
    p_value_column: str,
    tested_pairs_column: str,
    tested_pairs_name: str,
    test_name: str,
) -> None:
    """Add the correct paired exact-test result beneath one subplot.

    :param ax: Axis receiving the reader-facing test label.
    :param summary: One matched provider-comparison summary row.
    :param p_value_column: Exact-test p-value column in ``summary``.
    :param tested_pairs_column: Number of non-tied or discordant pairs.
    :param tested_pairs_name: Reader-facing description of those pairs.
    :param test_name: Concise exact paired test name.
    :return: None after adding the exact paired-test annotation.
    """
    if summary.empty:
        return
    row = summary.iloc[0]
    p_value = row[p_value_column]
    tested_pairs = int(row[tested_pairs_column])
    shared_regesta = int(row["shared_regesta"])
    if pd.isna(p_value):
        ax.set_xlabel(
            f"Exact paired {test_name}:\nunavailable (no {tested_pairs_name} pairs)",
            fontsize=FIGURE_FONTSIZE - 1,
        )
        return
    ax.set_xlabel(
        f"Exact paired {test_name}:\n"
        f"p = {float(p_value):.3f} "
        f"({tested_pairs} {tested_pairs_name} of {shared_regesta})",
        fontsize=FIGURE_FONTSIZE - 1,
    )


def _absolute_pair_observations(pairs: pd.DataFrame) -> pd.DataFrame:
    """Expand valid pair rows into one absolute observation per side.

    :param pairs: Rows from both primary paired comparisons.
    :return: Long-form absolute metric observations.
    """
    records: list[dict[str, Any]] = []
    for pair in pairs.to_dict(orient="records"):
        if pair["valid_pair"] is not True:
            continue
        comparison = str(pair["comparison"])
        for side in ("left", "right"):
            label_key = (comparison, side)
            if label_key not in ABSOLUTE_PAIR_SIDE_LABELS:
                raise ValueError(
                    f"Unsupported pair side for absolute plot: {label_key!r}"
                )
            records.append(
                {
                    "regest_id": pair["regest_id"],
                    "provider_label": pair["provider_label"],
                    "comparison": comparison,
                    "pair_side": side,
                    "condition_label": ABSOLUTE_PAIR_SIDE_LABELS[label_key],
                    "reuse_share": pair[f"{side}_reuse_share"],
                    "novel_schema_resources": pair[
                        f"{side}_novel_schema_declaration_count"
                    ],
                    "triples": pair[f"{side}_triples"],
                    "duration_minutes": (
                        float(pair[f"{side}_duration_seconds"]) / 60.0
                    ),
                }
            )
    return ut.frame_from_records(
        records,
        columns=(
            "regest_id",
            "provider_label",
            "comparison",
            "pair_side",
            "condition_label",
            "reuse_share",
            "novel_schema_resources",
            "triples",
            "duration_minutes",
        ),
    )


def _paired_line_panel_data(
    observations: pd.DataFrame,
    *,
    comparison: str,
    metric: str,
) -> pd.DataFrame:
    """Select complete paired metric observations in categorical x-axis order.

    :param observations: Expanded valid-pair observations from both conditions.
    :param comparison: One named two-condition comparison.
    :param metric: Numeric observation field plotted on the y-axis.
    :return: Complete paired rows ordered from the left to the right condition.
    """
    expected_conditions = [
        ABSOLUTE_PAIR_SIDE_LABELS[(comparison, side)]
        for side in ("left", "right")
    ]
    numeric = ut.numeric_column(observations, metric)
    selected = observations.loc[
        (ut.series_column(observations, "comparison") == comparison)
        & numeric.notna()
    ].copy()
    selected[metric] = numeric.loc[selected.index]
    complete_pair_ids = (
        selected.groupby(["provider_label", "regest_id"])["condition_label"]
        .agg(set)
        .loc[lambda conditions: conditions == set(expected_conditions)]
        .index
    )
    selected = selected.set_index(["provider_label", "regest_id"])
    selected = selected.loc[
        selected.index.isin(complete_pair_ids)
    ].reset_index()
    selected["condition_label"] = pd.Categorical(
        selected["condition_label"],
        categories=expected_conditions,
        ordered=True,
    )
    return selected.sort_values(
        ["provider_label", "regest_id", "condition_label"]
    )


def _paired_comparison_panel_data(
    observations: pd.DataFrame,
    *,
    metric: str,
) -> pd.DataFrame:
    """Prepare two separate paired comparisons for one metric.

    :param observations: Expanded valid-pair observations from both conditions.
    :param metric: Numeric observation field plotted on the y-axis.
    :return: One endpoint per valid pair with four comparison-specific x slots.
    """
    frames = [
        _paired_line_panel_data(
            observations,
            comparison=comparison,
            metric=metric,
        )
        for comparison in PAIR_SHEETS.values()
    ]
    panel_data = pd.concat(frames, ignore_index=True)
    endpoint_mapping = panel_data.apply(
        lambda row: PAIRED_COMPARISON_ENDPOINTS[
            (row["comparison"], row["pair_side"])
        ],
        axis="columns",
    )
    panel_data["plot_x"] = endpoint_mapping.map(lambda value: value[0])
    panel_data["endpoint_label"] = endpoint_mapping.map(lambda value: value[1])
    panel_data["trajectory_id"] = (
        ut.series_column(panel_data, "provider_label").astype(str)
        + "\x00"
        + ut.series_column(panel_data, "regest_id").astype(str)
        + "\x00"
        + ut.series_column(panel_data, "comparison").astype(str)
    )
    return _with_trajectory_jitter(
        panel_data,
        trajectory_column="trajectory_id",
    )


def _with_trajectory_jitter(
    panel_data: pd.DataFrame,
    *,
    trajectory_column: str = "regest_id",
) -> pd.DataFrame:
    """Offset each paired trajectory consistently without changing its rank.

    :param panel_data: Long-form paired observations with centered ``plot_x``.
    :param trajectory_column: Stable value identifying one connected line.
    :return: Observations with a deterministic shared horizontal offset per
        provider/regest trajectory.
    """
    result = panel_data.copy()
    trajectory_keys = zip(
        ut.series_column(result, "provider_label"),
        ut.series_column(result, trajectory_column),
        strict=True,
    )
    result["plot_x"] = ut.numeric_column(result, "plot_x") + pd.Series(
        [
            _trajectory_jitter(provider, regest_id)
            for provider, regest_id in trajectory_keys
        ],
        index=result.index,
    )
    return result.sort_values(["provider_label", "regest_id", "plot_x"])


def _trajectory_unit_column(panel_data: pd.DataFrame) -> str:
    """Select the field that prevents lines crossing comparison groups.

    :param panel_data: Long-form paired observations for one panel.
    :return: Seaborn ``units`` column name.
    """
    return (
        "trajectory_id"
        if "trajectory_id" in panel_data.columns
        else "regest_id"
    )


def _configure_paired_comparison_axis(
    ax: Axes,
    *,
    tick_labels: tuple[str, str] | None = None,
) -> None:
    """Render two pairwise-comparison groups over four hidden endpoints.

    :param ax: Absolute paired-metric subplot.
    :param tick_labels: Optional compact labels for the two comparison groups.
    :return: None.
    """
    tick_positions = tuple(item[0] for item in PAIRED_COMPARISON_TICKS)
    labels = tick_labels or tuple(item[1] for item in PAIRED_COMPARISON_TICKS)
    ax.set_xticks(tick_positions, labels)
    ax.set_xlim(-0.5, 3.5)
    ax.axvline(
        1.5,
        color="#b0b0b0",
        linewidth=0.5,
        linestyle=":",
        zorder=0,
        label="_comparison-divider",
    )


def _trajectory_jitter(provider_label: object, regest_id: object) -> float:
    """Return a reproducible small x-offset for one paired trajectory.

    :param provider_label: Display identity of the provider.
    :param regest_id: Stable matched sample identifier.
    :return: Offset within ``±PAIR_TRAJECTORY_JITTER`` category units.
    """
    key = f"{provider_label}\x00{regest_id}".encode("utf-8")
    value = int.from_bytes(hashlib.sha256(key).digest()[:8], "big")
    unit_interval = value / ((1 << 64) - 1)
    return (unit_interval - 0.5) * 2.0 * PAIR_TRAJECTORY_JITTER


def _plot_paired_trajectory_points(
    ax: Axes,
    *,
    panel_data: pd.DataFrame,
    metric: str,
    provider_order: list[str],
    palette: dict[str, str],
) -> None:
    """Draw centralized-style endpoints at the jittered line coordinates.

    :param ax: Panel receiving the paired endpoints.
    :param panel_data: Jittered long-form paired observations.
    :param metric: Numeric y-axis field.
    :param provider_order: Provider labels in display order.
    :param palette: Provider colours keyed by display label.
    :return: None.
    """
    strip_kws = ut.DEFAULT_OVERLAY_STRIP_KWS
    for provider_label in provider_order:
        provider_data = panel_data.loc[
            ut.series_column(panel_data, "provider_label") == provider_label
        ]
        if provider_data.empty:
            continue
        collection = ax.scatter(
            provider_data["plot_x"],
            provider_data[metric],
            color=palette[provider_label],
            s=float(strip_kws["size"]) ** 2,
            linewidths=strip_kws["linewidth"],
            edgecolors=palette[provider_label],
            alpha=strip_kws["alpha"],
            zorder=strip_kws["zorder"],
        )
        ut.restyle_overlay_strip_collections(
            [collection],
            face_alpha=0.0,
            edge_alpha=0.6,
        )


def _format_compact_annotation_number(value: float) -> str:
    """Format a statistic without scientific notation for compact plot bands.

    :param value: Mean statistic displayed in a panel annotation.
    :return: Plain number below one thousand or a rounded ``k`` value above it.
    """
    magnitude = abs(float(value))
    if magnitude < 1_000:
        return f"{value:.3g}"
    compact = value / 1_000.0
    return f"{compact:.3g}k"


def _empty_panel(ax: Axes, title: str) -> None:
    """Mark a panel whose source workbook has no plottable observations.

    :param ax: Axes to replace with an explanatory message.
    :param title: Panel title.
    :return: None.
    """
    ax.set_title(title)
    ax.text(
        0.5,
        0.5,
        "No valid observations",
        transform=ax.transAxes,
        ha="center",
        va="center",
    )
    ax.set_axis_off()


def _finish_figure(
    figure: Figure,
    axes: Any,
    *,
    status_text: str,
    provider_count: int,
    layout_top: float = 0.90,
    layout_bottom: float = 0.01,
    subplot_wspace: float | None = None,
    subplot_hspace: float | None = None,
    legend_colors: dict[str, str] | None = None,
    show_legend: bool = True,
) -> None:
    """Apply one compact shared legend.

    :param figure: Figure to finalize.
    :param axes: Axes container holding provider legends.
    :param status_text: Retained source-export status, omitted from the figure.
    :param provider_count: Number of provider legend entries to retain.
    :param layout_top: Upper boundary reserved for the plotting grid.
    :param layout_bottom: Lower boundary reserved for local legends or labels.
    :param subplot_wspace: Optional relative horizontal space between subplots.
    :param subplot_hspace: Optional relative vertical space between subplots.
    :param legend_colors: Optional explicit legend labels and their colors.
    :param show_legend: Whether to place the collected legend in a shared top
        band.  Figures with multiple legend groups can reserve their own bands.
    :return: None.
    """
    del status_text
    handles, labels = ut.collect_legend_entries(axes)
    ut.remove_axis_legends(axes)
    if legend_colors is not None:
        handles, labels = ut.square_legend_entries(legend_colors)
    figure.tight_layout(rect=(0.01, layout_bottom, 0.99, layout_top))
    subplot_adjustments = {
        name: value
        for name, value in {
            "wspace": subplot_wspace,
            "hspace": subplot_hspace,
        }.items()
        if value is not None
    }
    if subplot_adjustments:
        figure.subplots_adjust(**subplot_adjustments)
    ut.refresh_above_xaxis_annotations(axes)
    if handles and show_legend:
        legend_ax = figure.add_axes((0.2, layout_top + 0.01, 0.6, 0.035))
        ut.add_top_band_figure_legend(
            figure,
            legend_ax,
            handles=handles[:provider_count],
            labels=labels[:provider_count],
            ncol=provider_count,
            frameon=False,
            show_handles=True,
        )


def _write_image_captions(
    path: Path,
    *,
    figure_stems: tuple[str, ...],
) -> None:
    """Write one reader-facing caption for every exported figure.

    Captions keep study-design explanations out of the visual plotting area,
    where they compete with panel titles, legends, and data annotations.

    :param path: Markdown destination beside the exported figure files.
    :param figure_stems: Figure stems in the same order as their exports.
    :return: None after writing the complete caption document.
    :raises ValueError: If an exported figure has no maintained caption.
    """
    captions: list[str] = []
    for stem in figure_stems:
        caption = FIGURE_CAPTIONS.get(stem)
        if caption is None:
            raise ValueError(f"No image caption is defined for {stem!r}.")
        captions.extend((f"## `{stem}`", "", caption, ""))
    path.write_text(
        "# Figure captions\n\n"
        "Use these captions with the correspondingly named PDF, PNG, or SVG "
        "figure files.\n\n" + "\n".join(captions),
        encoding="utf-8",
    )


def _source_display_path(path: Path) -> str:
    """Return a useful source label without recording a host-absolute path.

    :param path: Absolute source workbook path.
    :return: Run, analysis directory, and workbook filename.
    """
    parents = path.parents
    if len(parents) >= 2:
        return str(Path(parents[1].name) / parents[0].name / path.name)
    return path.name


def _build_parser() -> argparse.ArgumentParser:
    """Build the command-line interface.

    :return: Configured argument parser.
    """
    parser = argparse.ArgumentParser(
        description=(
            "Plot DMW–Haiu results from one or more exported overview.xlsx "
            "workbooks."
        )
    )
    parser.add_argument(
        "workbooks",
        nargs="+",
        type=Path,
        help="Exported overview.xlsx workbook paths.",
    )
    parser.add_argument(
        "--output-root",
        type=Path,
        help=(
            "Parent for plots-TIMESTAMP. Defaults to the nearest common "
            "workbook directory."
        ),
    )
    parser.add_argument(
        "--quality-review-workbook",
        type=Path,
        help=(
            "Optional evaluated historian-review workbook. Requires "
            "--quality-reveal-key and adds historian-quality figures."
        ),
    )
    parser.add_argument(
        "--quality-reveal-key",
        type=Path,
        help=(
            "Condition reveal key for --quality-review-workbook. Required "
            "when that workbook is supplied."
        ),
    )
    return parser


def main() -> None:
    """Run the workbook plotting command.

    :return: None.
    """
    args = _build_parser().parse_args()
    output_dir = plot_workbooks(
        args.workbooks,
        output_root=args.output_root,
        quality_review_workbook=args.quality_review_workbook,
        quality_reveal_key=args.quality_reveal_key,
    )
    print(output_dir)


if __name__ == "__main__":
    main()
