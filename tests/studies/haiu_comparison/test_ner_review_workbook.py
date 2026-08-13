"""Tests for the adaptive historian NER review workbook."""

from __future__ import annotations

import hashlib
import json
import shutil
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from dmw_experiments.studies.haiu_comparison.analysis.workbooks import (
    export_historian_ner_review_workbook,
)
from dmw_experiments.studies.haiu_comparison.model.inputs import (
    canonical_json_sha256,
)
from dmw_experiments.studies.haiu_comparison.operations.repository_paths import (
    RUN_TEMPLATE_ROOT,
)


def test_single_provider_workbook_marks_spans_and_keeps_review_fields(
    tmp_path: Path,
) -> None:
    """One provider produces inline markers plus structured correction rows."""
    root = _minimal_run(tmp_path, include_lmstudio=False)
    workbook_path = (
        root / "analysis/workbooks/historian_ner_review_academiccloud_test.xlsx"
    )

    paths = export_historian_ner_review_workbook(
        {"academiccloud": root / "raw-academiccloud"},
        workbook_path=workbook_path,
    )

    workbook = load_workbook(paths.workbook, read_only=False)
    assert workbook.sheetnames == [
        "NER_Guide",
        "AcademicCloud_NER",
        "NER_Entities",
        "NER_Missing",
        "_Provenance",
        "_Validation",
    ]
    assert workbook["AcademicCloud_NER"].max_row == 3
    assert workbook["NER_Entities"].max_row == 6
    assert workbook["NER_Missing"].max_row == 3
    assert workbook["_Provenance"].sheet_state == "hidden"
    assert workbook["_Validation"].sheet_state == "hidden"
    assert len(workbook["NER_Entities"].data_validations.dataValidation) == 3

    with ZipFile(paths.workbook) as archive:
        shared_strings = archive.read("xl/sharedStrings.xml").decode("utf-8")
    assert "[1]" in shared_strings
    assert "≈[1]" in shared_strings
    assert any(color.removeprefix("#") in shared_strings for color in _COLORS)

    manifest = json.loads(paths.manifest.read_text(encoding="utf-8"))
    provider = manifest["providers"]["academiccloud"]
    assert provider["entity_count"] == 5
    assert provider["status_counts"] == {
        "resolved": 5,
        "ambiguous": 0,
        "unmatched": 0,
    }


def test_two_provider_workbook_uses_separate_review_sheets(
    tmp_path: Path,
) -> None:
    """The adaptive export exposes both provider identities in one workbook."""
    root = _minimal_run(tmp_path, include_lmstudio=True)
    paths = export_historian_ner_review_workbook(
        {
            "academiccloud": root / "raw-academiccloud",
            "lmstudio": root / "raw-lmstudio",
        },
        workbook_path=(
            root / "analysis/workbooks/"
            "historian_ner_review_academiccloud_lmstudio_test.xlsx"
        ),
    )

    workbook = load_workbook(paths.workbook, read_only=True)
    assert "AcademicCloud_NER" in workbook.sheetnames
    assert "LM_Studio_NER" in workbook.sheetnames
    entity_providers = {
        row[0].value for row in workbook["NER_Entities"].iter_rows(min_row=2)
    }
    assert entity_providers == {"AcademicCloud", "LM Studio"}


def test_strict_export_requires_every_scheduled_annotation(
    tmp_path: Path,
) -> None:
    """Missing shared annotation evidence is only legal in partial mode."""
    root = _minimal_run(tmp_path, include_lmstudio=False)
    annotation = (
        root / "raw-academiccloud/intermediates-shared_annotations/"
        "hsp-100-s01/annotation.json"
    )
    annotation.unlink()

    with pytest.raises(ValueError, match="Missing shared NER annotation"):
        export_historian_ner_review_workbook(
            {"academiccloud": root / "raw-academiccloud"},
            workbook_path=root / "analysis/workbooks/strict.xlsx",
        )

    paths = export_historian_ner_review_workbook(
        {"academiccloud": root / "raw-academiccloud"},
        workbook_path=root / "analysis/workbooks/partial.xlsx",
        allow_partial=True,
    )
    workbook = load_workbook(paths.workbook, read_only=True)
    statuses = [
        row[4].value
        for row in workbook["AcademicCloud_NER"].iter_rows(min_row=2)
    ]
    assert statuses == ["Unavailable", "Unavailable"]


def test_partial_two_provider_export_uses_scheduled_union(
    tmp_path: Path,
) -> None:
    """A missing provider schedule still receives unavailable union rows."""
    root = _minimal_run(tmp_path, include_lmstudio=True)
    manifest_path = root / "raw-lmstudio/manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["run"]["regest_ids"] = []
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    (
        root / "raw-lmstudio/intermediates-shared_annotations/"
        "hsp-100-s01/annotation.json"
    ).unlink()

    paths = export_historian_ner_review_workbook(
        {
            "academiccloud": root / "raw-academiccloud",
            "lmstudio": root / "raw-lmstudio",
        },
        workbook_path=root / "analysis/workbooks/partial-union.xlsx",
        allow_partial=True,
    )

    workbook = load_workbook(paths.workbook, read_only=True)
    statuses = [
        row[4].value for row in workbook["LM_Studio_NER"].iter_rows(min_row=2)
    ]
    assert statuses == ["Unavailable", "Unavailable"]


def test_provider_source_drift_is_rejected(tmp_path: Path) -> None:
    """Provider review sheets cannot silently display different source text."""
    root = _minimal_run(tmp_path, include_lmstudio=True)
    source_path = root / "raw-lmstudio/provenance/raw-regests/hsp-100-s01.json"
    source = json.loads(source_path.read_text(encoding="utf-8"))
    source["header"] = "Changed source"
    source_path.write_text(json.dumps(source), encoding="utf-8")

    with pytest.raises(ValueError, match="Frozen source text disagrees"):
        export_historian_ner_review_workbook(
            {
                "academiccloud": root / "raw-academiccloud",
                "lmstudio": root / "raw-lmstudio",
            },
            workbook_path=root / "analysis/workbooks/drift.xlsx",
        )


def test_provider_guideline_drift_is_rejected(tmp_path: Path) -> None:
    """Two provider sheets must share one frozen annotation policy."""
    root = _minimal_run(tmp_path, include_lmstudio=True)
    guideline = root / "raw-lmstudio/provenance/annotation_guidelines.md"
    guideline.write_text("Different guideline\n", encoding="utf-8")
    provenance_path = root / "raw-lmstudio/provenance/manifest.json"
    provenance = json.loads(provenance_path.read_text(encoding="utf-8"))
    provenance["inputs"]["annotation_guidelines"]["sha256"] = hashlib.sha256(
        guideline.read_bytes()
    ).hexdigest()
    provenance_path.write_text(json.dumps(provenance), encoding="utf-8")

    with pytest.raises(ValueError, match="guideline hashes disagree"):
        export_historian_ner_review_workbook(
            {
                "academiccloud": root / "raw-academiccloud",
                "lmstudio": root / "raw-lmstudio",
            },
            workbook_path=root / "analysis/workbooks/guideline-drift.xlsx",
        )


_COLORS = (
    "#1F4E78",
    "#9C0006",
    "#548235",
    "#7030A0",
    "#C65911",
    "#006666",
    "#7F6000",
    "#404040",
)


def _minimal_run(tmp_path: Path, *, include_lmstudio: bool) -> Path:
    """Create one strict one-unit run from the repository template."""
    root = tmp_path / "ner-test"
    shutil.copytree(RUN_TEMPLATE_ROOT, root)
    run_toml = root / "run.toml"
    contract = run_toml.read_text(encoding="utf-8")
    contract = contract.replace('mode = "full"', 'mode = "smoke"')
    contract = contract.replace('run_id = "template"', 'run_id = "ner-test"')
    contract = contract.replace("limit = 0", "limit = 1")
    if not include_lmstudio:
        contract = contract.replace(
            "[executions.lmstudio]\nenabled = true",
            "[executions.lmstudio]\nenabled = false",
        )
    run_toml.write_text(contract, encoding="utf-8")
    _write_minimal_catalog(root)
    for execution in (
        ("academiccloud", "lmstudio")
        if include_lmstudio
        else ("academiccloud",)
    ):
        _write_provider_evidence(root, execution)
    return root


def _write_minimal_catalog(root: Path) -> None:
    """Replace the large template population with one self-hashed record."""
    record = {
        "input_unit_id": "hsp-100-s01",
        "source_regest_id": "100",
        "source_subentry_index": 0,
        "source_sublemma_number": 1,
        "header": "Albertus Maguntia",
        "sublemma": "Johannes de Lapide",
        "source_regest_content_sha256": "a" * 64,
    }
    record["content_sha256"] = canonical_json_sha256(record)
    payload = {
        "schema_version": 1,
        "unit_kind": "header_sublemma_pair",
        "description": "NER workbook test",
        "source": {"source_run_id": "test-source"},
        "selection": {
            "source_regest_count": 1,
            "input_unit_count": 1,
            "excluded_header_only_regest_count": 0,
            "excluded_header_only_regest_ids": [],
        },
        "records": [record],
    }
    payload["catalogue_content_sha256"] = canonical_json_sha256(payload)
    (root / "INPUTS/header_sublemma_input_catalog.json").write_text(
        json.dumps(payload), encoding="utf-8"
    )


def _write_provider_evidence(root: Path, execution: str) -> None:
    """Write complete source, guideline, manifest, and annotation fixtures."""
    raw = root / f"raw-{execution}"
    manifest = {
        "schema_version": 3,
        "run": {"regest_ids": ["hsp-100-s01"]},
    }
    (raw / "manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    guideline = raw / "provenance/annotation_guidelines.md"
    guideline.write_text("Frozen NER guideline\n", encoding="utf-8")
    guideline_hash = hashlib.sha256(guideline.read_bytes()).hexdigest()
    provenance = {
        "inputs": {
            "annotation_guidelines": {
                "path": (
                    f"raw-{execution}/provenance/annotation_guidelines.md"
                ),
                "sha256": guideline_hash,
            }
        }
    }
    (raw / "provenance/manifest.json").write_text(
        json.dumps(provenance), encoding="utf-8"
    )
    source_dir = raw / "provenance/raw-regests"
    source_dir.mkdir(parents=True)
    (source_dir / "hsp-100-s01.json").write_text(
        json.dumps(
            {
                "regest_id": "hsp-100-s01",
                "header": "Albertus Maguntia",
                "subentries": ["Johannes de Lapide"],
            }
        ),
        encoding="utf-8",
    )
    annotation_dir = raw / "intermediates-shared_annotations/hsp-100-s01"
    annotation_dir.mkdir(parents=True)
    annotation = {
        "schema_version": 1,
        "regest_id": "hsp-100-s01",
        "version": "1.5.8",
        "content": {
            "header_entities": [
                {"type": "Person", "value": "Albertus"},
                {"type": "Ort", "value": "Maguntia"},
                {"type": "Diözese", "value": "Maguntia"},
                {"type": "Heading", "value": "Albertus Maguntia"},
            ],
            "subentry_entities": [
                {
                    "type": "Person",
                    "value": "Johanes de Lapide",
                    "subentry_index": 0,
                }
            ],
        },
        "annotation_model": "test-model",
    }
    (annotation_dir / "annotation.json").write_text(
        json.dumps(annotation), encoding="utf-8"
    )
