from __future__ import annotations

import hashlib
import json
import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

from dmw_experiments.studies.haiu_comparison.export_legacy_preliminary_workbook import (
    export_legacy_run,
)
from dmw_experiments.studies.haiu_comparison.export_results_workbook import (
    export_run,
)


LEGACY_CONDITIONS = (
    "workflow_full_ontology",
    "workflow_rag",
    "direct_llm_raw_regest",
)


def test_legacy_adapter_exports_limited_workbooks_without_touching_raw(
    tmp_path: Path,
) -> None:
    run_dir = _legacy_run(tmp_path)
    source_hashes_before = _source_hashes(run_dir)

    paths = export_legacy_run(run_dir)

    assert source_hashes_before == _source_hashes(run_dir)
    workbook = load_workbook(paths.workbook, read_only=True)
    assert workbook.sheetnames == [
        "00_Legacy_About",
        "01_Condition_Summary",
        "02_DMW_Context_AB",
        "03_Legacy_Direct_Context",
        "04_Observations",
        "05_Context_Timing",
        "06_Schema_Declarations",
        "07_Novel_Declarations",
        "08_Historian_Cases",
        "99_Definitions_Limits",
    ]

    summary = _rows_by_value(workbook["01_Condition_Summary"], "condition")
    assert (
        summary["workflow_full_ontology"][
            "independently_parseable_turtle_count"
        ]
        == 1
    )
    assert (
        summary["workflow_full_ontology"]["legacy_syntax_disagreement_count"]
        == 1
    )
    assert (
        summary["workflow_rag"]["median_schema_reference_iri_reuse_share"] == 1
    )

    pairs = list(workbook["02_DMW_Context_AB"].values)
    headers = pairs[0]
    eligible_index = headers.index("legacy_diagnostic_pair_eligible")
    assert sum(row[eligible_index] is True for row in pairs[1:]) == 1

    direct_rows = list(workbook["03_Legacy_Direct_Context"].values)
    assert "not Haiu-RAG" in str(direct_rows[1][2])

    review = load_workbook(paths.review_workbook, read_only=True)
    review_rows = list(review["Masked_Review"].values)
    assert "generated_turtle" in review_rows[2]
    assert "workflow_rag" not in str(review_rows[3:])

    analysis = json.loads(paths.manifest.read_text(encoding="utf-8"))
    assert analysis["publication_eligible"] is False
    assert analysis["pair_denominator"] == 1
    assert analysis["reference"]["source"] == (
        "fenced full-ontology planner prompt block"
    )
    assert "raw_ttl/haiu_retrieved/111.ttl" in analysis["source_hashes"]
    assert paths.reveal_key.is_file()
    assert analysis["audit_csv_enabled"] is False
    assert analysis["metric_definitions"]
    assert paths.workbook.name == "overview.xlsx"
    assert paths.review_workbook.name == "masked_case_review.xlsx"
    assert paths.reveal_key.name == "reveal_key.json"
    assert paths.manifest.name == "analysis_manifest.json"
    assert paths.readme.name == "README.md"
    assert sorted(path.name for path in paths.workbook.parent.iterdir()) == [
        "README.md",
        "analysis_manifest.json",
        "masked_case_review.xlsx",
        "overview.xlsx",
        "reveal_key.json",
    ]
    assert not list(paths.workbook.parent.glob("*.csv"))
    assert not (paths.workbook.parent / "audit_csv").exists()
    assert "PRELIMINARY LEGACY PILOT" in paths.readme.read_text(
        encoding="utf-8"
    )
    assert "`masked_case_review.xlsx`" in paths.readme.read_text(
        encoding="utf-8"
    )


def test_legacy_adapter_writes_opt_in_audit_without_system_comparison(
    tmp_path: Path,
) -> None:
    run_dir = _legacy_run(tmp_path)

    paths = export_legacy_run(run_dir, audit_csv=True)

    audit_dir = paths.workbook.parent / "audit_csv"
    assert sorted(path.name for path in audit_dir.iterdir()) == [
        "observations.csv",
        "pairs.csv",
        "schema_declarations.csv",
    ]
    with (audit_dir / "pairs.csv").open(encoding="utf-8", newline="") as handle:
        comparisons = {row["comparison"] for row in csv.DictReader(handle)}
    assert comparisons == {"workflow_full_ontology vs workflow_rag"}

    with pytest.raises(ValueError, match="not empty"):
        export_legacy_run(run_dir)
    export_legacy_run(run_dir, overwrite=True)
    assert not audit_dir.exists()


def test_legacy_adapter_rejects_changed_prompt_reference_hash(
    tmp_path: Path,
) -> None:
    run_dir = _legacy_run(tmp_path)
    raw_path = run_dir / "raw/workflow_full_ontology/111.json"
    raw = json.loads(raw_path.read_text(encoding="utf-8"))
    raw["ontology_ref"]["ttl_sha256"] = "0" * 64
    raw_path.write_text(json.dumps(raw), encoding="utf-8")

    with pytest.raises(ValueError, match="does not match"):
        export_legacy_run(run_dir)


def test_strict_exporter_still_rejects_legacy_artifact_contract(
    tmp_path: Path,
) -> None:
    run_dir = _legacy_run(tmp_path)

    with pytest.raises(ValueError, match="provenance"):
        export_run(run_dir)


def _legacy_run(tmp_path: Path) -> Path:
    run_dir = tmp_path / "legacy-run"
    reference = (
        "@prefix : <http://hisqu.de/rg_ontology/ontology/> .\n"
        "@prefix owl: <http://www.w3.org/2002/07/owl#> .\n"
        ":Reference a owl:Class ."
    )
    reference_hash = _sha256(reference)
    _write_json(
        run_dir / "summaries/run_manifest.json",
        {
            "run_id": "legacy-smoke",
            "regest_ids": ["111", "222", "333"],
            "conditions": list(LEGACY_CONDITIONS),
        },
    )

    annotation = {
        "regest_id": "111",
        "version": "0.0.1",
        "header_entities": [{"type": "Person", "value": "Ada"}],
        "subentry_entities": [],
    }
    _workflow_row(
        run_dir=run_dir,
        condition="workflow_full_ontology",
        regest_id="111",
        reference=reference,
        reference_hash=reference_hash,
        annotation=annotation,
        turtle=":Reference a owl:Class .\n:i_111 a :Reference .\n",
        recorded_syntax=True,
        stage2_reduction=4_096,
    )
    _workflow_row(
        run_dir=run_dir,
        condition="workflow_rag",
        regest_id="111",
        reference=reference,
        reference_hash=reference_hash,
        annotation=annotation,
        turtle=":Reference a owl:Class .\n:i_111 a :Reference .\n",
        recorded_syntax=True,
        stage2_reduction=None,
        include_flat_retrieval=True,
    )
    _workflow_row(
        run_dir=run_dir,
        condition="workflow_full_ontology",
        regest_id="222",
        reference=reference,
        reference_hash=reference_hash,
        annotation={**annotation, "regest_id": "222"},
        turtle=":broken a\n",
        recorded_syntax=True,
        stage2_reduction=None,
    )
    _workflow_row(
        run_dir=run_dir,
        condition="workflow_rag",
        regest_id="222",
        reference=reference,
        reference_hash=reference_hash,
        annotation={**annotation, "regest_id": "222"},
        turtle=":Reference a owl:Class .\n:i_222 a :Reference .\n",
        recorded_syntax=True,
        stage2_reduction=None,
        include_flat_retrieval=True,
    )
    _failed_workflow_row(
        run_dir=run_dir,
        condition="workflow_full_ontology",
        regest_id="333",
        error_message="NER generation failed",
    )
    _failed_workflow_row(
        run_dir=run_dir,
        condition="workflow_rag",
        regest_id="333",
        error_message="Paired condition was not submitted",
        failure_stage="shared_annotation_precondition",
    )
    _direct_row(run_dir)
    return run_dir


def _workflow_row(
    *,
    run_dir: Path,
    condition: str,
    regest_id: str,
    reference: str,
    reference_hash: str,
    annotation: dict[str, object],
    turtle: str,
    recorded_syntax: bool,
    stage2_reduction: int | None,
    include_flat_retrieval: bool = False,
) -> None:
    prompt = (
        "# Kontextressourcen\n\n"
        "## Referenzontologie (Ausschnitt, Turtle)\n"
        "Dies ist ein relevanter Ausschnitt.\n"
        f"```ttl\n{reference}\n```\n"
    )
    adjustment = (
        {
            "reduction_tokens": stage2_reduction,
            "requested_max_tokens": 20_000,
            "effective_max_tokens": 20_000 - stage2_reduction,
        }
        if stage2_reduction is not None
        else None
    )
    payload: dict[str, object] = {
        "condition": condition,
        "regest_id": regest_id,
        "success": True,
        "model": "qwen3.5-397b-a17b",
        "annotation_model": "qwen3.6-27b",
        "context_mode_requested": (
            "full_ontology" if condition == "workflow_full_ontology" else "rag"
        ),
        "context_mode_effective": (
            "full_ontology" if condition == "workflow_full_ontology" else "rag"
        ),
        "duration_seconds": 10.0,
        "total_attempt_duration_seconds": 10.0,
        "total_elapsed_seconds": 11.0,
        "prompt_tokens": 100,
        "prompt_tokens_complete": False,
        "prompt_tokens_source": "estimated_partial_stage1_only",
        "output_tokens": 12,
        "output_tokens_source": "estimated",
        "turtle_syntax_valid": recorded_syntax,
        "ontology_ref": {"ttl_sha256": reference_hash},
        "provider_run_metadata": {
            "stage2": {"context_window_adjustment": adjustment}
        },
        "raw_response": {
            "annotation_review": {"data": annotation},
            "debug_output": {"prompts": {"user": prompt}},
        },
    }
    _write_json(
        run_dir / f"raw/{condition}/{regest_id}.json",
        payload,
    )
    _write(run_dir / f"raw_yaml/{condition}/{regest_id}.yaml", "{}\n")
    _write(run_dir / f"raw_ttl/{condition}/{regest_id}.ttl", turtle)
    if include_flat_retrieval:
        _write(
            run_dir / f"raw_ttl/haiu_retrieved/{regest_id}.ttl",
            ":Reference a owl:Class .\n",
        )
        _write(
            run_dir / f"raw_ttl/haiu_retrieved/{regest_id}.yaml",
            "snapshot_fidelity: legacy_unknown\n",
        )


def _failed_workflow_row(
    *,
    run_dir: Path,
    condition: str,
    regest_id: str,
    error_message: str,
    failure_stage: str | None = None,
) -> None:
    payload: dict[str, object] = {
        "condition": condition,
        "regest_id": regest_id,
        "success": False,
        "error_message": error_message,
        "failure_stage": failure_stage,
        "turtle_syntax_valid": None,
    }
    _write_json(run_dir / f"raw/{condition}/{regest_id}.json", payload)
    _write(run_dir / f"raw_yaml/{condition}/{regest_id}.yaml", "{}\n")


def _direct_row(run_dir: Path) -> None:
    payload = {
        "condition": "direct_llm_raw_regest",
        "regest_id": "111",
        "success": True,
        "model": "qwen3.5-397b-a17b",
        "context_mode_requested": "none",
        "context_mode_effective": "none",
        "duration_seconds": 6.0,
        "total_attempt_duration_seconds": 6.0,
        "total_elapsed_seconds": 7.0,
        "prompt_tokens": 20,
        "prompt_tokens_complete": True,
        "prompt_tokens_source": "provider",
        "output_tokens": 8,
        "output_tokens_source": "estimated",
        "turtle_syntax_valid": True,
    }
    _write_json(run_dir / "raw/direct_llm_raw_regest/111.json", payload)
    _write(run_dir / "raw_yaml/direct_llm_raw_regest/111.yaml", "{}\n")
    _write(
        run_dir / "raw_ttl/direct_llm_raw_regest/111.ttl",
        ":Other a owl:Class .\n:i_111 a :Other .\n",
    )


def _rows_by_value(sheet, key: str) -> dict[str, dict[str, object]]:
    rows = list(sheet.values)
    headers = [str(value) for value in rows[0]]
    return {
        str(row[headers.index(key)]): dict(zip(headers, row, strict=True))
        for row in rows[1:]
    }


def _source_hashes(run_dir: Path) -> dict[str, str]:
    files = []
    for directory in ("raw", "raw_yaml", "raw_ttl", "prompts"):
        root = run_dir / directory
        if root.is_dir():
            files.extend(path for path in root.rglob("*") if path.is_file())
    return {
        path.relative_to(run_dir).as_posix(): _sha256(path)
        for path in sorted(files)
    }


def _write_json(path: Path, payload: dict[str, object]) -> None:
    _write(path, json.dumps(payload, indent=2))


def _write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def _sha256(value: str | Path) -> str:
    payload = (
        value.read_bytes() if isinstance(value, Path) else value.encode("utf-8")
    )
    return hashlib.sha256(payload).hexdigest()
