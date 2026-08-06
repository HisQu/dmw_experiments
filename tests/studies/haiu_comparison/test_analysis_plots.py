import json
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import pytest
from matplotlib.colors import to_rgba
from openpyxl import Workbook

from dmw_experiments.studies.haiu_comparison.analysis.plots.results import (
    PROVIDER_COLORS,
    _classify_failure,
    _plot_outcomes,
    _plot_paired_absolute_metrics,
    _status_text,
    load_workbook_results,
    plot_workbooks,
)
from dmw_experiments.studies.haiu_comparison.analysis.plots.quality import (
    _plot_false_assignment_error_profile,
    _plot_false_assignment_incidence,
    _plot_pairwise_quality_grade_trajectories,
    _plot_quality_grade_overview,
    _plot_quality_grade_provider_interaction,
    _quality_pair_grade_panel_data,
)
from dmw_experiments.studies.haiu_comparison.analysis.quality.grades import (
    build_quality_grade_analysis,
)
from dmw_experiments.studies.haiu_comparison.analysis.quality.errors import (
    build_quality_error_analysis,
)


def test_plot_workbooks_writes_timestamped_pdf_and_png_figures(
    tmp_path: Path,
) -> None:
    academic = _write_workbook(
        tmp_path / "academic/analysis/overview.xlsx",
        profile_name="academiccloud-qwen36",
        chat_provider="academiccloud",
        quantization="FP8",
        partial=True,
    )
    local = _write_workbook(
        tmp_path / "lmstudio/analysis/overview.xlsx",
        profile_name="lmstudio-qwen36-q6",
        chat_provider="lmstudio",
        quantization="Q6",
        partial=True,
    )

    output_dir = plot_workbooks(
        [academic, local],
        output_root=tmp_path,
        timestamp="20260729T120000CEST",
    )

    assert output_dir.name == "plots-20260729T120000CEST"
    assert sorted(path.name for path in output_dir.iterdir()) == [
        "image-captions.md",
        "outcomes.pdf",
        "outcomes.png",
        "outcomes.svg",
        "paired-absolute-metrics.pdf",
        "paired-absolute-metrics.png",
        "paired-absolute-metrics.svg",
        "plot_manifest.json",
    ]
    for figure_path in output_dir.glob("*.pdf"):
        assert figure_path.read_bytes().startswith(b"%PDF")
    for figure_path in output_dir.glob("*.png"):
        assert figure_path.read_bytes().startswith(b"\x89PNG\r\n\x1a\n")
    for figure_path in output_dir.glob("*.svg"):
        assert "<svg" in figure_path.read_text(encoding="utf-8")

    manifest = json.loads(
        (output_dir / "plot_manifest.json").read_text(encoding="utf-8")
    )
    assert manifest["status"] == (
        "PARTIAL DIAGNOSTIC EXPORT — not publication evidence"
    )
    assert [item["provider_profile"] for item in manifest["inputs"]] == [
        "academiccloud-qwen36",
        "lmstudio-qwen36-q6",
    ]
    assert all(
        not item["source"].startswith("/") for item in manifest["inputs"]
    )
    assert manifest["caption_file"] == "image-captions.md"
    captions = (output_dir / "image-captions.md").read_text(encoding="utf-8")
    assert "## `outcomes`" in captions
    assert "## `paired-absolute-metrics`" in captions


def test_plot_workbooks_supports_one_enabled_provider(tmp_path: Path) -> None:
    """AcademicCloud-only studies retain outcomes and paired metrics."""
    academic = _write_workbook(
        tmp_path / "academic/analysis/overview.xlsx",
        profile_name="academiccloud-qwen36",
        chat_provider="academiccloud",
        quantization="FP8",
        partial=True,
    )

    output_dir = plot_workbooks(
        [academic],
        output_root=tmp_path,
        timestamp="20260807T000000CEST",
    )

    assert (output_dir / "outcomes.svg").is_file()
    assert (output_dir / "paired-absolute-metrics.svg").is_file()
    manifest = json.loads(
        (output_dir / "plot_manifest.json").read_text(encoding="utf-8")
    )
    assert [item["provider_profile"] for item in manifest["inputs"]] == [
        "academiccloud-qwen36"
    ]


def test_plot_workbooks_adds_paired_historian_grades_when_supplied(
    tmp_path: Path,
) -> None:
    academic = _write_workbook(
        tmp_path / "academic/analysis/overview.xlsx",
        profile_name="academiccloud-qwen36",
        chat_provider="academiccloud",
        quantization="FP8",
        partial=True,
    )
    local = _write_workbook(
        tmp_path / "lmstudio/analysis/overview.xlsx",
        profile_name="lmstudio-qwen36-q6",
        chat_provider="lmstudio",
        quantization="Q6",
        partial=True,
    )
    review_workbook, reveal_key = _write_evaluated_quality_review(tmp_path)

    output_dir = plot_workbooks(
        [academic, local],
        output_root=tmp_path,
        timestamp="20260729T121000CEST",
        quality_review_workbook=review_workbook,
        quality_reveal_key=reveal_key,
    )

    assert (output_dir / "paired-quality-grades.png").is_file()
    assert (
        (output_dir / "paired-quality-grades.pdf")
        .read_bytes()
        .startswith(b"%PDF")
    )
    assert (output_dir / "quality-grade-analysis.xlsx").is_file()
    assert not (output_dir / "quality-grade-boundary-crossings.png").exists()
    assert (output_dir / "quality-grade-provider-interaction.png").is_file()
    captions = (output_dir / "image-captions.md").read_text(encoding="utf-8")
    assert "## `paired-quality-grades`" in captions
    assert "## `quality-grade-provider-interaction`" in captions
    manifest = json.loads(
        (output_dir / "plot_manifest.json").read_text(encoding="utf-8")
    )
    quality_manifest = manifest["historian_quality_grades"]
    assert quality_manifest["workbook"].endswith(
        "evaluated_quality_review.xlsx"
    )
    assert quality_manifest["reveal_key"].endswith(
        "evaluated_quality_review_reveal_key.json"
    )
    assert (
        quality_manifest["analysis_workbook"] == "quality-grade-analysis.xlsx"
    )
    assert quality_manifest["graded_observations"] == 6
    assert quality_manifest["complete_triplets"] == 2
    assert quality_manifest["providers"] == [
        "AcademicCloud FP8",
        "LM Studio Q6",
    ]
    assert quality_manifest["error_count_observations"] == 0
    assert quality_manifest["interpretation_count_observations"] == 0
    assert quality_manifest["assertion_count_observations"] == 0


def test_plot_workbooks_adds_false_assignment_error_profile_when_counted(
    tmp_path: Path,
) -> None:
    """Export the count figure and its audit tables only after review input."""
    academic = _write_workbook(
        tmp_path / "academic/analysis/overview.xlsx",
        profile_name="academiccloud-qwen36",
        chat_provider="academiccloud",
        quantization="FP8",
        partial=True,
    )
    local = _write_workbook(
        tmp_path / "lmstudio/analysis/overview.xlsx",
        profile_name="lmstudio-qwen36-q6",
        chat_provider="lmstudio",
        quantization="Q6",
        partial=True,
    )
    review_workbook, reveal_key = _write_evaluated_quality_review(
        tmp_path,
        include_error_counts=True,
    )

    output_dir = plot_workbooks(
        [academic, local],
        output_root=tmp_path,
        timestamp="20260803T140000CEST",
        quality_review_workbook=review_workbook,
        quality_reveal_key=reveal_key,
    )

    assert (output_dir / "false-assignment-error-profile.png").is_file()
    assert (
        (output_dir / "false-assignment-error-profile.pdf")
        .read_bytes()
        .startswith(b"%PDF")
    )
    audit_workbook = pd.ExcelFile(output_dir / "quality-grade-analysis.xlsx")
    assert {
        "13_Error_Count_Inputs",
        "14_Error_Count_Coverage",
        "15_Matched_Interpretation_Pairs",
        "16_Matched_Assertion_Pairs",
        "17_Pooled_Interp_Incidence",
        "18_Pooled_Assert_Incidence",
        "19_Interp_Pair_Differences",
        "20_Assert_Pair_Differences",
        "21_Pooled_Interp_Changes",
        "22_Pooled_Assert_Changes",
    }.issubset(audit_workbook.sheet_names)
    manifest = json.loads(
        (output_dir / "plot_manifest.json").read_text(encoding="utf-8")
    )
    quality_manifest = manifest["historian_quality_grades"]
    assert quality_manifest["error_count_observations"] == 6
    assert quality_manifest["interpretation_count_observations"] == 6
    assert quality_manifest["assertion_count_observations"] == 6
    assert quality_manifest["matched_interpretation_pairs"] == 4
    assert quality_manifest["matched_assertion_pairs"] == 4


def test_quality_pair_grade_panel_data_keeps_only_planned_pairs() -> None:
    trajectories = _quality_pair_grade_panel_data(
        pd.DataFrame(
            [
                (
                    "DMW full ontology vs DMW RAG",
                    "workflow_full_ontology",
                    "workflow_rag",
                    "AcademicCloud FP8",
                    "1001",
                    4,
                    3,
                ),
                (
                    "DMW RAG vs standalone Haiu RAG",
                    "workflow_rag",
                    "haiu_rag_ontologizer",
                    "AcademicCloud FP8",
                    "1001",
                    3,
                    2,
                ),
            ],
            columns=(
                "comparison",
                "first_condition",
                "second_condition",
                "provider_label",
                "regest_id",
                "first_grade",
                "second_grade",
            ),
        )
    )

    assert trajectories["comparison"].tolist() == [
        "DMW: full vs RAG",
        "DMW: full vs RAG",
        "DMW RAG vs standalone",
        "DMW RAG vs standalone",
    ]
    assert trajectories["plot_x"].between(-0.1, 3.1).all()


def test_false_assignment_incidence_uses_wilson_intervals() -> None:
    """Plot rubric-defined false assignments over four paired endpoints."""
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 5),
                ("AcademicCloud FP8", "1001", "workflow_rag", 1),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 1),
                ("AcademicCloud FP8", "1002", "workflow_rag", 4),
                ("AcademicCloud FP8", "1002", "haiu_rag_ontologizer", 1),
                ("LM Studio Q6", "1001", "workflow_full_ontology", 4),
                ("LM Studio Q6", "1001", "workflow_rag", 5),
                ("LM Studio Q6", "1001", "haiu_rag_ontologizer", 6),
                ("LM Studio Q6", "1002", "workflow_full_ontology", 3),
                ("LM Studio Q6", "1002", "workflow_rag", 2),
                ("LM Studio Q6", "1002", "haiu_rag_ontologizer", 1),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )
    figure, axis = plt.subplots()
    try:
        labels = _plot_false_assignment_incidence(
            axis,
            analysis=analysis,
            provider_order=["AcademicCloud FP8", "LM Studio Q6"],
            palette={
                "AcademicCloud FP8": PROVIDER_COLORS[0],
                "LM Studio Q6": PROVIDER_COLORS[1],
            },
        )

        assert axis.get_title() == "C. False-assignment incidence"
        assert (
            axis.get_ylabel() == "Share of analyses\nwith ≥1 false assignment"
        )
        assert [tick.get_text() for tick in axis.get_xticklabels()] == [
            "DMW vs DMW+HAIU",
            "DMW+HAIU vs HAIU",
        ]
        assert (
            len([patch for patch in axis.patches if patch.get_width() > 0]) == 8
        )
        assert len(labels) == 8
        assert {label.get_text() for label in labels} == {
            "0/1",
            "0/2",
            "1/1",
            "1/2",
        }
        assert all(label.get_zorder() > 3.0 for label in labels)
        assert {label.get_ha() for label in labels} == {"center"}
        assert len(axis.collections) == 8
        assert all(
            collection.get_zorder() == 3.0 for collection in axis.collections
        )
        assert all(
            collection.get_alpha() == 1.0 for collection in axis.collections
        )
        assert all(
            np.allclose(
                collection.get_color()[0],
                (0.302, 0.302, 0.302, 1.0),
                atol=0.0001,
            )
            for collection in axis.collections
        )
    finally:
        plt.close(figure)


def test_false_assignment_error_profile_pools_provider_local_direct_pairs() -> (
    None
):
    """Pool only pairs already matched within one provider and regest."""
    analysis = build_quality_error_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 0, 0),
                ("AcademicCloud FP8", "1001", "workflow_rag", 2, 1),
                (
                    "AcademicCloud FP8",
                    "1001",
                    "haiu_rag_ontologizer",
                    8,
                    "3+",
                ),
                ("AcademicCloud FP8", "1002", "workflow_full_ontology", 1, 1),
                ("AcademicCloud FP8", "1002", "workflow_rag", 0, 0),
                ("AcademicCloud FP8", "1002", "haiu_rag_ontologizer", 2, 1),
                ("LM Studio Q6", "1001", "workflow_full_ontology", 1, 1),
                ("LM Studio Q6", "1001", "workflow_rag", 0, 0),
                ("LM Studio Q6", "1001", "haiu_rag_ontologizer", 2, 1),
            ],
            columns=(
                "provider_label",
                "regest_id",
                "condition",
                "false_assertions",
                "false_interpretations",
            ),
        )
    )
    figure = _plot_false_assignment_error_profile(analysis)
    try:
        (
            interpretation_incidence,
            assertion_incidence,
            interpretation_changes,
            assertion_changes,
        ) = figure.axes[:4]
        assert interpretation_incidence.get_title() == (
            "A. Independent false\ninterpretations"
        )
        assert assertion_incidence.get_title() == "B. False atomic\nassertions"
        assert interpretation_changes.get_title() == (
            "C. Paired interpretation\nchanges"
        )
        assert assertion_changes.get_title() == (
            "D. Paired false-assertion\nchanges"
        )
        assert len(interpretation_incidence.patches) == 16
        assert len(assertion_incidence.patches) == 20
        assert len(interpretation_changes.patches) == 6
        assert len(assertion_changes.patches) == 6
        assert interpretation_changes.get_ylabel() == ("Share of matched pairs")
        assert any(
            text.get_text() == "n:" for text in interpretation_incidence.texts
        )
        assert {"1", "2"}.issubset(
            {text.get_text() for text in interpretation_incidence.texts}
        )
        assert any(
            text.get_text() == "n:" for text in assertion_incidence.texts
        )
        assert any(
            text.get_text() == "n:" for text in interpretation_changes.texts
        )
        assert any(text.get_text() == "n:" for text in assertion_changes.texts)
        assert to_rgba("#D73027") in {
            tuple(patch.get_facecolor())
            for patch in interpretation_changes.patches
        }
        assert to_rgba("#D73027") in {
            tuple(patch.get_facecolor()) for patch in assertion_changes.patches
        }
        assert np.allclose(
            figure.get_size_inches(),
            (8.8, 3.5),
        )
        assert figure._suptitle is None
        all_visible_labels = {
            text.get_text()
            for axis in figure.axes[:4]
            for text in [*axis.texts, *axis.get_xticklabels()]
        }
        assert "DMW vs\nDMW+HAIU" in all_visible_labels
        assert "DMW+HAIU\nvs HAIU" in all_visible_labels
        count_legend = figure.legends[0]
        assert [text.get_text() for text in count_legend.get_texts()] == [
            "0 errors",
            "1 error",
            "2 errors",
            "3/3+ errors",
            "4+ errors",
        ]
        assert figure.axes[5].get_position().y0 > (
            interpretation_changes.get_position().y1 + 0.10
        )
    finally:
        plt.close(figure)


def test_load_workbook_results_rejects_missing_required_sheet(
    tmp_path: Path,
) -> None:
    path = _write_workbook(
        tmp_path / "overview.xlsx",
        profile_name="academiccloud-qwen36",
        chat_provider="academiccloud",
        quantization="FP8",
        partial=False,
    )
    workbook = pd.ExcelFile(path)
    frames = {
        sheet_name: pd.read_excel(workbook, sheet_name=sheet_name)
        for sheet_name in workbook.sheet_names
        if sheet_name != "09_Token_Accounting"
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in frames.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    try:
        load_workbook_results(path)
    except ValueError as exc:
        assert "09_Token_Accounting" in str(exc)
    else:
        raise AssertionError("Expected missing-sheet validation to fail.")


def test_classify_failure_keeps_context_and_provider_errors_distinct() -> None:
    assert (
        _classify_failure(
            pd.Series({"success": True, "turtle_syntax_valid": False})
        )
        == "Invalid Turtle"
    )


def test_outcome_plot_uses_one_row_and_separate_two_row_issue_legend(
    tmp_path: Path,
) -> None:
    workbooks = [
        load_workbook_results(
            _write_workbook(
                tmp_path / "academic/analysis/overview.xlsx",
                profile_name="academiccloud-qwen36",
                chat_provider="academiccloud",
                quantization="FP8",
                partial=True,
            )
        ),
        load_workbook_results(
            _write_workbook(
                tmp_path / "lmstudio/analysis/overview.xlsx",
                profile_name="lmstudio-qwen36-q6",
                chat_provider="lmstudio",
                quantization="Q6",
                partial=True,
            )
        ),
    ]
    provider_order = [workbook.provider_label for workbook in workbooks]
    palette = {
        label: PROVIDER_COLORS[index]
        for index, label in enumerate(provider_order)
    }
    observations = pd.concat(
        [workbook.observations for workbook in workbooks],
        ignore_index=True,
    )
    observations["error_message"] = observations["error_message"].astype(
        "object"
    )
    observations.loc[0, "turtle_syntax_valid"] = False
    observations.loc[1, "success"] = False
    observations.loc[1, "output_truncated"] = True
    observations.loc[2, "success"] = False
    observations.loc[2, "error_message"] = "provider timed out"
    figure = _plot_outcomes(
        pd.concat(
            [workbook.results for workbook in workbooks], ignore_index=True
        ),
        pd.concat(
            [workbook.pairs for workbook in workbooks], ignore_index=True
        ),
        observations=observations,
        provider_order=provider_order,
        palette=palette,
        status_text=_status_text(workbooks),
    )
    try:
        plot_axes = [ax for ax in figure.axes if ax.get_title()]

        assert len(plot_axes) == 4
        assert len(figure.legends) == 2
        assert figure.legends[0]._ncols == len(provider_order)
        assert figure.legends[1]._ncols == 2
    finally:
        plt.close(figure)
    assert (
        _classify_failure(
            pd.Series({"output_truncated": True, "error_message": ""})
        )
        == "Context / length"
    )
    assert (
        _classify_failure(
            pd.Series({"failure_code": "model_context_window_exceeded"})
        )
        == "Context / length"
    )
    assert (
        _classify_failure(
            pd.Series({"output_truncated": False, "error_message": "timed out"})
        )
        == "Provider timeout"
    )
    assert (
        _classify_failure(
            pd.Series(
                {"output_truncated": False, "error_message": "Connection error"}
            )
        )
        == "Connection error"
    )
    assert (
        _classify_failure(
            pd.Series(
                {
                    "output_truncated": False,
                    "error_message": "DMW reused a prior ontology result; every "
                    "condition must generate a fresh ontology observation.",
                }
            )
        )
        == "Provenance rejection"
    )


def test_paired_trajectory_dots_align_with_line_endpoints(
    tmp_path: Path,
) -> None:
    """Ensure categorical point placement cannot drift from line positions."""
    workbooks = [
        load_workbook_results(
            _write_workbook(
                tmp_path / "academic/analysis/overview.xlsx",
                profile_name="academiccloud-qwen36",
                chat_provider="academiccloud",
                quantization="FP8",
                partial=True,
            )
        ),
        load_workbook_results(
            _write_workbook(
                tmp_path / "lmstudio/analysis/overview.xlsx",
                profile_name="lmstudio-qwen36-q6",
                chat_provider="lmstudio",
                quantization="Q6",
                partial=True,
            )
        ),
    ]
    provider_order = [workbook.provider_label for workbook in workbooks]
    palette = {
        label: PROVIDER_COLORS[index]
        for index, label in enumerate(provider_order)
    }
    figure = _plot_paired_absolute_metrics(
        pd.concat(
            [workbook.pairs for workbook in workbooks], ignore_index=True
        ),
        observations=pd.concat(
            [workbook.observations for workbook in workbooks],
            ignore_index=True,
        ),
        provider_order=provider_order,
        palette=palette,
        status_text=_status_text(workbooks),
    )
    try:
        plot_axes = [ax for ax in figure.axes if ax.get_title()]
        assert len(plot_axes) == 6
        y_scales = {ax.get_title(): ax.get_yscale() for ax in plot_axes}
        assert y_scales["Schema-reference reuse"] == "linear"
        assert y_scales["Duration"] == "log"
        for ax in plot_axes:
            assert [tick.get_text() for tick in ax.get_xticklabels()] == [
                "DMW vs DMW+HAIU",
                "DMW+HAIU vs HAIU",
            ]
            line_points = np.vstack(
                [
                    np.column_stack([line.get_xdata(), line.get_ydata()])
                    for line in ax.lines
                    if line.get_label().startswith("_child")
                ]
            )
            dot_points = np.vstack(
                [collection.get_offsets() for collection in ax.collections]
            )
            assert all(
                np.any(np.all(np.isclose(dot_points, point), axis=1))
                for point in line_points
            )
            assert any(
                not np.isclose(point[0], round(float(point[0])))
                for point in line_points
            )
            assert all(
                np.allclose(collection.get_facecolors()[:, 3], 0.0)
                for collection in ax.collections
                if collection.get_facecolors().size
            )
            for line in ax.lines:
                if not line.get_label().startswith("_child"):
                    continue
                assert line.get_alpha() == 0.3
                x_values = line.get_xdata()
                assert (x_values.max() < 1.5) or (x_values.min() > 1.5)
    finally:
        plt.close(figure)


def test_pairwise_grade_sample_size_annotations_are_opaque() -> None:
    """Keep Panel A denominators legible above the paired trajectories."""
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 4),
                ("AcademicCloud FP8", "1001", "workflow_rag", 3),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 2),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )
    figure, axis = plt.subplots()
    try:
        _plot_pairwise_quality_grade_trajectories(
            axis,
            analysis=analysis,
            provider_order=["AcademicCloud FP8"],
            palette={"AcademicCloud FP8": PROVIDER_COLORS[0]},
        )
        sample_size_texts = [
            text for text in axis.texts if text.get_text() == "n:"
        ]
        assert sample_size_texts
        assert all(text.get_alpha() == 1.0 for text in sample_size_texts)
    finally:
        plt.close(figure)


def test_grade_distribution_legend_explains_each_grade() -> None:
    """Render the six rubric descriptions above the paired endpoint bars."""
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 1),
                ("AcademicCloud FP8", "1001", "workflow_rag", 2),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 3),
                ("LM Studio Q6", "1002", "workflow_full_ontology", 4),
                ("LM Studio Q6", "1002", "workflow_rag", 5),
                ("LM Studio Q6", "1002", "haiu_rag_ontologizer", 6),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )
    figure = _plot_quality_grade_overview(
        analysis,
        status_text="Test export",
    )
    try:
        assert figure.axes[0].get_title() == "A. Grade distribution"
        assert figure.axes[1].get_title() == "B. Paired grade changes"
        assert figure.get_size_inches().tolist() == [6.0, 4.7]
        assert np.isclose(
            figure.axes[0].get_position().height * figure.get_figheight(),
            2.0,
            atol=0.001,
        )
        assert len(figure.axes[0].patches) == 24
        assert any(text.get_text() == "n:" for text in figure.axes[0].texts)
        legend = next(
            legend
            for legend in figure.legends
            if any(
                text.get_text().startswith("Grade 1")
                for text in legend.get_texts()
            )
        )
        assert legend._ncols == 1
        assert [text.get_text() for text in legend.texts] == [
            "Grade 1 — correct, essentially complete",
            "Grade 2 — correct, minor omissions",
            "Grade 3 — incomplete, factually safe",
            "Grade 4 — ≥1 error; local/formal, patchable",
            "Grade 5 — ≥1 error; plausible historical error",
            "Grade 6 — ≥1 error; gross source-reading failure",
        ]
    finally:
        plt.close(figure)


def test_paired_grade_changes_pool_only_direct_provider_regest_pairs() -> None:
    """Use three mutually exclusive directions and matched denominators."""
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 5),
                ("AcademicCloud FP8", "1001", "workflow_rag", 3),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 3),
                ("LM Studio Q6", "1001", "workflow_full_ontology", 2),
                ("LM Studio Q6", "1001", "workflow_rag", 2),
                ("LM Studio Q6", "1001", "haiu_rag_ontologizer", 4),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )
    figure = _plot_quality_grade_overview(analysis, status_text="Test export")
    try:
        change_axis = figure.axes[1]
        assert change_axis.get_title() == "B. Paired grade changes"
        assert len(change_axis.patches) == 6
        assert any(text.get_text() == "n:" for text in change_axis.texts)
        assert to_rgba("#1B9E77") in {
            tuple(patch.get_facecolor()) for patch in change_axis.patches
        }
        assert to_rgba("#D73027") in {
            tuple(patch.get_facecolor()) for patch in change_axis.patches
        }
    finally:
        plt.close(figure)


def test_provider_interaction_adds_grade_and_false_assignment_rows() -> None:
    """Show central trends and paired tests for both interaction outcomes."""
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 5),
                ("AcademicCloud FP8", "1001", "workflow_rag", 3),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 2),
                ("AcademicCloud FP8", "1002", "workflow_full_ontology", 4),
                ("AcademicCloud FP8", "1002", "workflow_rag", 4),
                ("AcademicCloud FP8", "1002", "haiu_rag_ontologizer", 5),
                ("LM Studio Q6", "1001", "workflow_full_ontology", 2),
                ("LM Studio Q6", "1001", "workflow_rag", 3),
                ("LM Studio Q6", "1001", "haiu_rag_ontologizer", 1),
                ("LM Studio Q6", "1002", "workflow_full_ontology", 6),
                ("LM Studio Q6", "1002", "workflow_rag", 4),
                ("LM Studio Q6", "1002", "haiu_rag_ontologizer", 4),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )
    figure = _plot_quality_grade_provider_interaction(
        analysis,
        provider_order=["AcademicCloud FP8", "LM Studio Q6"],
        palette={
            "AcademicCloud FP8": PROVIDER_COLORS[0],
            "LM Studio Q6": PROVIDER_COLORS[1],
        },
        status_text="Test export",
    )
    assert figure is not None
    try:
        plot_axes = figure.axes[:6]
        assert len(plot_axes) == 6
        assert [axis.get_title() for axis in plot_axes[:3]] == [
            "DMW full",
            "DMW + Haiu",
            "Standalone Haiu",
        ]
        assert all(
            any(
                line.get_linewidth() == pytest.approx(2.8)
                and line.get_alpha() == pytest.approx(0.45)
                for line in axis.lines
            )
            for axis in plot_axes
        )
        assert all("Exact paired" in axis.get_xlabel() for axis in plot_axes)
        assert all(
            "McNemar test" in axis.get_xlabel() for axis in plot_axes[3:]
        )
    finally:
        plt.close(figure)


def test_paired_absolute_metrics_keep_zero_reuse_values_on_linear_axis(
    tmp_path: Path,
) -> None:
    """Keep valid zero-reuse observations visible on the linear reuse axis."""
    workbook = load_workbook_results(
        _write_workbook(
            tmp_path / "academic/analysis/overview.xlsx",
            profile_name="academiccloud-qwen36",
            chat_provider="academiccloud",
            quantization="FP8",
            partial=True,
        )
    )
    workbook.pairs.loc[workbook.pairs["valid_pair"], "right_reuse_share"] = 0.0
    provider_order = [workbook.provider_label]
    palette = {workbook.provider_label: PROVIDER_COLORS[0]}
    figure = _plot_paired_absolute_metrics(
        workbook.pairs,
        observations=workbook.observations,
        provider_order=provider_order,
        palette=palette,
        status_text=_status_text([workbook]),
    )
    try:
        reuse_axis = next(
            ax
            for ax in figure.axes
            if ax.get_title() == "Schema-reference reuse"
        )
        assert reuse_axis.get_yscale() == "linear"
        assert any(
            np.any(np.isclose(collection.get_offsets()[:, 1], 0.0))
            for collection in reuse_axis.collections
            if collection.get_offsets().size
        )
    finally:
        plt.close(figure)


def _write_workbook(
    path: Path,
    *,
    profile_name: str,
    chat_provider: str,
    quantization: str,
    partial: bool,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    about = workbook.active
    about.title = "00_About"
    about.append(["DMW–Haiu Qwen 3.6 27B comparison", None])
    about.append(
        [
            "Status",
            ("PARTIAL DIAGNOSTIC EXPORT" if partial else "PUBLICATION EXPORT"),
        ]
    )
    about.append(
        [
            "Provider profile",
            json.dumps(
                {
                    "name": profile_name,
                    "chat_provider": chat_provider,
                    "quantization": quantization,
                }
            ),
        ]
    )
    workbook.save(path)

    results = pd.DataFrame(
        [
            {
                "condition": condition,
                "observations": 2,
                "valid_completed_count": index + 0,
                "valid_completed_rate": index / 2,
            }
            for index, condition in enumerate(
                (
                    "workflow_full_ontology",
                    "workflow_rag",
                    "haiu_rag_ontologizer",
                )
            )
        ]
    )
    observations = pd.DataFrame(
        [
            {
                "condition": condition,
                "regest_id": f"1100000{(index % 2) + 1}",
                "success": True,
                "turtle_syntax_valid": True,
                "output_truncated": False,
                "failure_code": None,
                "error_message": None,
                "duration_seconds": 60 + index,
                "prompt_tokens": 1000 * (index + 1),
                "prompt_tokens_complete": True,
                "output_tokens": 0 if index == 0 else 100 * index,
            }
            for index, condition in enumerate(
                (
                    "workflow_full_ontology",
                    "workflow_full_ontology",
                    "workflow_rag",
                    "workflow_rag",
                    "haiu_rag_ontologizer",
                    "haiu_rag_ontologizer",
                )
            )
        ]
    )
    context_pairs = _pair_frame(
        "workflow_full_ontology",
        "workflow_rag",
    )
    system_pairs = _pair_frame(
        "workflow_rag",
        "haiu_rag_ontologizer",
    )
    token_accounting = pd.DataFrame(
        {
            "condition": results["condition"],
            "prompt_tokens_median": [1000, 2000, 3000],
            "output_tokens_median": [0, 100, 200],
        }
    )

    with pd.ExcelWriter(
        path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        results.to_excel(writer, sheet_name="01_Results", index=False)
        context_pairs.to_excel(
            writer,
            sheet_name="02_DMW_Context_AB",
            index=False,
        )
        system_pairs.to_excel(
            writer,
            sheet_name="03_DMW_vs_Haiu_RAG",
            index=False,
        )
        observations.to_excel(
            writer,
            sheet_name="04_Observations",
            index=False,
        )
        token_accounting.to_excel(
            writer,
            sheet_name="09_Token_Accounting",
            index=False,
        )
    return path


def _write_evaluated_quality_review(
    tmp_path: Path,
    *,
    include_error_counts: bool = False,
) -> tuple[Path, Path]:
    workbook_path = tmp_path / "evaluated_quality_review.xlsx"
    reveal_key_path = tmp_path / "evaluated_quality_review_reveal_key.json"
    workbook = Workbook()
    academic = workbook.active
    academic.title = "AcademicCloud"
    headers = ["review_id", "regest_id", "grade_1_best_6_worst"]
    if include_error_counts:
        headers.extend(("false_assertions", "false_interpretations"))
    academic.append(headers)
    academic_rows = [
        ["R0001", "11010116", 4],
        ["R0002", None, 2],
        ["R0003", None, 1],
    ]
    if include_error_counts:
        for row, values in zip(
            academic_rows,
            ((3, 2), (0, 0), (8, "3+")),
            strict=True,
        ):
            row.extend(values)
    for row in academic_rows:
        academic.append(row)
    local = workbook.create_sheet("LM Studio")
    local.append(headers)
    local_rows = [
        ["R0001", "11010116", 3],
        ["R0002", None, 4],
        ["R0003", None, 2],
    ]
    if include_error_counts:
        for row, values in zip(
            local_rows,
            ((1, 1), (0, 0), (2, 1)),
            strict=True,
        ):
            row.extend(values)
    for row in local_rows:
        local.append(row)
    workbook.save(workbook_path)
    reveal_key_path.write_text(
        json.dumps(
            {
                "AcademicCloud": {
                    "R0001": {
                        "regest_id": "11010116",
                        "condition": "workflow_full_ontology",
                    },
                    "R0002": {
                        "regest_id": "11010116",
                        "condition": "workflow_rag",
                    },
                    "R0003": {
                        "regest_id": "11010116",
                        "condition": "haiu_rag_ontologizer",
                    },
                },
                "LM Studio": {
                    "R0001": {
                        "regest_id": "11010116",
                        "condition": "workflow_full_ontology",
                    },
                    "R0002": {
                        "regest_id": "11010116",
                        "condition": "workflow_rag",
                    },
                    "R0003": {
                        "regest_id": "11010116",
                        "condition": "haiu_rag_ontologizer",
                    },
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    return workbook_path, reveal_key_path


def _pair_frame(left_condition: str, right_condition: str) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "regest_id": "11000001",
                "left_condition": left_condition,
                "right_condition": right_condition,
                "valid_pair": True,
                "duration_delta_seconds": 10,
                "left_duration_seconds": 60,
                "right_duration_seconds": 70,
                "reuse_share_delta": 0.1,
                "left_reuse_share": 0.8,
                "right_reuse_share": 0.9,
                "left_novel_schema_declaration_count": 1,
                "right_novel_schema_declaration_count": 2,
                "left_triples": 10,
                "right_triples": 12,
            },
            {
                "regest_id": "11000002",
                "left_condition": left_condition,
                "right_condition": right_condition,
                "valid_pair": False,
                "duration_delta_seconds": 20,
                "left_duration_seconds": 60,
                "right_duration_seconds": 80,
                "reuse_share_delta": None,
                "left_reuse_share": None,
                "right_reuse_share": 0.7,
                "left_novel_schema_declaration_count": None,
                "right_novel_schema_declaration_count": 1,
                "left_triples": None,
                "right_triples": 15,
            },
        ]
    )
