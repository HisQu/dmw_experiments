from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from dmw_experiments.studies.haiu_comparison.analysis.quality.errors import (
    build_quality_error_analysis,
)
from dmw_experiments.studies.haiu_comparison.analysis.quality.grades import (
    build_quality_grade_analysis,
)
from dmw_experiments.studies.haiu_comparison.analysis.workbooks.quality import (
    export_quality_grade_analysis_workbook,
)


def test_quality_grade_analysis_keeps_matched_wins_and_ties_auditable() -> None:
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 5),
                ("AcademicCloud FP8", "1001", "workflow_rag", 3),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 1),
                ("AcademicCloud FP8", "1002", "workflow_full_ontology", 4),
                ("AcademicCloud FP8", "1002", "workflow_rag", 4),
                ("AcademicCloud FP8", "1002", "haiu_rag_ontologizer", 6),
                ("LM Studio Q6", "1001", "workflow_full_ontology", 2),
                ("LM Studio Q6", "1001", "workflow_rag", 3),
                ("LM Studio Q6", "1001", "haiu_rag_ontologizer", 1),
                ("LM Studio Q6", "1002", "workflow_full_ontology", 6),
                ("LM Studio Q6", "1002", "workflow_rag", 4),
                ("LM Studio Q6", "1002", "haiu_rag_ontologizer", 5),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )

    summary = analysis.condition_summary.set_index("condition")
    assert summary.loc["haiu_rag_ontologizer", "outright_best_count"] == 2
    assert summary.loc["workflow_rag", "outright_best_count"] == 1
    assert summary.loc["workflow_full_ontology", "outright_best_count"] == 0
    assert len(analysis.complete_triplets) == 4

    full_vs_rag = analysis.direct_duel_summary.iloc[0]
    assert full_vs_rag["comparison"] == "DMW full ontology vs DMW RAG"
    assert full_vs_rag["first_better_count"] == 1
    assert full_vs_rag["second_better_count"] == 2
    assert full_vs_rag["tie_count"] == 1
    assert analysis.direct_duel_summary["comparison"].tolist() == [
        "DMW full ontology vs DMW RAG",
        "DMW RAG vs standalone Haiu RAG",
    ]
    pooled_changes = analysis.pooled_grade_change_distribution.set_index(
        ["comparison", "change_direction"]
    )
    assert (
        pooled_changes.loc[
            ("DMW full ontology vs DMW RAG", "improved"), "count"
        ]
        == 2
    )
    assert (
        pooled_changes.loc[
            ("DMW full ontology vs DMW RAG", "unchanged"), "count"
        ]
        == 1
    )
    assert (
        pooled_changes.loc[
            ("DMW full ontology vs DMW RAG", "worsened"), "count"
        ]
        == 1
    )

    interaction = analysis.provider_interaction_summary
    assert set(interaction["shared_regesta"]) == {2}
    assert len(analysis.provider_interaction_pairs) == 6
    full_ontology_interaction = interaction.loc[
        interaction["condition"] == "workflow_full_ontology"
    ].iloc[0]
    assert full_ontology_interaction["exact_sign_test_non_tied_pairs"] == 2
    assert full_ontology_interaction["exact_sign_test_p_value"] == 1.0
    rag_interaction = interaction.loc[
        interaction["condition"] == "workflow_rag"
    ].iloc[0]
    assert rag_interaction["exact_sign_test_non_tied_pairs"] == 0
    assert pd.isna(rag_interaction["exact_sign_test_p_value"])
    false_assignment_interaction = (
        analysis.provider_false_assignment_interaction_summary.set_index(
            "condition"
        )
    )
    full_false_assignment_interaction = false_assignment_interaction.loc[
        "workflow_full_ontology"
    ]
    assert full_false_assignment_interaction["shared_regesta"] == 2
    assert (
        full_false_assignment_interaction[
            "first_provider_false_assignment_count"
        ]
        == 2
    )
    assert (
        full_false_assignment_interaction[
            "second_provider_false_assignment_count"
        ]
        == 1
    )
    assert (
        full_false_assignment_interaction["exact_mcnemar_discordant_pairs"] == 1
    )
    assert full_false_assignment_interaction["exact_mcnemar_p_value"] == 1.0
    assert analysis.friedman_summary.iloc[0]["complete_triplets"] == 4

    false_assignment = analysis.false_assignment_pair_summary.set_index(
        ["comparison", "provider_label", "condition"]
    )
    local_full = false_assignment.loc[
        (
            "DMW full ontology vs DMW RAG",
            "LM Studio Q6",
            "workflow_full_ontology",
        )
    ]
    assert local_full["models"] == 2
    assert local_full["false_assignment_count"] == 1
    assert local_full["false_assignment_share"] == 0.5
    assert local_full["wilson_95_lower_share"] == pytest.approx(
        0.0945, abs=0.0001
    )
    assert local_full["wilson_95_upper_share"] == pytest.approx(
        0.9055, abs=0.0001
    )


def test_false_assignment_incidence_uses_direct_pair_denominators() -> None:
    """Keep each condition's false-assignment rate within its direct pair."""
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 4),
                ("AcademicCloud FP8", "1001", "workflow_rag", 3),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 2),
                ("AcademicCloud FP8", "1002", "workflow_rag", 4),
                ("AcademicCloud FP8", "1002", "haiu_rag_ontologizer", 2),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )

    summary = analysis.false_assignment_pair_summary.set_index(
        ["comparison", "provider_label", "condition"]
    )
    full_vs_rag = "DMW full ontology vs DMW RAG"
    rag_vs_haiu = "DMW RAG vs standalone Haiu RAG"
    assert (
        summary.loc[
            (full_vs_rag, "AcademicCloud FP8", "workflow_full_ontology"),
            "models",
        ]
        == 1
    )
    assert (
        summary.loc[
            (full_vs_rag, "AcademicCloud FP8", "workflow_rag"), "models"
        ]
        == 1
    )
    assert (
        summary.loc[
            (rag_vs_haiu, "AcademicCloud FP8", "workflow_rag"), "models"
        ]
        == 2
    )
    assert (
        summary.loc[
            (rag_vs_haiu, "AcademicCloud FP8", "haiu_rag_ontologizer"), "models"
        ]
        == 2
    )
    paired_distribution = analysis.paired_grade_distribution.set_index(
        ["comparison", "pair_side", "grade"]
    )
    assert paired_distribution.loc[(full_vs_rag, "second", 3), "models"] == 1
    assert paired_distribution.loc[(rag_vs_haiu, "first", 3), "models"] == 2


def test_quality_grade_analysis_workbook_exposes_calculations(
    tmp_path: Path,
) -> None:
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 4),
                ("AcademicCloud FP8", "1001", "workflow_rag", 3),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 2),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )

    workbook_path = export_quality_grade_analysis_workbook(
        analysis=analysis,
        path=tmp_path / "quality-grade-analysis.xlsx",
        quality_review_workbook=tmp_path / "evaluated_review.xlsx",
        quality_reveal_key=tmp_path / "reveal_key.json",
    )

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    assert workbook.sheetnames == [
        "00_Read_Me",
        "01_Graded_Observations",
        "02_Condition_Summary",
        "03_Grade_Distribution",
        "03B_Paired_Grade_Distribution",
        "04_False_Assignment_Incidence",
        "05_Direct_Duels",
        "06_Duel_Pairs",
        "06B_Pooled_Grade_Changes",
        "07_Complete_Triplets",
        "08_Provider_Summary",
        "09_Provider_Interaction",
        "09B_Provider_Grade_Trends",
        "09C_False_Assign_Interaction",
        "10_Provider_Pairs",
        "11_Exploratory_Friedman",
        "12_Methods",
    ]
    methods = list(workbook["12_Methods"].values)
    assert methods[1][0] == "Condition mean and median"
    assert methods[3][0] == "Practical review categories"
    practical_categories = str(methods[3][1])
    assert "bounded local or formal error" in practical_categories
    assert (
        "plausible but materially false historical misinformation"
        in practical_categories
    )
    assert "gross source-reading failure" in practical_categories
    assert methods[4][0] == "False-assignment incidence and Wilson interval"
    assert "first_minus_second_grade" in str(methods[5][1])
    assert methods[6][0] == "Pooled paired grade changes"
    assert methods[7][0] == "Paired grade distribution"
    assert "same condition" in str(methods[7][1])
    assert methods[10][0] == "Provider false-assignment interaction"


def test_quality_grade_analysis_retains_pair_only_direct_duels() -> None:
    analysis = build_quality_grade_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 4),
                ("AcademicCloud FP8", "1001", "workflow_rag", 3),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 2),
                ("AcademicCloud FP8", "1002", "workflow_rag", 5),
                ("AcademicCloud FP8", "1002", "haiu_rag_ontologizer", 3),
            ],
            columns=("provider_label", "regest_id", "condition", "grade"),
        )
    )

    assert len(analysis.complete_triplets) == 1
    rag_vs_haiu = analysis.direct_duel_summary.iloc[1]
    assert rag_vs_haiu["comparison"] == "DMW RAG vs standalone Haiu RAG"
    assert rag_vs_haiu["complete_pairs"] == 2
    assert rag_vs_haiu["second_better_count"] == 2


def test_quality_error_analysis_keeps_optional_count_denominators_separate() -> (
    None
):
    """Do not convert an uncounted optional review field into a zero error."""
    analysis = build_quality_error_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 0, 0),
                ("AcademicCloud FP8", "1002", "workflow_rag", None, 2),
                (
                    "AcademicCloud FP8",
                    "1003",
                    "haiu_rag_ontologizer",
                    5,
                    "3+",
                ),
                ("LM Studio Q6", "1001", "workflow_full_ontology", 1, None),
            ],
            columns=(
                "provider_label",
                "regest_id",
                "condition",
                "false_assertions",
                "false_interpretations",
            ),
        )
    )

    coverage = analysis.study_overview.set_index("measure")["models"]
    assert coverage["reviews_with_any_error_count"] == 4
    assert coverage["reviews_with_interpretation_count"] == 3
    assert coverage["reviews_with_assertion_count"] == 3
    assert coverage["matched_interpretation_pairs"] == 0
    assert coverage["matched_assertion_pairs"] == 0
    assert analysis.pooled_interpretation_incidence.empty
    assert analysis.pooled_assertion_incidence.empty
    assert analysis.interpretation_pair_differences.empty
    assert analysis.assertion_pair_differences.empty
    assert analysis.pooled_interpretation_change_distribution.empty
    assert analysis.pooled_assertion_change_distribution.empty

    with pytest.raises(ValueError, match="0, 1, 2, or 3\\+"):
        build_quality_error_analysis(
            pd.DataFrame(
                [
                    (
                        "AcademicCloud FP8",
                        "1004",
                        "workflow_full_ontology",
                        0,
                        3,
                    )
                ],
                columns=(
                    "provider_label",
                    "regest_id",
                    "condition",
                    "false_assertions",
                    "false_interpretations",
                ),
            )
        )


def test_quality_error_analysis_pairs_each_count_measure_separately() -> None:
    """Compare only regesta counted at both endpoints of one direct duel."""
    analysis = build_quality_error_analysis(
        pd.DataFrame(
            [
                ("AcademicCloud FP8", "1001", "workflow_full_ontology", 1, 1),
                ("AcademicCloud FP8", "1001", "workflow_rag", 0, 0),
                ("AcademicCloud FP8", "1001", "haiu_rag_ontologizer", 3, 2),
                ("AcademicCloud FP8", "1002", "workflow_full_ontology", 0, 0),
                ("AcademicCloud FP8", "1002", "workflow_rag", 2, 1),
                (
                    "AcademicCloud FP8",
                    "1002",
                    "haiu_rag_ontologizer",
                    None,
                    "3+",
                ),
            ],
            columns=(
                "provider_label",
                "regest_id",
                "condition",
                "false_assertions",
                "false_interpretations",
            ),
        )
    )

    interpretation_pairs = analysis.matched_interpretation_pairs
    assertion_pairs = analysis.matched_assertion_pairs
    assert len(interpretation_pairs) == 4
    assert len(assertion_pairs) == 3
    assertion_incidence = analysis.pooled_assertion_incidence
    standalone_first = assertion_incidence.loc[
        (assertion_incidence["comparison"] == "DMW RAG vs standalone Haiu RAG")
        & (assertion_incidence["pair_side"] == "first")
    ]
    assert standalone_first["models"].eq(1).all()
    standalone_second = assertion_incidence.loc[
        (assertion_incidence["comparison"] == "DMW RAG vs standalone Haiu RAG")
        & (assertion_incidence["pair_side"] == "second")
    ].set_index("error_count_band")
    assert standalone_second.loc["3", "count"] == 1
    assert standalone_second.loc["4+", "count"] == 0
    interpretation_differences = analysis.interpretation_pair_differences
    assert interpretation_differences["error_count_difference"].tolist() == [
        -1,
        1,
        2,
        2,
    ]
    assert interpretation_differences["change_direction"].tolist() == [
        "improved",
        "worsened",
        "worsened",
        "worsened",
    ]
    interpretation_changes = (
        analysis.pooled_interpretation_change_distribution.set_index(
            ["comparison", "change_direction"]
        )
    )
    assert (
        interpretation_changes.loc[
            ("DMW full ontology vs DMW RAG", "improved"), "count"
        ]
        == 1
    )
    assert (
        interpretation_changes.loc[
            ("DMW RAG vs standalone Haiu RAG", "worsened"), "share"
        ]
        == 1.0
    )
