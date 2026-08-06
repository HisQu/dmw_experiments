import csv
import hashlib
import json
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from rdflib import Literal, URIRef
from rdflib.namespace import RDF, RDFS

from dmw_experiments.studies.haiu_comparison.export_results_workbook import (
    _SourceOrderedGraph,
    _historian_review_headers,
    _human_readable_main_entry,
    _human_readable_relationships,
    _shared_input_lineage,
    export_provider_historian_evaluation_sidecar,
    export_provider_historian_review_workbook,
    export_run,
)


CONDITIONS = (
    "workflow_full_ontology",
    "workflow_rag",
    "haiu_rag_ontologizer",
)


def _export_run(run_root: Path, **kwargs: object):
    """Export the AcademicCloud execution from a copied-run fixture."""
    return export_run(run_root / "raw-academiccloud", **kwargs)


def _export_provider_historian_review_workbook(
    academiccloud_root: Path,
    lmstudio_root: Path,
    **kwargs: object,
):
    """Export provider review data from two copied-run fixtures."""
    return export_provider_historian_review_workbook(
        academiccloud_root / "raw-academiccloud",
        lmstudio_root / "raw-academiccloud",
        **kwargs,
    )


def test_pair_review_headers_expose_source_regest_and_sublemma() -> None:
    packet = {
        "review_id": "R0001",
        "regest_id": "hsp-100-s01",
        "source_regest_id": "100",
        "source_sublemma_number": "1",
        "regest_text": "Header\nSublemma",
    }

    headers = _historian_review_headers([[packet]])

    assert headers[:5] == (
        "review_id",
        "regest_id",
        "source_regest_id",
        "source_sublemma_number",
        "regest_text",
    )


def test_pair_review_requires_lineage_shared_by_all_conditions() -> None:
    lineage = {
        "source_regest_id": "100",
        "source_sublemma_number": 1,
    }
    graph = _SourceOrderedGraph()

    shared = _shared_input_lineage(
        [
            ({"input_lineage": lineage}, graph),
            ({"input_lineage": dict(lineage)}, graph),
        ]
    )

    assert shared == {
        "source_regest_id": "100",
        "source_sublemma_number": "1",
    }


def test_historian_review_uses_generated_name_before_internal_iri() -> None:
    """Render standalone names in resources and triples without changing RDF."""
    namespace = "https://example.test/ontology/"
    event = URIRef(f"{namespace}i_10200370_event_indulgence")
    person = URIRef(f"{namespace}i_10200370_person_albertus_aep")
    event_type = URIRef(f"{namespace}Gnadenerweis")
    person_type = URIRef(f"{namespace}Person")
    name_property = URIRef(f"{namespace}hat_Namen")
    participant_property = URIRef(f"{namespace}hat_Beteiligten")
    graph = _SourceOrderedGraph()
    graph.add((event, RDF.type, event_type))
    graph.add((event, name_property, Literal("Jubiläumsablass")))
    graph.add((person, RDF.type, person_type))
    graph.add((person, name_property, Literal("Albertus aep. Magdeburg")))
    graph.add((participant_property, RDFS.label, Literal("hat Beteiligten")))
    graph.add((event, participant_property, person))

    person_display = _human_readable_main_entry(
        graph,
        person,
        bnode_labels={},
        reference_labels={},
    )
    relationships = _human_readable_relationships(graph, reference_labels={})

    assert person_display == "Albertus aep. Magdeburg"
    assert relationships[-1].subject == "Jubiläumsablass"
    assert relationships[-1].predicate == "hat Beteiligten"
    assert relationships[-1].object == "Albertus aep. Magdeburg"


def test_historian_review_replaces_unnamed_structural_resource_iris() -> None:
    """Render structural records from their declared type and source order."""
    namespace = "https://example.test/ontology/"
    first_record = URIRef(f"{namespace}i_10200370_regest_1")
    second_record = URIRef(f"{namespace}i_10200370_regest_2")
    record_type = URIRef(f"{namespace}Regest")
    related_to = URIRef(f"{namespace}bezieht_sich_auf")
    graph = _SourceOrderedGraph()
    graph.add((first_record, RDF.type, record_type))
    graph.add((second_record, RDF.type, record_type))
    graph.add((record_type, RDFS.label, Literal("Regest")))
    graph.add((related_to, RDFS.label, Literal("bezieht sich auf")))
    graph.add((first_record, related_to, second_record))

    first_display = _human_readable_main_entry(
        graph,
        first_record,
        bnode_labels={},
        reference_labels={},
    )
    relationships = _human_readable_relationships(graph, reference_labels={})

    assert first_display == "Regest 1"
    assert relationships[-1].subject == "Regest 1"
    assert relationships[-1].object == "Regest 2"


def test_exporter_writes_both_pair_comparisons_and_historian_packets(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)

    paths = _export_run(run_dir)

    workbook = load_workbook(paths.workbook, read_only=True)
    assert workbook.sheetnames == [
        "00_About",
        "01_Results",
        "02_DMW_Context_AB",
        "03_DMW_vs_Haiu_RAG",
        "04_Observations",
        "05_Timing_Context",
        "06_Schema_Declarations",
        "07_Novel_Declarations",
        "08_Exploratory_Cases",
        "09_Token_Accounting",
        "99_Definitions",
    ]
    pair_rows = list(workbook["03_DMW_vs_Haiu_RAG"].values)
    assert pair_rows[1][pair_rows[0].index("valid_pair")] is True
    assert (
        pair_rows[1][pair_rows[0].index("right_novel_schema_declaration_count")]
        == 1
    )
    result_rows = list(workbook["01_Results"].values)
    assert result_rows[1][result_rows[0].index("valid_completed_count")] == 1
    assert result_rows[1][result_rows[0].index("valid_completed_rate")] == 1

    historian_review = load_workbook(
        paths.historian_review_workbook, read_only=False
    )
    assert historian_review.sheetnames == ["Historian_Review"]
    historian_sidecar = load_workbook(
        paths.historian_review_evaluation_sidecar, read_only=False
    )
    assert historian_sidecar.sheetnames == [
        "Review_Guide",
        "Ontology_Classes",
        "Ontology_Properties",
        "Ontology_Individuals",
    ]
    guide_rows = list(historian_sidecar["Review_Guide"].values)
    guide_text = str(guide_rows)
    assert "workflow_full_ontology" in guide_text
    assert "workflow_rag" in guide_text
    assert "haiu_rag_ontologizer" in guide_text
    assert (
        "1 complete regest triplets; 0 additional planned two-condition pairs; "
        "3 condition-masked review rows." in guide_text
    )
    assert "Completed clean-run analysis export" in guide_text
    assert "Comparison emphasis" in guide_text
    assert "Grading Scheme for Regest Models" in guide_text
    assert "cannot receive a grade better than 4" in guide_text
    assert "plausible but materially false historical proposition" in guide_text
    assert "Gross source-reading failure" in guide_text
    assert "Grade 4 versus grade 5" in guide_text
    assert "Neither count mechanically determines the grade" in guide_text
    assert "false_assertions" in guide_text
    assert "false_interpretations" in guide_text
    assert "Substantive historical assertions" in guide_text
    assert "Frozen ontology reference" in guide_text
    assert "Workbook layout" in guide_text
    assert (
        "Retrieval-condition rows received only per-regest subsets"
        in guide_text
    )
    assert (
        "Bold underlined text occurs only in the current condition"
        in guide_text
    )
    assert "academiccloud-qwen36" not in guide_text
    assert "smoke" not in guide_text
    class_rows = list(historian_sidecar["Ontology_Classes"].values)
    assert class_rows[0] == (
        "Class",
        "Definition",
        "Parent class(es)",
        "RG expression(s)",
        "RG example",
    )
    assert class_rows[1:] == [
        ("Reference", "Reference definition", None, None, None),
        (
            "Reference Second",
            "Second definition",
            "Reference",
            "sec.",
            "Example regest",
        ),
    ]
    property_rows = list(historian_sidecar["Ontology_Properties"].values)
    assert property_rows[0] == (
        "Property type",
        "Property",
        "Definition",
        "Domain",
        "Range",
        "RG expression(s)",
    )
    assert property_rows[1:] == [
        (
            "Object property",
            "Relates to",
            None,
            "Reference",
            "Reference Second",
            None,
        ),
        ("Annotation property", "RG expression", None, None, None, None),
    ]
    individual_rows = list(historian_sidecar["Ontology_Individuals"].values)
    assert individual_rows == [
        ("Individual", "Type(s)", "Definition", "RG expression(s)"),
        ("Controlled value", "Reference Second", None, "ctrl."),
    ]
    historian_sheet = historian_review["Historian_Review"]
    historian_rows = list(historian_sheet.values)
    historian_headers = historian_rows[0]
    assert historian_headers == (
        "review_id",
        "regest_id",
        "regest_text",
        "grade_1_best_6_worst",
        "historian_verdict_and_notes",
        "false_assertions",
        "false_interpretations",
        "classes",
        "object_properties",
        "individuals",
        "relationships",
        "datatype_properties",
        "annotation_properties",
        "rdf_properties",
        "other_resources",
    )
    assert "workflow_rag" not in str(historian_rows[1:])
    historian_key = json.loads(
        paths.historian_review_reveal_key.read_text(encoding="utf-8")
    )
    review_id = next(
        value
        for value, packet in historian_key.items()
        if packet["condition"] == "workflow_rag"
    )
    historian_packet = next(
        row for row in historian_rows[1:] if row[0] == review_id
    )
    packet = dict(zip(historian_headers, historian_packet, strict=True))
    assert (
        historian_rows[1][2] == "Header text\nFirst subentry\nSecond subentry"
    )
    assert [row[1] for row in historian_rows[1:]] == ["11010116", None, None]
    assert [row[2] for row in historian_rows[1:]] == [
        "Header text\nFirst subentry\nSecond subentry",
        None,
        None,
    ]
    assert len({str(row[0]) for row in historian_rows[1:]}) == 3
    assert packet["classes"] == "Reference Second — First Class — Reference"
    assert packet["object_properties"] == "Relates to"
    assert packet["datatype_properties"] == "Has code"
    assert packet["annotation_properties"] == "Review note"
    assert packet["rdf_properties"] == "Generic property"
    assert packet["individuals"] == "Source instance — Target instance"
    assert packet["relationships"] == (
        "• Source instance — Relates to → Target instance\n"
        "• Target instance — Relates to → Source instance"
    )
    assert packet["grade_1_best_6_worst"] is None
    assert packet["historian_verdict_and_notes"] is None
    assert packet["false_assertions"] is None
    assert packet["false_interpretations"] is None
    assert historian_sheet.column_dimensions["C"].width == pytest.approx(
        40.7109375
    )
    assert historian_sheet.column_dimensions["D"].width <= 13
    assert historian_sheet.column_dimensions["E"].width == pytest.approx(
        32.28515625
    )
    assert historian_sheet.column_dimensions["F"].width <= 16
    assert historian_sheet.column_dimensions["G"].width <= 19
    assert historian_sheet.column_dimensions["K"].width >= 90
    assert historian_sheet["C2"].alignment.wrap_text is True
    assert historian_sheet["C2"].alignment.vertical == "top"
    assert historian_sheet.freeze_panes == "H2"
    assert historian_sheet.data_validations.count == 3
    assert sorted(
        str(cell_range) for cell_range in historian_sheet.merged_cells.ranges
    ) == ["B2:B4", "C2:C4"]
    assert not historian_sheet.tables
    with ZipFile(paths.historian_review_workbook) as archive:
        shared_strings = archive.read("xl/sharedStrings.xml").decode("utf-8")
    assert "<u/>" in shared_strings
    assert "<b/>" in shared_strings
    assert 'rgb="FF1F4E78"' in shared_strings
    assert 'rgb="FFC65911"' in shared_strings
    assert 'rgb="FF7030A0"' in shared_strings

    analysis = json.loads(paths.manifest.read_text(encoding="utf-8"))
    assert analysis["pair_denominators"] == {
        "dmw_full_vs_rag": 1,
        "workflow_rag_vs_haiu_rag": 1,
    }
    assert analysis["exporter_sha256"]
    assert analysis["audit_csv_enabled"] is False
    assert analysis["metric_definitions"]
    assert paths.workbook.name == "overview.xlsx"
    assert (
        paths.historian_review_workbook.name
        == "masked_historian_quality_review.xlsx"
    )
    assert (
        paths.historian_review_evaluation_sidecar.name
        == "masked_historian_quality_review_evaluation_sidecar.xlsx"
    )
    assert (
        paths.historian_review_reveal_key.name
        == "historian_quality_review_reveal_key.json"
    )
    assert paths.manifest.name == "analysis_manifest.json"
    assert paths.readme.name == "README.md"
    assert sorted(path.name for path in paths.workbook.parent.iterdir()) == [
        "README.md",
        "analysis_manifest.json",
        "historian_quality_review_reveal_key.json",
        "masked_historian_quality_review.xlsx",
        "masked_historian_quality_review_evaluation_sidecar.xlsx",
        "overview.xlsx",
    ]
    assert not list(paths.workbook.parent.glob("*.csv"))
    assert not (paths.workbook.parent / "audit_csv").exists()
    assert "Start with `overview.xlsx`" in paths.readme.read_text(
        encoding="utf-8"
    )
    assert "evaluation_sidecar.xlsx" in paths.readme.read_text(encoding="utf-8")


def test_historian_review_uses_direct_turtle_parsing_for_triplet_selection(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    raw_path = run_dir / "raw-academiccloud/result-workflow_rag/11010116.json"
    raw = json.loads(raw_path.read_text(encoding="utf-8"))
    raw["turtle_syntax_valid"] = False
    _write_json(raw_path, raw)

    paths = _export_run(run_dir)

    workbook = load_workbook(paths.historian_review_workbook, read_only=True)
    rows = list(workbook["Historian_Review"].values)
    assert len(rows[1:]) == 3


def test_exporter_reconciles_full_turtle_generation_input_context(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    raw_path = (
        run_dir
        / "raw-academiccloud/result-workflow_full_ontology/11010116.json"
    )
    raw = json.loads(raw_path.read_text(encoding="utf-8"))
    raw["prompt_tokens"] = 12
    raw["prompt_tokens_source"] = "estimated_stage_bundles_only"
    raw["generation_budget"] = {
        "stage1": {"provider_prompt_tokens": 100},
        "stage2": {"provider_prompt_tokens": 150},
    }
    _write_json(raw_path, raw)

    paths = _export_run(run_dir)

    workbook = load_workbook(paths.workbook, read_only=True)
    observations = list(workbook["04_Observations"].values)
    headers = observations[0]
    full_row = next(
        row
        for row in observations[1:]
        if row[headers.index("condition")] == "workflow_full_ontology"
    )
    assert full_row[headers.index("prompt_tokens")] == 150
    assert full_row[headers.index("prompt_tokens_source")] == "provider"


def test_historian_review_sorts_contiguous_three_row_regest_groups(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    _add_complete_regest(run_dir, "11010115")

    paths = _export_run(run_dir)

    workbook = load_workbook(paths.historian_review_workbook, read_only=False)
    review = workbook["Historian_Review"]
    rows = list(review.values)
    assert [row[1] for row in rows[1:]] == [
        "11010115",
        None,
        None,
        "11010116",
        None,
        None,
    ]
    assert sorted(
        str(cell_range) for cell_range in review.merged_cells.ranges
    ) == [
        "B2:B4",
        "B5:B7",
        "C2:C4",
        "C5:C7",
    ]
    reveal_key = json.loads(
        paths.historian_review_reveal_key.read_text(encoding="utf-8")
    )
    first_group_ids = [row[0] for row in rows[1:4]]
    second_group_ids = [row[0] for row in rows[4:7]]
    assert {
        reveal_key[review_id]["regest_id"] for review_id in first_group_ids
    } == {"11010115"}
    assert {
        reveal_key[review_id]["regest_id"] for review_id in second_group_ids
    } == {"11010116"}


def test_historian_review_adds_planned_pair_without_complete_triplet(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    _add_complete_regest(run_dir, "11010115")
    _write(
        run_dir / "raw-academiccloud/result-haiu_rag_ontologizer/11010115.ttl",
        "This is not Turtle.",
    )

    paths = _export_run(run_dir)

    sidecar = load_workbook(
        paths.historian_review_evaluation_sidecar, read_only=False
    )
    guide_text = str(list(sidecar["Review_Guide"].values))
    assert (
        "1 complete regest triplets; 1 additional planned two-condition pairs; "
        "5 condition-masked review rows." in guide_text
    )
    workbook = load_workbook(paths.historian_review_workbook, read_only=False)
    review = workbook["Historian_Review"]
    rows = list(review.values)
    assert [row[1] for row in rows[1:]] == [
        "11010115",
        None,
        "11010116",
        None,
        None,
    ]
    assert sorted(
        str(cell_range) for cell_range in review.merged_cells.ranges
    ) == ["B2:B3", "B4:B6", "C2:C3", "C4:C6"]
    reveal_key = json.loads(
        paths.historian_review_reveal_key.read_text(encoding="utf-8")
    )
    pair_review_ids = [row[0] for row in rows[1:3]]
    assert {
        reveal_key[review_id]["condition"] for review_id in pair_review_ids
    } == {"workflow_full_ontology", "workflow_rag"}


def test_historian_review_adds_rag_haiu_pair_without_full_turtle(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    _write(
        run_dir
        / "raw-academiccloud/result-workflow_full_ontology/11010116.ttl",
        "This is not Turtle.",
    )

    paths = _export_run(run_dir)

    workbook = load_workbook(paths.historian_review_workbook, read_only=False)
    rows = list(workbook["Historian_Review"].values)
    assert [row[1] for row in rows[1:]] == ["11010116", None]
    reveal_key = json.loads(
        paths.historian_review_reveal_key.read_text(encoding="utf-8")
    )
    assert {reveal_key[row[0]]["condition"] for row in rows[1:]} == {
        "workflow_rag",
        "haiu_rag_ontologizer",
    }


def test_historian_review_excludes_regest_without_a_parseable_triplet(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)
    _write(
        run_dir / "raw-academiccloud/result-workflow_rag/11010116.ttl",
        "This is not Turtle.",
    )

    paths = _export_run(run_dir)

    workbook = load_workbook(paths.historian_review_workbook, read_only=True)
    rows = list(workbook["Historian_Review"].values)
    assert rows[1][0] == (
        "No complete triplets or directly parseable planned two-condition pairs "
        "are available."
    )
    assert (
        json.loads(
            paths.historian_review_reveal_key.read_text(encoding="utf-8")
        )
        == {}
    )


def test_provider_historian_review_keeps_provider_sheets_and_exports_separate(
    tmp_path: Path,
) -> None:
    academiccloud_run = _complete_run(tmp_path / "academiccloud")
    lmstudio_run = _complete_run(tmp_path / "lmstudio")
    academiccloud_export = _export_run(academiccloud_run)
    lmstudio_export = _export_run(lmstudio_run)
    individual_hashes = {
        path: _sha256(path.read_text(encoding="utf-8"))
        for path in (
            academiccloud_export.historian_review_reveal_key,
            lmstudio_export.historian_review_reveal_key,
        )
    }

    paths = _export_provider_historian_review_workbook(
        academiccloud_run,
        lmstudio_run,
        workbook_path=tmp_path / "historian_quality_review_providers.xlsx",
    )

    workbook = load_workbook(paths.workbook, read_only=False)
    assert workbook.sheetnames == ["AcademicCloud", "LM Studio"]
    sidecar = load_workbook(paths.evaluation_sidecar, read_only=False)
    assert sidecar.sheetnames == [
        "Review_Guide",
        "Ontology_Classes",
        "Ontology_Properties",
        "Ontology_Individuals",
    ]
    guide_text = str(list(sidecar["Review_Guide"].values))
    assert "Provider layout" in guide_text
    assert "AcademicCloud review sample" in guide_text
    assert "LM Studio review sample" in guide_text
    assert "Grading Scheme for Regest Models" in guide_text
    assert "A missing assignment is an omission" in guide_text
    assert "false_assertions" in guide_text
    assert "false_interpretations" in guide_text
    assert "Frozen ontology reference" in guide_text
    assert "Workbook layout" in guide_text
    assert (
        "Retrieval-condition rows received only per-regest subsets"
        in guide_text
    )
    for sheet_name in ("AcademicCloud", "LM Studio"):
        review = workbook[sheet_name]
        assert review.freeze_panes == "H2"
        assert list(review.values)[0] == (
            "review_id",
            "regest_id",
            "regest_text",
            "grade_1_best_6_worst",
            "historian_verdict_and_notes",
            "false_assertions",
            "false_interpretations",
            "classes",
            "object_properties",
            "individuals",
            "relationships",
            "datatype_properties",
            "annotation_properties",
            "rdf_properties",
            "other_resources",
        )
    reveal_key = json.loads(paths.reveal_key.read_text(encoding="utf-8"))
    assert set(reveal_key) == {"AcademicCloud", "LM Studio"}
    assert all(
        packet["condition"] in CONDITIONS
        for provider_key in reveal_key.values()
        for packet in provider_key.values()
    )
    assert {
        path: _sha256(path.read_text(encoding="utf-8"))
        for path in individual_hashes
    } == individual_hashes
    assert (
        paths.evaluation_sidecar.name
        == "historian_quality_review_providers_evaluation_sidecar.xlsx"
    )

    with pytest.raises(ValueError, match="already exists"):
        _export_provider_historian_review_workbook(
            academiccloud_run,
            lmstudio_run,
            workbook_path=paths.workbook,
        )
    _export_provider_historian_review_workbook(
        academiccloud_run,
        lmstudio_run,
        workbook_path=paths.workbook,
        overwrite=True,
    )


def test_provider_evaluation_sidecar_never_replaces_manual_review_rows(
    tmp_path: Path,
) -> None:
    """Refresh guide material without touching its separately reviewed input."""
    academiccloud_run = _complete_run(tmp_path / "academiccloud")
    lmstudio_run = _complete_run(tmp_path / "lmstudio")
    paths = _export_provider_historian_review_workbook(
        academiccloud_run,
        lmstudio_run,
        workbook_path=tmp_path / "historian_quality_review_providers.xlsx",
    )
    original_review_hash = hashlib.sha256(
        paths.workbook.read_bytes()
    ).hexdigest()
    original_key_hash = _sha256(paths.reveal_key.read_text(encoding="utf-8"))

    sidecar = export_provider_historian_evaluation_sidecar(
        academiccloud_run / "raw-academiccloud",
        lmstudio_run / "raw-academiccloud",
        review_workbook_path=paths.workbook,
        overwrite=True,
    )

    assert sidecar == paths.evaluation_sidecar
    assert (
        hashlib.sha256(paths.workbook.read_bytes()).hexdigest()
        == original_review_hash
    )
    assert (
        _sha256(paths.reveal_key.read_text(encoding="utf-8"))
        == original_key_hash
    )
    assert load_workbook(sidecar, read_only=True).sheetnames == [
        "Review_Guide",
        "Ontology_Classes",
        "Ontology_Properties",
        "Ontology_Individuals",
    ]


def test_provider_historian_review_rejects_mixed_frozen_ontologies(
    tmp_path: Path,
) -> None:
    """Reject a shared vocabulary sheet when provider snapshots differ."""
    academiccloud_run = _complete_run(tmp_path / "academiccloud")
    lmstudio_run = _complete_run(tmp_path / "lmstudio")
    reference_path = (
        lmstudio_run
        / "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/reference_ontology.ttl"
    )
    changed_reference = reference_path.read_text(encoding="utf-8").replace(
        "Reference definition", "Changed reference definition"
    )
    _write(reference_path, changed_reference)
    provenance_path = (
        lmstudio_run
        / "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/provenance_manifest.json"
    )
    provenance = json.loads(provenance_path.read_text(encoding="utf-8"))
    provenance["inputs"]["reference_ontology"]["sha256"] = _sha256(
        changed_reference
    )
    _write_json(provenance_path, provenance)

    with pytest.raises(
        ValueError, match="different frozen reference ontologies"
    ):
        _export_provider_historian_review_workbook(
            academiccloud_run,
            lmstudio_run,
            workbook_path=tmp_path / "historian_quality_review_providers.xlsx",
        )


def test_exporter_writes_opt_in_audit_csv_and_cleans_stale_audit(
    tmp_path: Path,
) -> None:
    run_dir = _complete_run(tmp_path)

    paths = _export_run(run_dir, audit_csv=True)

    audit_dir = paths.workbook.parent / "audit_csv"
    assert sorted(path.name for path in audit_dir.iterdir()) == [
        "observations.csv",
        "pairs.csv",
        "schema_declarations.csv",
    ]
    with (audit_dir / "pairs.csv").open(encoding="utf-8", newline="") as handle:
        comparisons = {row["comparison"] for row in csv.DictReader(handle)}
    assert comparisons == {
        "DMW full ontology vs DMW RAG",
        "DMW RAG vs standalone Haiu RAG",
    }

    with pytest.raises(ValueError, match="not empty"):
        _export_run(run_dir)
    stale_review = paths.workbook.parent / "masked_case_review.xlsx"
    stale_key = paths.workbook.parent / "reveal_key.json"
    stale_review.write_text("obsolete", encoding="utf-8")
    stale_key.write_text("obsolete", encoding="utf-8")
    _export_run(run_dir, overwrite=True)
    assert not audit_dir.exists()
    assert not stale_review.exists()
    assert not stale_key.exists()


def test_exporter_rejects_missing_retrieval_sidecar(tmp_path: Path) -> None:
    run_dir = _complete_run(tmp_path)
    (
        run_dir
        / "raw-academiccloud/intermediates-workflow_rag/11010116.retrieved.yaml"
    ).unlink()

    with pytest.raises(ValueError, match="missing retrieval .yaml"):
        _export_run(run_dir)


def test_exporter_ignores_retry_pending_raw_checkpoint(tmp_path: Path) -> None:
    run_dir = _complete_run(tmp_path)
    condition = "haiu_rag_ontologizer"
    regest_id = "11010116"
    raw_path = (
        run_dir / f"raw-academiccloud/result-{condition}/{regest_id}.json"
    )
    raw = json.loads(raw_path.read_text(encoding="utf-8"))
    raw.update(
        {
            "success": False,
            "turtle_syntax_valid": None,
            "error_message": "Provider call is being retried.",
        }
    )
    _write_json(raw_path, raw)
    _write_json(
        run_dir
        / f"raw-academiccloud/intermediates-{condition}/{regest_id}.attempt.json",
        {
            "condition": condition,
            "regest_id": regest_id,
            "status": "retry_pending",
            "attempt": 1,
            "success": False,
        },
    )

    paths = _export_run(run_dir, allow_partial=True)

    analysis = json.loads(paths.manifest.read_text(encoding="utf-8"))
    assert analysis["row_count"] == 2
    assert (
        f"raw-academiccloud/result-{condition}/{regest_id}.json"
        not in analysis["source_raw_sha256"]
    )


def _complete_run(tmp_path: Path) -> Path:
    run_dir = tmp_path / "run"
    _write(run_dir / "run.toml", 'run_id = "test"\n')
    reference = (
        "@prefix : <https://example.org/> .\n"
        "@prefix owl: <http://www.w3.org/2002/07/owl#> .\n"
        "@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .\n"
        "@prefix skos: <http://www.w3.org/2004/02/skos/core#> .\n"
        ":stringInRG a owl:AnnotationProperty ; "
        'rdfs:label "RG expression"@en .\n'
        ':Reference a owl:Class ; rdfs:label "Reference"@de ; '
        'skos:definition "Reference definition"@en .\n'
        ':Second a owl:Class ; rdfs:label "Reference Second"@de ; '
        ':stringInRG "sec."@la ; rdfs:subClassOf :Reference ; '
        'skos:definition "Second definition"@en ; '
        'skos:example "Example regest"@la .\n'
        ':relates a owl:ObjectProperty ; rdfs:label "Relates to"@en ; '
        "rdfs:domain :Reference ; rdfs:range :Second .\n"
        ":controlled a owl:NamedIndividual, :Second ; "
        'rdfs:label "Controlled value"@en ; :stringInRG "ctrl."@la .\n'
    )
    _write(
        run_dir
        / "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/reference_ontology.ttl",
        reference,
    )
    regest = {
        "schema_version": 1,
        "source": "preflight_frozen_raw_regest_snapshot",
        "regest_id": "11010116",
        "header": "Header text",
        "subentries": ["First subentry", "Second subentry"],
        "content_sha256": "ignored-by-export-fixture",
    }
    regest_path = (
        run_dir
        / "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/raw_regests/11010116.json"
    )
    _write_json(regest_path, regest)
    raw_regests_manifest = {
        "schema_version": 1,
        "source": "preflight_frozen_raw_regest_snapshot",
        "records": {
            "11010116": {
                "path": "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/raw_regests/11010116.json",
                "sha256": _sha256(regest_path.read_text(encoding="utf-8")),
            }
        },
    }
    raw_regests_manifest_path = (
        run_dir
        / "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/raw_regests_manifest.json"
    )
    _write_json(raw_regests_manifest_path, raw_regests_manifest)
    _write_json(
        run_dir
        / "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/provenance_manifest.json",
        {
            "schema_version": 1,
            "inputs": {
                "reference_ontology": {
                    "path": "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/reference_ontology.ttl",
                    "sha256": _sha256(reference),
                }
            },
        },
    )
    _write_json(
        run_dir / "environment/academiccloud-run-manifest.json",
        {
            "run_id": "smoke",
            "regest_ids": ["11010116"],
            "conditions": list(CONDITIONS),
            "provider_profile": {"name": "academiccloud-qwen36"},
            "raw_regest_snapshot": {
                "path": "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/raw_regests_manifest.json",
                "sha256": _sha256(
                    raw_regests_manifest_path.read_text(encoding="utf-8")
                ),
                "count": 1,
            },
        },
    )
    for condition in CONDITIONS:
        turtle = _generated_turtle(condition)
        payload = {
            "condition": condition,
            "regest_id": "11010116",
            "success": True,
            "raw_ttl_capture_complete": True,
            "turtle_syntax_valid": True,
            "turtle_triple_count": 2,
            "duration_seconds": 1.5,
            "stage2_output_reduced": False,
            "prompt_tokens": 12,
            "prompt_tokens_complete": True,
            "condition_order": list(CONDITIONS),
            "condition_order_position": CONDITIONS.index(condition),
        }
        payload["raw_ttl_output"] = turtle
        _write_json(
            run_dir / f"raw-academiccloud/result-{condition}/11010116.json",
            payload,
        )
        _write(
            run_dir / f"raw-academiccloud/result-{condition}/11010116.yaml",
            "{}\n",
        )
        _write(
            run_dir / f"raw-academiccloud/result-{condition}/11010116.ttl",
            turtle,
        )
        if condition in {"workflow_rag", "haiu_rag_ontologizer"}:
            _write(
                run_dir / f"raw-academiccloud/intermediates-{condition}/"
                "11010116.retrieved.ttl",
                turtle,
            )
            _write(
                run_dir / f"raw-academiccloud/intermediates-{condition}/"
                "11010116.retrieved.yaml",
                "snapshot_fidelity: native_full_graph\n",
            )
    return run_dir


def _generated_turtle(condition: str) -> str:
    if condition == "workflow_rag":
        return "\n".join(
            (
                "@prefix : <https://example.org/> .",
                "@prefix owl: <http://www.w3.org/2002/07/owl#> .",
                "@prefix rdf: <http://www.w3.org/1999/02/22-rdf-syntax-ns#> .",
                "@prefix rdfs: <http://www.w3.org/2000/01/rdf-schema#> .",
                ':Second a owl:Class ; rdfs:label "Second Class" .',
                ':First a owl:Class ; rdfs:label "First Class" .',
                ":Reference a owl:Class .",
                ':relates a owl:ObjectProperty ; rdfs:label "Relates to" .',
                ':hasCode a owl:DatatypeProperty ; rdfs:label "Has code" .',
                ':note a owl:AnnotationProperty ; rdfs:label "Review note" .',
                ':generic a rdf:Property ; rdfs:label "Generic property" .',
                ':instance a :Second ; rdfs:label "Source instance" .',
                ':target a :First ; rdfs:label "Target instance" .',
                ":instance :relates :target .",
                ":target :relates :instance .",
            )
        )
    declaration = (
        ":Reference a owl:Class ."
        if condition != "haiu_rag_ontologizer"
        else ":NewConcept a owl:Class ."
    )
    return "\n".join(
        (
            "@prefix : <https://example.org/> .",
            "@prefix owl: <http://www.w3.org/2002/07/owl#> .",
            declaration,
        )
    )


def _add_complete_regest(run_dir: Path, regest_id: str) -> None:
    source_regest_id = "11010116"
    regest_path = (
        run_dir
        / f"raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/raw_regests/{regest_id}.json"
    )
    _write_json(
        regest_path,
        {
            "schema_version": 1,
            "source": "preflight_frozen_raw_regest_snapshot",
            "regest_id": regest_id,
            "header": f"Header text {regest_id}",
            "subentries": ["First subentry", "Second subentry"],
            "content_sha256": "ignored-by-export-fixture",
        },
    )
    snapshot_path = (
        run_dir
        / "raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/raw_regests_manifest.json"
    )
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    snapshot["records"][regest_id] = {
        "path": f"raw-academiccloud/intermediates-haiu_rag_ontologizer/provenance/raw_regests/{regest_id}.json",
        "sha256": _sha256(regest_path.read_text(encoding="utf-8")),
    }
    _write_json(snapshot_path, snapshot)

    run_manifest_path = run_dir / "environment/academiccloud-run-manifest.json"
    run_manifest = json.loads(run_manifest_path.read_text(encoding="utf-8"))
    run_manifest["regest_ids"].append(regest_id)
    run_manifest["raw_regest_snapshot"]["sha256"] = _sha256(
        snapshot_path.read_text(encoding="utf-8")
    )
    run_manifest["raw_regest_snapshot"]["count"] = len(
        run_manifest["regest_ids"]
    )
    _write_json(run_manifest_path, run_manifest)

    for condition in CONDITIONS:
        source_raw_path = (
            run_dir
            / f"raw-academiccloud/result-{condition}/{source_regest_id}.json"
        )
        raw = json.loads(source_raw_path.read_text(encoding="utf-8"))
        raw["regest_id"] = regest_id
        _write_json(
            run_dir / f"raw-academiccloud/result-{condition}/{regest_id}.json",
            raw,
        )
        for suffix in (".yaml", ".ttl"):
            source_path = run_dir / (
                f"raw-academiccloud/result-{condition}/"
                f"{source_regest_id}{suffix}"
            )
            _write(
                run_dir
                / f"raw-academiccloud/result-{condition}/{regest_id}{suffix}",
                source_path.read_text(encoding="utf-8"),
            )
        if condition in {"workflow_rag", "haiu_rag_ontologizer"}:
            for suffix in (".ttl", ".yaml"):
                source_path = (
                    run_dir
                    / f"raw-academiccloud/intermediates-{condition}/{source_regest_id}.retrieved{suffix}"
                )
                _write(
                    run_dir
                    / f"raw-academiccloud/intermediates-{condition}/{regest_id}.retrieved{suffix}",
                    source_path.read_text(encoding="utf-8"),
                )


def _write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def _write_json(path: Path, payload: dict[str, object]) -> None:
    _write(path, json.dumps(payload, indent=2))


def _sha256(content: str) -> str:
    return hashlib.sha256(content.encode("utf-8")).hexdigest()
